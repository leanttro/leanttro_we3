"""
ESTOQUE WPP — Backend unificado
Gerenciador de estoque, custos e vendas via WhatsApp (formulário ou IA/Groq)
"""
import os, json, csv, io, bcrypt, re, asyncio
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from typing import Optional

import psycopg2
import psycopg2.extras
from fastapi import FastAPI, HTTPException, Depends, Header, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
import jwt
import httpx
import openpyxl
import pandas as pd

# ─────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────
SECRET_KEY    = os.getenv("SECRET_KEY", "troca-isso-em-producao")
ADMIN_TOKEN   = os.getenv("ADMIN_TOKEN", "troca-isso-tambem")
GROQ_API_KEY_1 = os.getenv("GROQ_API_KEY_1", os.getenv("GROQ_API_KEY", ""))
GROQ_API_KEY_2 = os.getenv("GROQ_API_KEY_2", "")
BAILEYS_URL   = os.getenv("BAILEYS_URL", "http://baileys:3000")
ALGORITHM     = "HS256"
TOKEN_EXP     = 24 * 7  # horas
DB_URL        = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@postgres:5432/postgres")
GROQ_API_URL  = "https://api.groq.com/openai/v1/chat/completions"
TIMEZONE_PADRAO = ZoneInfo("America/Sao_Paulo")

GROQ_MODELOS_FALLBACK = [
    "openai/gpt-oss-20b",
    "openai/gpt-oss-120b",
    "qwen/qwen3.6-27b",
]

# Ordem lógica atualizada pelo cliente. Os números aqui precisam bater com o
# roteamento em processar_texto() lá embaixo.
MENU_TEXTO = (
    "📋 *Menu*\n"
    "1️⃣ Montar orçamento\n"
    "2️⃣ Registrar venda\n"
    "3️⃣ Cadastrar produto\n"
    "4️⃣ Cadastrar matéria-prima\n"
    "5️⃣ Resumo do dia\n"
    "6️⃣ Registrar entrada de estoque\n"
    "7️⃣ Visão geral do estoque\n"
    "8️⃣ Configurar resumo automático\n"
    "9️⃣ Editar Matéria-Prima\n"
    "🔟 Editar Estoque ou dados do Produto\n"
    "1️⃣1️⃣ Ajuda — o que cada opção faz\n"
    "1️⃣2️⃣ Calculadora de custos fixos\n"
    "1️⃣3️⃣ Agenda / Compromissos\n"
    "0️⃣ Abrir menu\n\n"
    "Responda com o número da opção."
)

TEXTO_AJUDA = (
    "ℹ️ *Como usar o sistema*\n\n"
    "1️⃣ *Montar orçamento* — monta um orçamento com um ou mais produtos, aplica desconto/aumento "
    "se você quiser e gera o texto pra enviar ao cliente.\n"
    "2️⃣ *Venda* — registra uma venda (diminui o estoque e, se o produto tiver receita, "
    "desconta as matérias-primas usadas automaticamente).\n"
    "3️⃣ *Cadastrar produto* — cria um novo produto (e opcionalmente já monta a receita dele).\n"
    "4️⃣ *Cadastrar matéria-prima* — cria um novo insumo usado nas receitas.\n"
    "5️⃣ *Resumo do dia* — total de entradas/vendas/ajustes de hoje.\n"
    "6️⃣ *Entrada de estoque* — registra chegada de mercadoria (aumenta o estoque).\n"
    "7️⃣ *Visão geral* — lista todos os produtos e matérias-primas com o estoque atual.\n"
    "8️⃣ *Resumo automático* — escolha até 2 horários por dia pra receber o resumo (opção 5) sem precisar pedir.\n"
    "9️⃣ *Editar Matéria-Prima* — mostra a receita atual do produto (se já tiver) e deixa você editar a quantidade "
    "de um item, adicionar mais matéria-prima, ou remover um item, tudo na mesma conversa.\n"
    "🔟 *Editar Estoque ou dados do Produto* — ao escolher essa opção, você escolhe entre:\n"
    "   1 - Editar estoque (quantidade) de um ou mais produtos de uma vez;\n"
    "   2 - Editar nome, custo unitário, preço de venda e/ou SKU de um produto específico "
    "(você escolhe quais campos mudar, informa os novos valores e confirma antes de salvar);\n"
    "   3 - Editar estoque (quantidade) de uma ou mais matérias-primas de uma vez.\n"
    "1️⃣1️⃣ *Ajuda* — este texto que você está lendo agora.\n"
    "1️⃣2️⃣ *Calculadora de custos fixos* — soma as contas fixas do negócio (aluguel, luz, internet etc.) "
    "e mostra o total, com a lista de cada conta. É diferente da calculadora que aparece durante o cadastro "
    "de produto (opção 3): aquela usa o custo fixo pra sugerir o preço de venda de UM produto específico "
    "(rateando o custo fixo pelo volume esperado e aplicando uma margem de lucro); esta aqui é só uma soma "
    "simples das contas, sem gerar preço de produto nenhum.\n"
    "1️⃣3️⃣ *Agenda / Compromissos* — crie compromissos (com ou sem vínculo a um cliente), veja os compromissos "
    "de hoje ou da semana, ou cancele um compromisso. Se você configurar um lembrete, recebe um aviso por "
    "WhatsApp X minutos antes do horário marcado.\n\n"
    "A qualquer momento, digite *menu* ou *0* para voltar aqui."
)

# ─────────────────────────────────────────
#  MÓDULOS POR CLIENTE (estoque / agenda)
# ─────────────────────────────────────────
# MENU_TEXTO acima é o menu completo (estoque + item 13 de agenda). Quando o
# cliente só tem o módulo "estoque" ativo, usamos a versão sem o item 13 pra
# não poluir o menu com uma opção que ele não tem.
MENU_TEXTO_SEM_AGENDA = MENU_TEXTO.replace("1️⃣3️⃣ Agenda / Compromissos\n", "")

AGENDA_MENU_TEXTO = (
    "📅 *Agenda*\n"
    "1 - Criar compromisso\n"
    "2 - Compromissos de hoje\n"
    "3 - Compromissos da semana\n"
    "4 - Cancelar um compromisso\n\n"
    "Responda com o número."
)

TEXTO_ESCOLHER_MODULO = (
    "👋 O que você quer fazer?\n"
    "1️⃣ 📦 Vendas & Estoque\n"
    "2️⃣ 📅 Agenda\n\n"
    "Responda com o número."
)

def obter_modulos_cliente(cliente: dict) -> list:
    """Módulos ativos do cliente (['estoque'], ['agenda'] ou ambos).
    Nullable no banco — sem valor definido, cai no default 'estoque' pra não
    quebrar clientes que já existiam antes desse conceito existir."""
    modulos = cliente.get("modulos")
    if not modulos:
        return ["estoque"]
    return list(modulos)

def etapa_raiz_para_modulos(modulos: list) -> str:
    """Se o cliente tem só 1 módulo ativo, vai direto pro menu daquele módulo
    (sem forçar escolha). Se tem os 2, primeiro escolhe qual módulo usar."""
    tem_estoque = "estoque" in modulos
    tem_agenda = "agenda" in modulos
    if tem_estoque and tem_agenda:
        return "escolher_modulo"
    if tem_agenda and not tem_estoque:
        return "agenda_menu"
    return "menu"

def texto_raiz_para_modulos(modulos: list) -> str:
    tem_estoque = "estoque" in modulos
    tem_agenda = "agenda" in modulos
    if tem_estoque and tem_agenda:
        return TEXTO_ESCOLHER_MODULO
    if tem_agenda and not tem_estoque:
        return AGENDA_MENU_TEXTO
    return MENU_TEXTO_SEM_AGENDA

# ─────────────────────────────────────────
#  BANCO
# ─────────────────────────────────────────
def get_conn_raw():
    return psycopg2.connect(DB_URL, cursor_factory=psycopg2.extras.RealDictCursor)

def get_db():
    conn = get_conn_raw()
    try:
        yield conn
    finally:
        conn.close()

def db_one(conn, sql, params=()):
    cur = conn.cursor()
    cur.execute(sql, params)
    row = cur.fetchone()
    return dict(row) if row else None

def db_all(conn, sql, params=()):
    cur = conn.cursor()
    cur.execute(sql, params)
    return [dict(r) for r in cur.fetchall()]

def db_exec(conn, sql, params=()):
    cur = conn.cursor()
    cur.execute(sql, params)
    conn.commit()
    try:
        row = cur.fetchone()
        return dict(row) if row else None
    except Exception:
        return None

# ─────────────────────────────────────────
#  AUTH — clientes (painel)
# ─────────────────────────────────────────
security = HTTPBearer()

def criar_token(cliente_id: int, email: str) -> str:
    payload = {
        "sub": str(cliente_id),
        "email": email,
        "exp": datetime.utcnow() + timedelta(hours=TOKEN_EXP)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def get_current_cliente(
    creds: HTTPAuthorizationCredentials = Depends(security),
    conn=Depends(get_db)
):
    try:
        payload = jwt.decode(creds.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        cliente_id = int(payload["sub"])
    except Exception:
        raise HTTPException(status_code=401, detail="Token inválido")
    cliente = db_one(conn, "SELECT * FROM clientes WHERE id = %s AND ativo = TRUE", (cliente_id,))
    if not cliente:
        raise HTTPException(status_code=401, detail="Conta não encontrada ou inativa")
    return cliente

def check_admin(authorization: str = Header(default="")):
    """
    Aceita tanto 'Authorization: Bearer <token>' (usado pelo login.html)
    quanto o token puro no header, para compatibilidade.
    """
    token = authorization.replace("Bearer ", "").strip() if authorization else ""
    if not token or token != ADMIN_TOKEN:
        raise HTTPException(status_code=403, detail="Token de admin inválido")
    return True

# ─────────────────────────────────────────
#  SCHEMAS
# ─────────────────────────────────────────
class LoginBody(BaseModel):
    email: str
    senha: str

class AdminCriarClienteBody(BaseModel):
    nome_negocio: str
    email: str
    senha: str
    plano: str = "formulario"  # formulario | ia
    modulos: Optional[list[str]] = None  # estoque | agenda — default ['estoque']

class AdminPlanoBody(BaseModel):
    plano: str

class AdminModulosBody(BaseModel):
    modulos: list[str]  # subconjunto de: estoque, agenda

class AdminAtivoBody(BaseModel):
    ativo: bool

class AtendimentoClienteFinalBody(BaseModel):
    ativado: bool

class AdminNumeroBody(BaseModel):
    numero: str
    nome: Optional[str] = None

class AdminTrocaNumeroBody(BaseModel):
    numero_novo: str
    motivo: Optional[str] = None

class ProdutoBody(BaseModel):
    nome: str
    sku: Optional[str] = None
    custo_unitario: Optional[float] = 0
    preco_venda: Optional[float] = 0
    estoque_atual: Optional[float] = 0
    unidade: Optional[str] = "un"

class MovimentacaoManualBody(BaseModel):
    produto_id: int
    tipo: str  # entrada | saida | venda | ajuste
    quantidade: float
    valor_unitario: Optional[float] = 0
    cliente_negocio_id: Optional[int] = None  # vínculo opcional com clientes_negocio (só relevante p/ venda)

class MateriaPrimaBody(BaseModel):
    nome: str
    sku: Optional[str] = None
    custo_unitario: Optional[float] = 0
    estoque_atual: Optional[float] = 0
    unidade: Optional[str] = "un"
    estoque_minimo: Optional[float] = None  # opcional — se None, não gera alerta

class MovimentacaoMateriaPrimaBody(BaseModel):
    materia_prima_id: int
    tipo: str  # entrada | saida | ajuste
    quantidade: float
    valor_unitario: Optional[float] = 0

class ReceitaItemBody(BaseModel):
    materia_prima_id: int
    quantidade_necessaria: float

class ReceitaBody(BaseModel):
    itens: list[ReceitaItemBody]

class AgendamentoBody(BaseModel):
    produto_id: int
    data_agendamento: str  # ISO format: YYYY-MM-DD
    hora: str  # HH:MM
    quantidade: float
    notas: Optional[str] = None

class ProdutoEditarBody(BaseModel):
    custo_unitario: Optional[float] = None
    preco_venda: Optional[float] = None

class OrcamentoItemBody(BaseModel):
    produto_id: int
    quantidade: float

class OrcamentoBody(BaseModel):
    itens: list[OrcamentoItemBody]
    desconto_tipo: Optional[str] = None  # "valor" | "percentual" | None
    desconto_valor: Optional[float] = 0
    observacoes: Optional[str] = None
    nome_cliente: Optional[str] = None
    cliente_negocio_id: Optional[int] = None  # vínculo opcional com clientes_negocio

class ClienteNegocioBody(BaseModel):
    nome: str
    telefone: Optional[str] = None

class CustoFixoBody(BaseModel):
    nome_da_conta: str
    valor: float

class AgendaCompromissoBody(BaseModel):
    titulo: str
    data: str  # ISO format: YYYY-MM-DD
    hora_inicio: str  # HH:MM
    hora_fim: Optional[str] = None       # HH:MM — se ausente, usa duracao_minutos (ou 30min padrão)
    duracao_minutos: Optional[int] = None
    cliente_negocio_id: Optional[int] = None  # vínculo opcional com clientes_negocio
    notas: Optional[str] = None
    lembrete_minutos_antes: Optional[int] = None  # se None, sem lembrete
    status: Optional[str] = "agendado"  # agendado | concluido | cancelado

class AgendaStatusBody(BaseModel):
    status: str  # agendado | concluido | cancelado

class CalculadoraCustoBody(BaseModel):
    produto_id: Optional[int] = None       # se informado e custo_variavel não vier, usa o custo_unitario do produto
    custo_variavel: Optional[float] = 0    # custo variável por unidade, informado manualmente
    volume_esperado: float                 # volume de vendas esperado por mês
    margem: Optional[float] = None         # margem de lucro desejada em % (padrão 30%)

# ─────────────────────────────────────────
#  HELPERS DE NEGÓCIO
# ─────────────────────────────────────────
def formatar_moeda(valor: float) -> str:
    """Formata número no padrão R$ 1.234,56"""
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_qtd(qtd: float) -> str:
    """Mostra quantidade sem casas decimais quando é um número inteiro"""
    return str(int(qtd)) if float(qtd).is_integer() else str(qtd)

def normalizar_numero(remote_jid_ou_numero: str) -> str:
    n = remote_jid_ou_numero.split("@")[0]
    return re.sub(r"\D", "", n)

def buscar_numero_autorizado(conn, numero: str):
    return db_one(conn, "SELECT * FROM numeros_autorizados WHERE numero = %s AND ativo = TRUE", (numero,))

def get_or_create_sessao(conn, numero_autorizado_id: int):
    sessao = db_one(conn, "SELECT * FROM sessoes_conversa WHERE numero_autorizado_id = %s", (numero_autorizado_id,))
    if sessao:
        return sessao
    return db_exec(conn, """
        INSERT INTO sessoes_conversa (numero_autorizado_id, etapa_atual, dados_parciais)
        VALUES (%s, 'menu', '{}') RETURNING *
    """, (numero_autorizado_id,))

def salvar_sessao(conn, numero_autorizado_id: int, etapa: str, dados: dict):
    db_exec(conn, """
        UPDATE sessoes_conversa
        SET etapa_atual = %s, dados_parciais = %s, atualizado_em = NOW()
        WHERE numero_autorizado_id = %s
    """, (etapa, json.dumps(dados), numero_autorizado_id))

# ── Sessão de CLIENTE FINAL (visitante público, não é funcionário) ──
def get_or_create_sessao_cliente_final(conn, cliente_id: int, numero: str):
    sessao = db_one(conn, """
        SELECT * FROM sessoes_cliente_final WHERE cliente_id = %s AND numero = %s
    """, (cliente_id, numero))
    if sessao:
        return sessao
    return db_exec(conn, """
        INSERT INTO sessoes_cliente_final (cliente_id, numero, etapa_atual, dados_parciais)
        VALUES (%s, %s, 'menu_cliente_final', '{}') RETURNING *
    """, (cliente_id, numero))

def salvar_sessao_cliente_final(conn, cliente_id: int, numero: str, etapa: str, dados: dict):
    db_exec(conn, """
        UPDATE sessoes_cliente_final
        SET etapa_atual = %s, dados_parciais = %s, atualizado_em = NOW()
        WHERE cliente_id = %s AND numero = %s
    """, (etapa, json.dumps(dados), cliente_id, numero))

def buscar_clientes_com_atendimento_ativado(conn):
    """Todos os clientes (empresas) que ligaram o toggle de atendimento
    automático ao cliente final."""
    return db_all(conn, """
        SELECT * FROM clientes WHERE ativo = TRUE AND atendimento_cliente_final_ativado = TRUE
    """)

def buscar_produto_por_nome(conn, cliente_id: int, nome: str):
    return db_one(conn, """
        SELECT * FROM produtos
        WHERE cliente_id = %s AND ativo = TRUE AND nome ILIKE %s
        ORDER BY id LIMIT 1
    """, (cliente_id, f"%{nome.strip()}%"))

def buscar_materia_prima_por_nome(conn, cliente_id: int, nome: str):
    return db_one(conn, """
        SELECT * FROM materias_primas
        WHERE cliente_id = %s AND ativo = TRUE AND nome ILIKE %s
        ORDER BY id LIMIT 1
    """, (cliente_id, f"%{nome.strip()}%"))

def fmt_num(valor) -> str:
    """Formata um número (Decimal ou float) sem casas decimais falsas
    (ex: evita mostrar '2.000' quando é só o número 2, o que em pt-BR
    é lido como 'dois mil'). Usa vírgula como separador decimal."""
    v = float(valor or 0)
    if v == int(v):
        texto = str(int(v))
    else:
        texto = f"{v:.3f}".rstrip("0").rstrip(".")
    return texto.replace(".", ",")

def listar_produtos_cliente(conn, cliente_id: int):
    return db_all(conn, """
        SELECT id, nome, preco_venda FROM produtos
        WHERE cliente_id = %s AND ativo = TRUE
        ORDER BY nome
    """, (cliente_id,))

def listar_materias_primas_cliente(conn, cliente_id: int):
    return db_all(conn, """
        SELECT id, nome, unidade FROM materias_primas
        WHERE cliente_id = %s AND ativo = TRUE
        ORDER BY nome
    """, (cliente_id,))

def montar_lista_numerada(itens, titulo: str, rodape: str = "Responda com o número.", mostrar_preco: bool = False) -> str:
    linhas = [titulo, ""]
    for i, item in enumerate(itens, start=1):
        if mostrar_preco:
            preco = float(item.get("preco_venda") or 0)
            linhas.append(f"{i}. {item['nome']} — {formatar_moeda(preco)}")
        else:
            linhas.append(f"{i}. {item['nome']}")
    linhas.append("")
    linhas.append(rodape)
    return "\n".join(linhas)

def montar_menu_edicao_receita(dados: dict) -> str:
    """Renderiza o estado atual da receita sendo editada (opção 9 — Editar
    Matéria-Prima) e as ações disponíveis: editar quantidade de um item
    existente, adicionar matéria-prima nova, remover um item, ou salvar."""
    itens = dados.get("receita_itens", [])
    nome_produto = dados.get("receita_produto_nome", "")
    if itens:
        linhas = "\n".join(
            f"{i}. {fmt_num(item['quantidade'])} {item['unidade']} de {item['nome']}"
            for i, item in enumerate(itens, start=1)
        )
        corpo = f"📋 Receita atual de *{nome_produto}*:\n{linhas}"
    else:
        corpo = f"*{nome_produto}* ainda não tem nenhuma matéria-prima na receita."
    rodape = (
        "\n\nDigite o número de um item para editar a quantidade, "
        "*novo* para adicionar uma matéria-prima, "
        "*remover N* para tirar um item (ex: remover 2), "
        "ou *pronto* para salvar."
    )
    return corpo + rodape

def parse_selecao_multipla(texto: str, max_idx: int):
    """Aceita '1' ou '1,3,5' ou '1 3 5' e devolve a lista de índices (1-based)
    válidos, sem repetir, na ordem digitada. Retorna None se algo for inválido."""
    partes = re.split(r"[,\s]+", texto.strip())
    indices = []
    for p in partes:
        if not p:
            continue
        if not p.isdigit():
            return None
        idx = int(p)
        if idx < 1 or idx > max_idx:
            return None
        if idx not in indices:
            indices.append(idx)
    return indices or None

def montar_texto_confirmacao_carrinho(dados: dict, tipo: str) -> str:
    """Monta o texto de confirmação (SIM/NÃO) do carrinho inteiro."""
    carrinho = dados.get("carrinho", [])
    linhas = []
    total_geral = 0.0
    for item in carrinho:
        if tipo == "ajuste":
            linhas.append(f"- {item['produto_nome']} → {fmt_num(item['quantidade'])}")
        else:
            item_total = round(item["quantidade"] * item["valor_unitario"], 2)
            total_geral += item_total
            linhas.append(
                f"- {fmt_num(item['quantidade'])} × {item['produto_nome']} "
                f"a R$ {item['valor_unitario']:.2f} (R$ {item_total:.2f})"
            )
    corpo = "\n".join(linhas)
    acao = {"entrada": "a entrada", "venda": "a venda", "saida": "a saída", "ajuste": "os ajustes"}[tipo]
    if tipo == "ajuste":
        return f"Confirma {acao}?\n{corpo}\n\nResponda SIM ou NÃO."
    return f"Confirma {acao}?\n{corpo}\n\nTotal: R$ {total_geral:.2f}\n\nResponda SIM ou NÃO."

def avancar_fila_ou_confirmar(conn, numero_autorizado_id: int, dados: dict, tipo: str) -> str:
    """Chamado depois que um item do carrinho (produto + quantidade [+ valor])
    foi completado. Se ainda sobra produto na fila, pergunta a quantidade do
    próximo; se não sobra mais nada, monta o resumo do carrinho inteiro pra
    confirmação (SIM/NÃO) — exceto em vendas, onde antes passa pelo passo
    opcional de identificar o cliente (TAREFA 2)."""
    fila = dados.get("fila_produtos_ids", [])
    if fila:
        proximo_id = fila.pop(0)
        produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s", (proximo_id,))
        dados["fila_produtos_ids"] = fila
        dados["produto_id"] = produto["id"]
        dados["produto_nome"] = produto["nome"]
        salvar_sessao(conn, numero_autorizado_id, f"{tipo}_quantidade", dados)
        return f"Quantidade de *{produto['nome']}*?"

    if tipo == "venda":
        salvar_sessao(conn, numero_autorizado_id, "venda_cliente", dados)
        return "Quem é o cliente? Pode digitar o nome, telefone, ou 'pular'."

    salvar_sessao(conn, numero_autorizado_id, "confirmando", dados)
    return montar_texto_confirmacao_carrinho(dados, tipo)

# ─────────────────────────────────────────
#  ORÇAMENTO — helpers compartilhados (painel HTTP + bot WhatsApp)
# ─────────────────────────────────────────
def calcular_itens_orcamento(conn, cliente_id: int, itens: list) -> tuple:
    """itens: lista de dicts {'produto_id': int, 'quantidade': float}.
    Retorna (itens_detalhados, subtotal). Lança ValueError se algum produto não existir."""
    itens_detalhados = []
    subtotal = 0.0
    for item in itens:
        produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s",
                          (item["produto_id"], cliente_id))
        if not produto:
            raise ValueError(f"Produto {item['produto_id']} não encontrado")
        preco_unitario = float(produto["preco_venda"] or 0)
        subtotal_item = preco_unitario * item["quantidade"]
        subtotal += subtotal_item
        itens_detalhados.append({
            "produto_id": produto["id"],
            "nome": produto["nome"],
            "quantidade": item["quantidade"],
            "preco_unitario": preco_unitario,
            "subtotal": subtotal_item,
        })
    return itens_detalhados, subtotal

def calcular_ajuste_preco(subtotal: float, ajuste_tipo: Optional[str], ajuste_valor_informado: float) -> float:
    """Retorna o valor de ajuste a somar ao subtotal.
    Negativo = desconto, positivo = aumento, zero = manteve igual.
    ajuste_tipo: 'desconto_percentual' | 'desconto_valor' | 'aumento_percentual' | 'aumento_valor' | None"""
    if ajuste_tipo == "desconto_percentual":
        ajuste = -(subtotal * (ajuste_valor_informado / 100))
    elif ajuste_tipo == "desconto_valor":
        ajuste = -ajuste_valor_informado
    elif ajuste_tipo == "aumento_percentual":
        ajuste = subtotal * (ajuste_valor_informado / 100)
    elif ajuste_tipo == "aumento_valor":
        ajuste = ajuste_valor_informado
    else:
        ajuste = 0.0
    if subtotal + ajuste < 0:  # nunca deixa o total ficar negativo
        ajuste = -subtotal
    return ajuste

def montar_texto_orcamento(cliente: dict, itens_detalhados: list, subtotal: float, ajuste_tipo: Optional[str],
                            ajuste_calculado: float, ajuste_valor_informado: float, total: float,
                            observacoes: Optional[str] = None, nome_cliente: Optional[str] = None) -> str:
    linhas = ["🧾 *Orçamento*"]
    if nome_cliente:
        linhas.append(f"Para: {nome_cliente}")
    linhas.append("")
    for item in itens_detalhados:
        linhas.append(f"▪️ {formatar_qtd(item['quantidade'])}x {item['nome']} — {formatar_moeda(item['subtotal'])}")
    linhas.append("")
    linhas.append(f"Subtotal: {formatar_moeda(subtotal)}")
    if ajuste_calculado < 0:
        if ajuste_tipo == "desconto_percentual":
            linhas.append(f"Desconto ({formatar_qtd(ajuste_valor_informado)}%): -{formatar_moeda(abs(ajuste_calculado))}")
        else:
            linhas.append(f"Desconto: -{formatar_moeda(abs(ajuste_calculado))}")
    elif ajuste_calculado > 0:
        if ajuste_tipo == "aumento_percentual":
            linhas.append(f"Acréscimo ({formatar_qtd(ajuste_valor_informado)}%): +{formatar_moeda(ajuste_calculado)}")
        else:
            linhas.append(f"Acréscimo: +{formatar_moeda(ajuste_calculado)}")
    linhas.append(f"*Total: {formatar_moeda(total)}*")
    if observacoes:
        linhas.append("")
        linhas.append(observacoes)
    return "\n".join(linhas)

def salvar_orcamento_db(conn, cliente_id: int, nome_cliente, itens_detalhados: list, subtotal: float,
                         ajuste_tipo, ajuste_calculado: float, total: float, texto_formatado: str,
                         observacoes, cliente_negocio_id: Optional[int] = None) -> dict:
    return db_exec(conn, """
        INSERT INTO orcamentos (cliente_id, nome_cliente, itens, subtotal, desconto_tipo, desconto_valor, total, texto_formatado, observacoes, cliente_negocio_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING *
    """, (
        cliente_id, nome_cliente, json.dumps(itens_detalhados), subtotal,
        ajuste_tipo, ajuste_calculado, total, texto_formatado, observacoes, cliente_negocio_id
    ))

def parse_ajuste_preco_texto(texto: str, subtotal: float):
    """Interpreta a resposta do usuário no passo de ajuste de preço do bot.
    Aceita 'não'/'nao'/'n'/'0'/'manter'/'igual' (mantém o preço igual);
    '-10' ou '-10%' (desconto em R$ ou %); '+10' ou '+10%' (aumento em R$ ou %);
    e, por compatibilidade, '10' ou '10%' sem sinal também é tratado como desconto.
    Retorna (tipo, valor_informado) — tipo é None quando não há ajuste — ou None se não entendeu."""
    t = texto.strip().lower()
    if t in ("não", "nao", "n", "0", "manter", "igual", "mesma coisa", "sem ajuste", "nenhum", "pular"):
        return (None, 0.0)
    t_limpo = t.replace("r$", "").strip()

    sinal = None
    if t_limpo.startswith("+"):
        sinal = "aumento"
        t_limpo = t_limpo[1:].strip()
    elif t_limpo.startswith("-"):
        sinal = "desconto"
        t_limpo = t_limpo[1:].strip()
    else:
        sinal = "desconto"  # sem sinal explícito: mantém o comportamento antigo (desconto)

    if t_limpo.endswith("%"):
        try:
            valor = float(t_limpo[:-1].replace(",", "."))
            return (f"{sinal}_percentual", valor)
        except ValueError:
            return None
    try:
        valor = float(t_limpo.replace(",", "."))
        return (f"{sinal}_valor", valor)
    except ValueError:
        return None

def aplicar_movimentacao(conn, cliente_id, produto_id, numero_autorizado_id, tipo, quantidade, valor_unitario, origem, mensagem_original=None, cliente_negocio_id=None):
    quantidade = float(quantidade)
    valor_unitario = float(valor_unitario or 0)
    valor_total = round(quantidade * valor_unitario, 2)

    mov = db_exec(conn, """
        INSERT INTO movimentacoes
            (cliente_id, produto_id, numero_autorizado_id, tipo, quantidade, valor_unitario, valor_total, origem, mensagem_original, cliente_negocio_id)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (cliente_id, produto_id, numero_autorizado_id, tipo, quantidade, valor_unitario, valor_total, origem, mensagem_original, cliente_negocio_id))
    movimentacao_id = mov["id"] if mov else None

    alertas = []
    if tipo == "entrada":
        db_exec(conn, "UPDATE produtos SET estoque_atual = estoque_atual + %s, custo_unitario = %s WHERE id = %s",
                (quantidade, valor_unitario, produto_id))
    elif tipo in ("saida", "venda"):
        db_exec(conn, "UPDATE produtos SET estoque_atual = estoque_atual - %s WHERE id = %s", (quantidade, produto_id))
        if tipo == "venda" and valor_unitario:
            db_exec(conn, "UPDATE produtos SET preco_venda = %s WHERE id = %s", (valor_unitario, produto_id))
        if tipo == "venda":
            # Toda venda de um produto que tem receita cadastrada desconta
            # automaticamente a matéria-prima usada (ficha técnica / BOM).
            # Passa o movimentacao_id pra poder reverter isso com precisão depois (TAREFA 2 — dashboard).
            alertas = baixar_materia_prima_por_receita(conn, cliente_id, produto_id, numero_autorizado_id, quantidade, origem,
                                                        movimentacao_id=movimentacao_id)
    elif tipo == "ajuste":
        db_exec(conn, "UPDATE produtos SET estoque_atual = %s WHERE id = %s", (quantidade, produto_id))

    return valor_total, alertas

# ─────────────────────────────────────────
#  MATÉRIA-PRIMA / RECEITA (ficha técnica)
# ─────────────────────────────────────────
def aplicar_movimentacao_materia_prima(conn, cliente_id, materia_prima_id, numero_autorizado_id, tipo,
                                        quantidade, valor_unitario, origem, mensagem_original=None,
                                        produto_id_origem=None, movimentacao_id=None):
    quantidade = float(quantidade)
    valor_unitario = float(valor_unitario or 0)
    valor_total = round(quantidade * valor_unitario, 2)

    db_exec(conn, """
        INSERT INTO movimentacoes_materia_prima
            (cliente_id, materia_prima_id, numero_autorizado_id, tipo, quantidade, valor_unitario,
             valor_total, origem, mensagem_original, produto_id_origem, movimentacao_id)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (cliente_id, materia_prima_id, numero_autorizado_id, tipo, quantidade, valor_unitario,
          valor_total, origem, mensagem_original, produto_id_origem, movimentacao_id))

    if tipo == "entrada":
        db_exec(conn, "UPDATE materias_primas SET estoque_atual = estoque_atual + %s, custo_unitario = %s WHERE id = %s",
                (quantidade, valor_unitario, materia_prima_id))
    elif tipo in ("saida", "baixa_receita"):
        db_exec(conn, "UPDATE materias_primas SET estoque_atual = estoque_atual - %s WHERE id = %s",
                (quantidade, materia_prima_id))
    elif tipo == "ajuste":
        db_exec(conn, "UPDATE materias_primas SET estoque_atual = %s WHERE id = %s", (quantidade, materia_prima_id))

    # Alerta de estoque baixo — só dispara pra quem definiu um estoque mínimo
    # (campo opcional; quem não usa a função simplesmente nunca recebe isso).
    alerta = None
    if tipo in ("saida", "baixa_receita"):
        mp = db_one(conn, "SELECT nome, unidade, estoque_atual, estoque_minimo FROM materias_primas WHERE id = %s",
                    (materia_prima_id,))
        if mp and mp["estoque_minimo"] is not None and float(mp["estoque_atual"]) < float(mp["estoque_minimo"]):
            alerta = (f"⚠️ *{mp['nome']}* está com {fmt_num(mp['estoque_atual'])} {mp['unidade']}, "
                      f"abaixo do mínimo de {fmt_num(mp['estoque_minimo'])} {mp['unidade']}.")

    return valor_total, alerta

def baixar_materia_prima_por_receita(conn, cliente_id, produto_id, numero_autorizado_id, quantidade_vendida, origem, movimentacao_id=None):
    """Ao vender 1 ou mais unidades de um produto, desconta a matéria-prima
    de cada item da receita cadastrada, proporcionalmente à quantidade vendida.
    Produtos sem receita cadastrada simplesmente não têm nenhum item aqui.
    `movimentacao_id`, quando informado, fica gravado em cada baixa pra permitir
    reverter tudo com precisão depois (TAREFA 2 — excluir venda no dashboard).
    Retorna a lista de alertas de estoque baixo disparados (pode ser vazia)."""
    itens = db_all(conn, """
        SELECT r.materia_prima_id, r.quantidade_necessaria, m.custo_unitario
        FROM receita_itens r
        JOIN materias_primas m ON m.id = r.materia_prima_id
        WHERE r.produto_id = %s AND m.ativo = TRUE
    """, (produto_id,))
    alertas = []
    for item in itens:
        qtd_baixa = round(float(item["quantidade_necessaria"]) * float(quantidade_vendida), 4)
        _, alerta = aplicar_movimentacao_materia_prima(
            conn, cliente_id, item["materia_prima_id"], numero_autorizado_id,
            "baixa_receita", qtd_baixa, item["custo_unitario"],
            origem=origem, produto_id_origem=produto_id, movimentacao_id=movimentacao_id
        )
        if alerta:
            alertas.append(alerta)
    return alertas

# ─────────────────────────────────────────
#  CLIENTES DO NEGÓCIO (clientes finais — TAREFA 2)
# ─────────────────────────────────────────
def buscar_ou_criar_cliente_negocio(conn, cliente_id: int, texto: str):
    """Recebe o texto digitado no passo opcional 'Quem é o cliente?' (nome OU
    telefone) e retorna o registro em clientes_negocio, criando se necessário.
    Se parecer um telefone e já existir um cliente com esse telefone pra esse
    cliente_id, reaproveita o registro existente. Retorna None se o texto for
    vazio ou 'pular' — nesse caso o chamador simplesmente segue sem vincular."""
    if not texto:
        return None
    texto = texto.strip()
    if not texto or texto.lower() == "pular":
        return None

    digitos = re.sub(r"\D", "", texto)
    if len(digitos) >= 8:
        existente = db_one(conn, "SELECT * FROM clientes_negocio WHERE cliente_id = %s AND telefone = %s",
                            (cliente_id, digitos))
        if existente:
            return existente
        return db_exec(conn, """
            INSERT INTO clientes_negocio (cliente_id, nome, telefone) VALUES (%s,%s,%s) RETURNING *
        """, (cliente_id, texto, digitos))

    existente = db_one(conn, """
        SELECT * FROM clientes_negocio WHERE cliente_id = %s AND nome ILIKE %s ORDER BY id LIMIT 1
    """, (cliente_id, texto))
    if existente:
        return existente
    return db_exec(conn, """
        INSERT INTO clientes_negocio (cliente_id, nome, telefone) VALUES (%s,%s,NULL) RETURNING *
    """, (cliente_id, texto))

# ─────────────────────────────────────────
#  AGENDA / COMPROMISSOS
# ─────────────────────────────────────────
def parse_data_br(texto: str):
    """Aceita 'DD/MM' ou 'DD/MM/AAAA' (ano opcional, assume o ano corrente).
    Retorna um date ou None se inválido."""
    texto = texto.strip()
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?", texto)
    if not m:
        return None
    dia, mes, ano = int(m.group(1)), int(m.group(2)), m.group(3)
    hoje = datetime.now(TIMEZONE_PADRAO).date()
    if ano:
        ano = int(ano)
        if ano < 100:
            ano += 2000
    else:
        ano = hoje.year
    try:
        return datetime(ano, mes, dia).date()
    except ValueError:
        return None

def calcular_hora_fim(hora_inicio_str: str, hora_fim_str: str = None, duracao_minutos: int = None) -> str:
    """Recebe 'HH:MM'. Se hora_fim_str vier preenchida, usa ela direto. Senão,
    soma duracao_minutos (ou 30min padrão) à hora_inicio e devolve 'HH:MM'."""
    if hora_fim_str:
        return hora_fim_str[:5]
    h, m = map(int, hora_inicio_str[:5].split(":"))
    dur = int(duracao_minutos) if duracao_minutos else 30
    fim_min = (h * 60 + m + dur) % (24 * 60)
    return f"{fim_min // 60:02d}:{fim_min % 60:02d}"

def _horarios_se_sobrepoem(ini1: str, fim1: str, ini2: str, fim2: str) -> bool:
    def to_min(s):
        h, m = map(int, str(s)[:5].split(":"))
        return h * 60 + m
    return to_min(ini1) < to_min(fim2) and to_min(ini2) < to_min(fim1)

def listar_compromissos_dia(conn, cliente_id: int, data):
    return db_all(conn, """
        SELECT a.*, cn.nome AS cliente_negocio_nome
        FROM agenda_compromissos a
        LEFT JOIN clientes_negocio cn ON cn.id = a.cliente_negocio_id
        WHERE a.cliente_id = %s AND a.data = %s AND a.status != 'cancelado'
        ORDER BY a.hora_inicio
    """, (cliente_id, data))

def verificar_conflito_agenda(conn, cliente_id: int, data, hora_inicio: str, hora_fim: str, excluir_id: int = None):
    """Checa (sem bloquear) se já existe outro compromisso não-cancelado do mesmo
    cliente, no mesmo dia, cujo horário se sobrepõe. Retorna a lista de conflitos."""
    existentes = listar_compromissos_dia(conn, cliente_id, data)
    conflitos = []
    for c in existentes:
        if excluir_id and c["id"] == excluir_id:
            continue
        fim_existente = calcular_hora_fim(str(c["hora_inicio"])[:5], str(c["hora_fim"])[:5] if c["hora_fim"] else None)
        if _horarios_se_sobrepoem(hora_inicio, hora_fim, str(c["hora_inicio"])[:5], fim_existente):
            conflitos.append(c)
    return conflitos

def listar_clientes_negocio_cliente(conn, cliente_id: int):
    return db_all(conn, "SELECT id, nome FROM clientes_negocio WHERE cliente_id = %s ORDER BY nome", (cliente_id,))

def listar_compromissos_futuros_cliente(conn, cliente_id: int):
    hoje = datetime.now(TIMEZONE_PADRAO).date()
    return db_all(conn, """
        SELECT * FROM agenda_compromissos
        WHERE cliente_id = %s AND status = 'agendado' AND data >= %s
        ORDER BY data, hora_inicio
    """, (cliente_id, hoje))

def texto_lista_compromissos_periodo(conn, cliente_id: int, dias: int) -> str:
    """dias=0 -> só hoje; dias=7 -> hoje até +7 dias (semana)."""
    hoje = datetime.now(TIMEZONE_PADRAO).date()
    fim = hoje + timedelta(days=dias)
    compromissos = db_all(conn, """
        SELECT a.*, cn.nome AS cliente_negocio_nome
        FROM agenda_compromissos a
        LEFT JOIN clientes_negocio cn ON cn.id = a.cliente_negocio_id
        WHERE a.cliente_id = %s AND a.status = 'agendado' AND a.data BETWEEN %s AND %s
        ORDER BY a.data, a.hora_inicio
    """, (cliente_id, hoje, fim))
    titulo = "📅 *Compromissos de hoje*" if dias == 0 else "📅 *Compromissos da semana*"
    if not compromissos:
        return f"{titulo}\n\nNenhum compromisso."
    linhas = [titulo, ""]
    for c in compromissos:
        linha = f"- {c['data'].strftime('%d/%m')} {str(c['hora_inicio'])[:5]} — {c['titulo']}"
        if c.get("cliente_negocio_nome"):
            linha += f" ({c['cliente_negocio_nome']})"
        linhas.append(linha)
    return "\n".join(linhas)

def montar_resumo_agenda_confirmacao(conn, cliente_id: int, dados: dict) -> str:
    data_obj = datetime.fromisoformat(dados["agenda_data"]).date()
    hora_inicio = dados["agenda_hora"]
    hora_fim = calcular_hora_fim(hora_inicio)
    linhas = [
        "Confirma o compromisso?", "",
        f"📌 {dados['agenda_titulo']}",
        f"📅 {data_obj.strftime('%d/%m/%Y')} às {hora_inicio}",
    ]
    if dados.get("agenda_cliente_negocio_nome"):
        linhas.append(f"👤 {dados['agenda_cliente_negocio_nome']}")
    if dados.get("agenda_lembrete_minutos"):
        linhas.append(f"⏰ Lembrete {dados['agenda_lembrete_minutos']} min antes")
    conflitos = verificar_conflito_agenda(conn, cliente_id, data_obj, hora_inicio, hora_fim)
    if conflitos:
        linhas.append("")
        linhas.append("⚠️ Já existe outro compromisso nesse horário:")
        for c in conflitos:
            linhas.append(f"- {c['titulo']} ({str(c['hora_inicio'])[:5]})")
    linhas.append("")
    linhas.append("Responda SIM ou NÃO.")
    return "\n".join(linhas)

# ─────────────────────────────────────────
#  CALCULADORA DE CUSTO/PREÇO (TAREFA 1)
# ─────────────────────────────────────────
def calcular_resultado_calculadora(dados: dict) -> dict:
    """Aplica as fórmulas da calculadora de preço em cima do que já foi
    coletado em `dados` (via formulário passo-a-passo ou já extraído pela IA).
    Não bate no banco — é só matemática em cima do dicionário de sessão."""
    custo_variavel = float(dados.get("custo_variavel") or 0)
    if not custo_variavel and dados.get("calc_receita_itens"):
        custo_variavel = sum(
            float(item["quantidade"]) * float(item.get("custo_unitario") or 0)
            for item in dados["calc_receita_itens"]
        )

    custo_fixo_mensal = float(dados.get("calc_custo_fixo_mensal") or 0)
    volume_esperado = float(dados.get("calc_volume_esperado") or 0)
    margem = dados.get("calc_margem")
    margem = float(margem) if margem is not None else 30.0
    margem_ajustada = margem >= 100
    if margem_ajustada:
        margem = 99.0  # evita divisão por zero/negativa — margem de 100%+ não faz sentido matemático aqui

    custo_fixo_rateado = (custo_fixo_mensal / volume_esperado) if volume_esperado > 0 else 0.0
    custo_total_unitario = custo_variavel + custo_fixo_rateado
    preco_sugerido = custo_total_unitario / (1 - margem / 100)

    return {
        "custo_variavel": round(custo_variavel, 2),
        "custo_fixo_rateado": round(custo_fixo_rateado, 2),
        "custo_total_unitario": round(custo_total_unitario, 2),
        "preco_sugerido": round(preco_sugerido, 2),
        "margem": margem,
        "margem_ajustada": margem_ajustada,
    }

def texto_resultado_calculadora(dados: dict, resultado: dict) -> str:
    nome = dados.get("prod_nome", "produto")
    aviso_margem = (
        f"⚠️ A margem informada era 100% ou mais, o que não é matematicamente possível "
        f"(o preço nunca cobriria o custo) — ajustei automaticamente para {fmt_num(resultado['margem'])}%.\n\n"
        if resultado.get("margem_ajustada") else ""
    )
    return (
        f"🧮 *Cálculo de preço — {nome}*\n"
        f"Custo variável (por unidade): R$ {resultado['custo_variavel']:.2f}\n"
        f"Custo fixo rateado (por unidade): R$ {resultado['custo_fixo_rateado']:.2f}\n"
        f"Custo total por unidade: R$ {resultado['custo_total_unitario']:.2f}\n"
        f"Margem desejada: {fmt_num(resultado['margem'])}%\n"
        f"*Preço sugerido: R$ {resultado['preco_sugerido']:.2f}*\n\n"
        f"{aviso_margem}"
        "⚠️ O custo fixo mensal e o volume esperado não ficam salvos — é só pra calcular o preço agora "
        "(a cada novo produto, esses valores são pedidos de novo).\n\n"
        "Confirma usar este custo e preço no cadastro? Responda SIM ou NÃO."
    )

# ─────────────────────────────────────────
#  EDIÇÃO DE CAMPOS DO PRODUTO (nome/custo/preço/SKU) — TAREFA 1 (opção 10.2)
# ─────────────────────────────────────────
# Cada item é (nome_da_coluna_no_banco, rótulo mostrado pro usuário). A ordem
# aqui define a numeração do submenu (1, 2, 3, 4).
CAMPOS_EDITAVEIS_PRODUTO = [
    ("nome", "Nome"),
    ("custo_unitario", "Custo unitário"),
    ("preco_venda", "Preço de venda"),
    ("sku", "SKU"),
]
CAMPO_LABEL_PRODUTO = dict(CAMPOS_EDITAVEIS_PRODUTO)
CAMPOS_NUMERICOS_PRODUTO = {"custo_unitario", "preco_venda"}

def texto_menu_campos_editar(nome_produto: str) -> str:
    linhas = [f"Quais campos você quer editar em *{nome_produto}*?", ""]
    for i, (_, label) in enumerate(CAMPOS_EDITAVEIS_PRODUTO, start=1):
        linhas.append(f"{i} - {label}")
    linhas.append("")
    linhas.append("Responda com o(s) número(s). Pra mais de um, separe por vírgula (ex: 2,3).")
    return "\n".join(linhas)

def parse_valor_campo_produto(campo: str, texto: str):
    """Converte o texto digitado pro tipo certo do campo. Retorna (ok, valor)."""
    texto = texto.strip()
    if not texto:
        return False, None
    if campo in CAMPOS_NUMERICOS_PRODUTO:
        try:
            return True, float(texto.replace(",", "."))
        except ValueError:
            return False, None
    return True, texto  # nome / sku: texto livre

def montar_resumo_edicao_produto(dados: dict) -> str:
    """Monta o resumo 'de → para' de cada campo escolhido, pra confirmação SIM/NÃO."""
    atual = dados.get("editar_produto_atual", {})
    novos = dados.get("editar_campos_novos", {})
    linhas = [f"Confirma as alterações em *{dados.get('editar_produto_nome')}*?", ""]
    for campo, valor_novo in novos.items():
        valor_atual = atual.get(campo)
        if campo in CAMPOS_NUMERICOS_PRODUTO:
            atual_fmt = f"R$ {float(valor_atual or 0):.2f}"
            novo_fmt = f"R$ {float(valor_novo):.2f}"
        else:
            atual_fmt = valor_atual if valor_atual else "(vazio)"
            novo_fmt = valor_novo
        linhas.append(f"- {CAMPO_LABEL_PRODUTO[campo]}: {atual_fmt} → {novo_fmt}")
    linhas.append("")
    linhas.append("Responda SIM ou NÃO.")
    return "\n".join(linhas)

def aplicar_edicao_produto(conn, cliente_id: int, produto_id: int, campos_novos: dict):
    """Grava no banco só os campos escolhidos pelo usuário — os demais campos
    do produto ficam intactos (TAREFA 1). `campos_novos` só pode conter chaves
    vindas de CAMPO_LABEL_PRODUTO (nunca texto livre do usuário), então é
    seguro usá-las como nome de coluna no SQL."""
    campos_novos = {k: v for k, v in (campos_novos or {}).items() if k in CAMPO_LABEL_PRODUTO}
    if not campos_novos:
        return
    sets = [f"{campo} = %s" for campo in campos_novos]
    params = list(campos_novos.values()) + [produto_id, cliente_id]
    sql = f"UPDATE produtos SET {', '.join(sets)} WHERE id = %s AND cliente_id = %s"
    db_exec(conn, sql, tuple(params))

# ─────────────────────────────────────────
#  CALCULADORA DE CUSTOS FIXOS (somatória simples) — TAREFA 3 (opção 12)
# ─────────────────────────────────────────
# Diferente da calculadora de custo/preço acima (que roda dentro do cadastro
# de produto e sugere um preço de venda pra UM produto, rateando custo fixo
# por volume + margem): esta aqui só soma as contas fixas do negócio.
def parse_conta_fixa_texto(texto: str):
    """Aceita 'Aluguel 2000' ou 'Aluguel R$ 2.000,00' (nome + valor no final da
    mensagem). Retorna (nome, valor) ou None se não conseguir separar os dois."""
    texto = texto.strip().replace("R$", "").replace("r$", "").strip()
    if not texto:
        return None
    partes = texto.rsplit(" ", 1)
    if len(partes) != 2:
        return None
    nome, valor_str = partes
    nome = nome.strip()
    valor_str = valor_str.strip().replace(".", "").replace(",", ".")
    if not nome or not valor_str:
        return None
    try:
        valor = float(valor_str)
    except ValueError:
        return None
    if valor < 0:
        return None
    return nome, valor

def montar_texto_resultado_custos_fixos(itens: list, total: float) -> str:
    linhas = ["🧮 *Total de contas fixas*", ""]
    for item in itens:
        linhas.append(f"- {item['nome']}: R$ {item['valor']:.2f}")
    linhas.append("")
    linhas.append(f"*Total: R$ {total:.2f}*")
    return "\n".join(linhas)

# ─────────────────────────────────────────
#  IA (GROQ) — extração estruturada
# ─────────────────────────────────────────
def get_groq_key(cliente: dict) -> str:
    if cliente.get("groq_key_override"):
        return cliente["groq_key_override"]
    return GROQ_API_KEY_1 or GROQ_API_KEY_2

PROMPT_UNIFICADO = """Você é o classificador único de intenção de um sistema de estoque via WhatsApp. Analise a
mensagem do usuário e responda APENAS com um JSON válido, sem nenhum texto antes ou depois, classificando a
intenção em UMA destas opções:
- "entrada": chegada de mercadoria (aumenta estoque)
- "venda": venda de um produto (diminui estoque)
- "saida": saída de estoque que não é venda (perda, quebra, uso interno etc.)
- "cadastro_produto": o usuário quer cadastrar um produto novo
- "edicao_produto": o usuário quer editar nome, custo, preço de venda e/ou SKU de um produto JÁ existente
- "custos_fixos": o usuário quer somar contas fixas do negócio (aluguel, luz, internet etc.), sem relação com um produto específico
- "nenhuma": a mensagem não se encaixa em nenhuma das opções acima

Responda sempre no formato abaixo, preenchendo com null os campos que não se aplicam à intenção detectada:
{
  "intencao": "entrada" | "venda" | "saida" | "cadastro_produto" | "edicao_produto" | "custos_fixos" | "nenhuma",

  "produto": "nome do produto (só para entrada/venda/saida) ou null",
  "quantidade": numero ou null,
  "valor_unitario": numero ou null,
  "cliente": "nome ou telefone do cliente — só quando intencao é venda e a mensagem menciona claramente quem comprou (ex: 'vendi 3 bolos pra Maria'), senão null",

  "nome": "nome do NOVO produto (só para cadastro_produto) ou null",
  "unidade": "un/kg/l/etc (só para cadastro_produto) ou null",
  "custo_unitario": numero ou null,
  "preco_venda": numero ou null,
  "quer_calculadora": true/false/null,
  "custo_variavel_unitario": numero ou null,
  "custo_fixo_mensal": numero ou null,
  "volume_esperado_mensal": numero ou null,
  "margem_percentual": numero ou null,

  "produto_editar": "nome do produto JÁ EXISTENTE que o usuário quer editar (só para edicao_produto) ou null",
  "nome_novo": "novo nome do produto, se quiser mudar, ou null",
  "custo_unitario_novo": numero ou null,
  "preco_venda_novo": numero ou null,
  "sku_novo": "novo SKU, se quiser mudar, ou null",

  "contas_fixas": [{"nome": "nome da conta", "valor": numero}, ...] (só para custos_fixos, uma entrada por conta) ou null
}

Regras importantes:
- Em "cadastro_produto", "quer_calculadora" deve ser true quando o usuário pede ajuda pra calcular o preço
  (menciona aluguel, luz, custo fixo, quantas unidades espera vender, margem de lucro etc.) e não informou
  custo_unitario/preco_venda diretamente.
- Em "edicao_produto", só use essa intenção quando o produto mencionado parecer já existir no estoque (ex:
  "muda o preço do bolo pra 30", "atualiza o SKU do brigadeiro pra BRG-01", "renomeia X pra Y", "o custo do
  brigadeiro agora é 2 e o nome mudou pra Brigadeiro Gourmet"). Preencha só os campos "_novo" que o usuário
  realmente mencionou; deixe os demais null.
- Em "custos_fixos", preencha "contas_fixas" com uma entrada por conta mencionada (nome + valor). NÃO some os
  valores você mesmo — apenas liste cada conta separadamente, a soma é feita fora da IA.
- Se não conseguir identificar algum campo com confiança, use null nesse campo.
- Se a mensagem não se encaixar em nenhuma intenção, responda apenas: {"intencao": "nenhuma"}
"""

async def chamar_groq_json(texto_usuario: str, groq_key: str, prompt: str = PROMPT_UNIFICADO) -> Optional[dict]:
    if not groq_key:
        print("⚠️ GROQ: nenhuma chave configurada (env var vazia)")
        return None
    messages = [
        {"role": "system", "content": prompt},
        {"role": "user", "content": texto_usuario}
    ]
    headers = {"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"}
    async with httpx.AsyncClient(timeout=25) as client:
        for modelo in GROQ_MODELOS_FALLBACK:
            payload = {"model": modelo, "temperature": 0.1, "max_tokens": 500, "messages": messages}
            try:
                resp = await client.post(GROQ_API_URL, headers=headers, json=payload)
                if resp.status_code != 200:
                    print(f"⚠️ GROQ [{modelo}] status {resp.status_code}: {resp.text[:300]}")
                    continue
                data = resp.json()
                bruto = data["choices"][0]["message"]["content"].strip()
                bruto = re.sub(r"^```json|```$", "", bruto, flags=re.MULTILINE).strip()
                return json.loads(bruto)
            except Exception as e:
                print(f"⚠️ GROQ [{modelo}] exceção: {e}")
                continue
    print("⚠️ GROQ: todos os modelos falharam")
    return None

async def enviar_whatsapp(destino: str, texto: str):
    # 'destino' pode ser um número puro (ex: "5511999998888") ou um JID completo
    # (ex: "224713024491669@lid" ou "5511999998888@s.whatsapp.net"). O baileys
    # só reconstrói "@s.whatsapp.net" quando não há "@" no valor — por isso é
    # essencial repassar o JID original (com @lid) quando ele existir, em vez
    # de normalizar para dígitos antes de responder.
    payload = {"number": destino, "message": texto}
    async with httpx.AsyncClient(timeout=10) as client:
        try:
            await asyncio.wait_for(client.post(f"{BAILEYS_URL}/disparar", json=payload), timeout=8)
        except asyncio.TimeoutError:
            print(f"⚠️ Timeout ao enviar WhatsApp para {destino}")
        except Exception as e:
            print(f"⚠️ Erro ao enviar WhatsApp para {destino}: {e}")

async def notificar_admin_novo_orcamento(conn, cliente_id: int, cliente_negocio_nome: str, produto_nome: str, quantidade: str, total_formatado: str, numero_cliente: str):
    """Envia notificação para todos os números autorizados do cliente sobre novo orçamento."""
    try:
        numeros_admin = db_all(conn, "SELECT numero FROM numeros_autorizados WHERE cliente_id = %s AND ativo = TRUE", (cliente_id,))
        if not numeros_admin:
            return
        
        mensagem = (
            f"📋 *NOVO ORÇAMENTO* 🔔\n\n"
            f"👤 Cliente: {cliente_negocio_nome}\n"
            f"📱 Telefone: {numero_cliente}\n"
            f"🛍️ Produto: {produto_nome}\n"
            f"📦 Quantidade: {quantidade}\n"
            f"💰 Total: {total_formatado}\n\n"
            f"✅ Verifique no admin para confirmar e enviar ao cliente."
        )
        
        payload = {"number": numeros_admin[0]["numero"], "message": mensagem}
        async with httpx.AsyncClient(timeout=5) as client:
            try:
                await asyncio.wait_for(client.post(f"{BAILEYS_URL}/disparar", json=payload), timeout=5)
            except Exception as e:
                print(f"⚠️ Erro ao enviar notificação: {e}")
    except Exception as e:
        print(f"⚠️ Erro ao notificar admin: {e}")

# ─────────────────────────────────────────
#  MÁQUINA DE ESTADOS — modo formulário
# ─────────────────────────────────────────
ETAPAS_TIPO = {"entrada": "entrada", "venda": "venda", "saida": "saida", "ajuste": "ajuste"}

def resposta_menu(modulos=None):
    """modulos=None mantém o comportamento antigo (menu completo), usado nos
    pontos internos das etapas do módulo estoque, que só são alcançáveis por
    quem já tem esse módulo ativo. Quando modulos é passado explicitamente e
    não inclui 'agenda', o item 13 (Agenda) é omitido."""
    if modulos is not None and "agenda" not in modulos:
        return MENU_TEXTO_SEM_AGENDA
    return MENU_TEXTO

# ═════════════════════════════════════════
#  ATENDIMENTO AO CLIENTE FINAL (visitante público, mesmo número WhatsApp)
# ═════════════════════════════════════════
# Ativado por empresa via toggle no dashboard (atendimento_cliente_final_ativado).
# Fluxo bem mais enxuto que o do funcionário: só ver produtos/serviços, pedir
# orçamento e chamar atendente humano — nada de estoque, custo, cadastro etc.
# "produtos" aqui cobre tanto produto físico quanto serviço — o texto do menu
# não usa "cardápio" porque nem todo cliente é confeitaria/comida.

MENU_CLIENTE_FINAL = (
    "👋 Olá! Em que posso ajudar?\n\n"
    "1️⃣ Ver produtos/serviços\n"
    "2️⃣ Pedir orçamento\n"
    "3️⃣ Falar com atendente\n\n"
    "Responda com o número da opção."
)

def texto_produtos_servicos_cliente_final(conn, cliente_id: int) -> str:
    produtos = listar_produtos_cliente(conn, cliente_id)
    if not produtos:
        return "No momento não temos produtos/serviços cadastrados. Digite *3* para falar com um atendente."
    return montar_lista_numerada(
        produtos, "📦 *Produtos/serviços disponíveis:*",
        rodape="Digite *menu* para voltar.", mostrar_preco=True
    )

async def processar_texto_cliente_final(conn, cliente: dict, numero: str, texto: str) -> str:
    """Fluxo simplificado pro cliente final (visitante). `numero` já vem
    normalizado (só dígitos). Sessão é isolada por (cliente_id, numero) na
    tabela sessoes_cliente_final — não interfere em nada da sessão de
    funcionário (sessoes_conversa)."""
    cliente_id = cliente["id"]
    sessao = get_or_create_sessao_cliente_final(conn, cliente_id, numero)
    etapa = sessao["etapa_atual"]
    dados = sessao["dados_parciais"] if isinstance(sessao["dados_parciais"], dict) else json.loads(sessao["dados_parciais"] or "{}")
    texto_low = texto.strip().lower()

    if texto_low in ("menu", "0", "cancelar"):
        salvar_sessao_cliente_final(conn, cliente_id, numero, "menu_cliente_final", {})
        return MENU_CLIENTE_FINAL

    if etapa == "menu_cliente_final":
        escolha = texto.strip()

        if escolha == "1":
            salvar_sessao_cliente_final(conn, cliente_id, numero, "menu_cliente_final", {})
            return texto_produtos_servicos_cliente_final(conn, cliente_id)

        if escolha == "2":
            produtos = listar_produtos_cliente(conn, cliente_id)
            if not produtos:
                return "No momento não temos produtos/serviços cadastrados. Digite *3* para falar com um atendente."
            salvar_sessao_cliente_final(conn, cliente_id, numero, "cf_orc_escolha_produto",
                                         {"produtos_ids": [p["id"] for p in produtos]})
            return montar_lista_numerada(
                produtos, "📋 *Pedir orçamento*\nQual produto/serviço você quer?",
                rodape="Responda com o número.", mostrar_preco=True
            )

        if escolha == "3":
            salvar_sessao_cliente_final(conn, cliente_id, numero, "cf_aguardando_humano", {})
            return (
                "📱 Ok! Pode escrever sua mensagem que um atendente vai te responder por aqui assim que possível.\n\n"
                "Digite *menu* a qualquer momento para voltar às opções."
            )

        return "Não entendi. " + MENU_CLIENTE_FINAL

    # ── Pedido de orçamento: escolheu o produto, agora pede a quantidade ──
    if etapa == "cf_orc_escolha_produto":
        produtos_ids = dados.get("produtos_ids", [])
        try:
            idx = int(texto.strip())
            assert 1 <= idx <= len(produtos_ids)
        except (ValueError, AssertionError):
            return f"Manda só o número do produto (1 a {len(produtos_ids)}), ou *menu* para voltar."
        produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s",
                          (produtos_ids[idx - 1], cliente_id))
        if not produto:
            salvar_sessao_cliente_final(conn, cliente_id, numero, "menu_cliente_final", {})
            return "Esse produto não está mais disponível.\n\n" + MENU_CLIENTE_FINAL
        salvar_sessao_cliente_final(conn, cliente_id, numero, "cf_orc_quantidade",
                                     {"produto_id": produto["id"], "produto_nome": produto["nome"],
                                      "preco_unitario": float(produto["preco_venda"] or 0)})
        return f"Quantas unidades de *{produto['nome']}* você quer?"

    # ── Pedido de orçamento: informou quantidade, monta e mostra o resumo ──
    if etapa == "cf_orc_quantidade":
        try:
            quantidade = float(texto.strip().replace(",", "."))
            assert quantidade > 0
        except (ValueError, AssertionError):
            return "Manda só a quantidade (um número), por favor."
        preco_unitario = float(dados.get("preco_unitario") or 0)
        total = quantidade * preco_unitario
        dados["quantidade"] = quantidade
        dados["total"] = total
        salvar_sessao_cliente_final(conn, cliente_id, numero, "cf_orc_confirmar", dados)
        return (
            f"🧾 *Resumo do orçamento*\n\n"
            f"{dados['produto_nome']} — {formatar_qtd(quantidade)} un.\n"
            f"Valor unitário: {formatar_moeda(preco_unitario)}\n"
            f"*Total: {formatar_moeda(total)}*\n\n"
            "Confirma o envio desse orçamento? (sim/não)"
        )

    # ── Confirmação final do orçamento — salva vinculado ao negócio, sem tocar estoque ──
    if etapa == "cf_orc_confirmar":
        if texto_low in ("sim", "s", "confirmar", "confirmo"):
            cliente_negocio = db_one(conn, "SELECT * FROM clientes_negocio WHERE cliente_id = %s AND telefone = %s",
                                      (cliente_id, numero))
            if not cliente_negocio:
                cliente_negocio = db_exec(conn, """
                    INSERT INTO clientes_negocio (cliente_id, nome, telefone)
                    VALUES (%s, %s, %s) RETURNING *
                """, (cliente_id, f"Visitante {numero}", numero))

            itens_json = json.dumps([{
                "produto_id": dados.get("produto_id"),
                "nome": dados.get("produto_nome"),
                "quantidade": dados.get("quantidade"),
                "preco_unitario": dados.get("preco_unitario"),
            }])
            total = float(dados.get("total") or 0)
            texto_formatado = (
                f"🧾 Orçamento — {cliente.get('nome_negocio','')}\n"
                f"{dados.get('produto_nome')} x{formatar_qtd(dados.get('quantidade') or 0)} = {formatar_moeda(total)}"
            )
            db_exec(conn, """
                INSERT INTO orcamentos (cliente_id, nome_cliente, itens, subtotal, total, texto_formatado, cliente_negocio_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (cliente_id, cliente_negocio["nome"], itens_json, total, total, texto_formatado, cliente_negocio["id"]))

            salvar_sessao_cliente_final(conn, cliente_id, numero, "menu_cliente_final", {})
            
            # Notificar admin sobre novo orçamento (não-bloqueante)
            try:
                asyncio.create_task(notificar_admin_novo_orcamento(
                    conn, 
                    cliente_id, 
                    cliente_negocio["nome"], 
                    dados.get("produto_nome", ""), 
                    formatar_qtd(dados.get("quantidade") or 0),
                    formatar_moeda(total),
                    numero
                ))
            except Exception as e:
                print(f"⚠️ Erro ao criar task de notificação: {e}")
            return "✅ Orçamento enviado! Em breve alguém confirma com você.\n\n" + MENU_CLIENTE_FINAL

        if texto_low in ("não", "nao", "n"):
            salvar_sessao_cliente_final(conn, cliente_id, numero, "menu_cliente_final", {})
            return "Sem problemas, cancelado.\n\n" + MENU_CLIENTE_FINAL

        return "Responda *sim* ou *não*."

    # ── Aguardando atendente humano: qualquer coisa que o visitante mandar aqui
    # fica registrada como orçamento pendente/observação; a dona lê pelo WhatsApp
    # normal, já que a conversa continua no mesmo número. Não trava o visitante. ──
    if etapa == "cf_aguardando_humano":
        return "Recebido! Um atendente vai te responder por aqui. Digite *menu* se quiser ver as opções de novo."

    # fallback: qualquer etapa desconhecida volta pro menu
    salvar_sessao_cliente_final(conn, cliente_id, numero, "menu_cliente_final", {})
    return MENU_CLIENTE_FINAL

async def processar_texto(conn, cliente: dict, numero_autorizado: dict, texto: str) -> str:
    numero_autorizado_id = numero_autorizado["id"]
    sessao = get_or_create_sessao(conn, numero_autorizado_id)
    etapa = sessao["etapa_atual"]
    dados = sessao["dados_parciais"] if isinstance(sessao["dados_parciais"], dict) else json.loads(sessao["dados_parciais"] or "{}")
    texto_low = texto.strip().lower()
    modulos = obter_modulos_cliente(cliente)

    if texto_low in ("menu", "cancelar", "0"):
        etapa_raiz = etapa_raiz_para_modulos(modulos)
        salvar_sessao(conn, numero_autorizado_id, etapa_raiz, {})
        return texto_raiz_para_modulos(modulos)

    # ── ETAPA: escolher módulo (só aparece pra cliente com estoque + agenda ativos) ──
    if etapa == "escolher_modulo":
        if texto.strip() == "1" and "estoque" in modulos:
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return resposta_menu(modulos)
        if texto.strip() == "2" and "agenda" in modulos:
            salvar_sessao(conn, numero_autorizado_id, "agenda_menu", {})
            return AGENDA_MENU_TEXTO
        return TEXTO_ESCOLHER_MODULO

    # ── ETAPA: MENU ──
    if etapa == "menu":
        if texto.strip() == "10":
            salvar_sessao(conn, numero_autorizado_id, "editar_menu", {})
            return (
                "🔟 *Editar Estoque ou dados do Produto*\n"
                "1 - Editar estoque (quantidade) de produto\n"
                "2 - Editar nome, custo, preço de venda ou SKU de produto\n"
                "3 - Editar estoque (quantidade) de matéria-prima\n\n"
                "Responda com o número."
            )

        if texto.strip() == "12":
            salvar_sessao(conn, numero_autorizado_id, "custosfixos_item", {"custosfixos_itens": []})
            return (
                "🧮 *Calculadora de custos fixos*\n"
                "Soma as contas fixas do seu negócio (aluguel, luz, internet etc.) — diferente da calculadora "
                "que aparece no cadastro de produto, esta não sugere preço, só soma.\n\n"
                "Manda o nome e o valor de cada conta (ex: Aluguel 2000). Digite PRONTO quando terminar."
            )

        if texto.strip() == "13" and "agenda" in modulos:
            salvar_sessao(conn, numero_autorizado_id, "agenda_menu", {})
            return (
                "📅 *Agenda*\n"
                "1 - Criar compromisso\n"
                "2 - Compromissos de hoje\n"
                "3 - Compromissos da semana\n"
                "4 - Cancelar um compromisso\n\n"
                "Responda com o número."
            )

        opcoes = {"6": "entrada", "2": "venda"}
        if texto.strip() in opcoes:
            tipo = opcoes[texto.strip()]
            produtos = listar_produtos_cliente(conn, cliente["id"])
            if not produtos:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return ("Você ainda não tem nenhum produto cadastrado. "
                        "Cadastre pelo painel (+ Novo Produto) e volte aqui depois.\n\n") + resposta_menu(modulos)
            dados = {"tipo": tipo, "produtos_ids": [p["id"] for p in produtos]}
            salvar_sessao(conn, numero_autorizado_id, f"{tipo}_produto", dados)
            rodape = "Responda com o número. Pra mais de um produto, separe por vírgula (ex: 1,3,5)."
            return montar_lista_numerada(produtos, "Qual produto?", rodape=rodape)

        if texto.strip() == "5":
            return await gerar_resumo_dia(conn, cliente["id"], modulos)

        if texto.strip() == "7":
            return gerar_visao_geral(conn, cliente["id"], modulos)

        if texto.strip() == "3":
            salvar_sessao(conn, numero_autorizado_id, "prod_nome", {})
            return "Qual o nome do novo produto?"

        if texto.strip() == "4":
            salvar_sessao(conn, numero_autorizado_id, "mp_nome", {})
            return "Qual o nome da nova matéria-prima?"

        if texto.strip() == "9":
            produtos = listar_produtos_cliente(conn, cliente["id"])
            if not produtos:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return ("Você ainda não tem nenhum produto cadastrado. "
                        "Cadastre pelo painel (+ Novo Produto) e volte aqui depois.\n\n") + resposta_menu(modulos)
            dados = {"produtos_ids": [p["id"] for p in produtos]}
            salvar_sessao(conn, numero_autorizado_id, "receita_produto_escolha", dados)
            return montar_lista_numerada(produtos, "De qual produto você quer montar/editar a receita?")

        if texto.strip() == "8":
            config = obter_config_resumo_automatico(conn, cliente["id"])
            salvar_sessao(conn, numero_autorizado_id, "config_resumo_horarios", {})
            return (
                "🔟 *Resumo automático*\n\n"
                f"Horários atuais: {formatar_horarios_config(config)}\n\n"
                "Digite até 2 horários no formato HH:MM separados por vírgula "
                "(ex: 12:00,20:00) para receber o resumo do dia automaticamente nesses horários.\n"
                "Digite *desativar* para desligar o envio automático."
            )

        if texto.strip() == "11" or texto_low in ("ajuda", "help"):
            return TEXTO_AJUDA + "\n\n" + resposta_menu(modulos)

        if texto.strip() == "1":
            produtos = listar_produtos_cliente(conn, cliente["id"])
            if not produtos:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return ("Você ainda não tem nenhum produto cadastrado. "
                        "Cadastre pelo painel (+ Novo Produto) e volte aqui depois.\n\n") + resposta_menu(modulos)
            dados = {"produtos_ids": [p["id"] for p in produtos]}
            salvar_sessao(conn, numero_autorizado_id, "orc_escolha_produtos", dados)
            rodape = "Responda com o(s) número(s) dos produtos do orçamento. Pra mais de um, separe por vírgula (ex: 1,3,5)."
            return montar_lista_numerada(produtos, "🧾 *Montar orçamento*\nQuais produtos entram?", rodape=rodape, mostrar_preco=True)

        # modo IA: UMA ÚNICA chamada à Groq classifica a intenção e já devolve
        # os campos relevantes pra ela — sem chamadas adicionais depois (TAREFA 2).
        if cliente["plano"] == "ia" and "estoque" in modulos:
            extraido = await chamar_groq_json(texto, get_groq_key(cliente))
            intencao = extraido.get("intencao") if extraido else None

            if intencao in ("entrada", "venda", "saida"):
                extraido["tipo"] = intencao  # compatibilidade com preparar_confirmacao_ia
                return await preparar_confirmacao_ia(conn, cliente, numero_autorizado_id, extraido, texto, modulos)

            if intencao == "cadastro_produto":
                resposta_ia = await iniciar_cadastro_produto_ia(conn, cliente, numero_autorizado_id, extraido)
                if resposta_ia:
                    return resposta_ia

            elif intencao == "edicao_produto":
                resposta_ia = await iniciar_edicao_produto_ia(conn, cliente, numero_autorizado_id, extraido, modulos)
                if resposta_ia:
                    return resposta_ia

            elif intencao == "custos_fixos":
                resposta_ia = iniciar_calculadora_custos_fixos_ia(extraido, modulos)
                if resposta_ia:
                    return resposta_ia

        return "Olá, bem-vindo(a) ao Painel do Seu Negócio!\n\n" + resposta_menu(modulos)

    # ── ETAPA: submenu da opção 10 — estoque (quantidade) ou dados do produto — TAREFA 1 ──
    if etapa == "editar_menu":
        if texto.strip() == "1":
            produtos = listar_produtos_cliente(conn, cliente["id"])
            if not produtos:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return ("Você ainda não tem nenhum produto cadastrado. "
                        "Cadastre pelo painel (+ Novo Produto) e volte aqui depois.\n\n") + resposta_menu(modulos)
            # Exatamente o mesmo fluxo de ajuste de estoque de sempre (múltiplos produtos, fila).
            dados = {"tipo": "ajuste", "produtos_ids": [p["id"] for p in produtos]}
            salvar_sessao(conn, numero_autorizado_id, "ajuste_produto", dados)
            rodape = "Responda com o número. Pra mais de um produto, separe por vírgula (ex: 1,3,5)."
            return montar_lista_numerada(produtos, "Qual produto?", rodape=rodape)

        if texto.strip() == "2":
            produtos = listar_produtos_cliente(conn, cliente["id"])
            if not produtos:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return ("Você ainda não tem nenhum produto cadastrado. "
                        "Cadastre pelo painel (+ Novo Produto) e volte aqui depois.\n\n") + resposta_menu(modulos)
            dados = {"produtos_ids": [p["id"] for p in produtos]}
            salvar_sessao(conn, numero_autorizado_id, "editar_campo_produto_escolha", dados)
            return montar_lista_numerada(produtos, "Qual produto você quer editar (nome/custo/preço/SKU)?")

        if texto.strip() == "3":
            materias = listar_materias_primas_cliente(conn, cliente["id"])
            if not materias:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return "Você ainda não tem matéria-prima cadastrada.\n\n" + resposta_menu(modulos)
            dados = {"materias_ids": [m["id"] for m in materias]}
            salvar_sessao(conn, numero_autorizado_id, "ajuste_mp_escolha", dados)
            rodape = "Responda com o número. Pra mais de uma, separe por vírgula (ex: 1,3,5)."
            return montar_lista_numerada(materias, "Qual matéria-prima?", rodape=rodape)

        return "Responda 1 (editar estoque de produto), 2 (editar dados do produto) ou 3 (editar estoque de matéria-prima)."

    # ── ETAPA: edição de produto — escolhendo QUAL produto (um só por vez) — TAREFA 1 ──
    if etapa == "editar_campo_produto_escolha":
        produtos_ids = dados.get("produtos_ids", [])
        try:
            idx = int(texto.strip())
            assert 1 <= idx <= len(produtos_ids)
        except (ValueError, AssertionError):
            return f"Manda só o número do produto (1 a {len(produtos_ids)})."
        produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s",
                          (produtos_ids[idx - 1], cliente["id"]))
        if not produto:
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Esse produto não existe mais.\n\n" + resposta_menu(modulos)
        dados = {
            "editar_produto_id": produto["id"],
            "editar_produto_nome": produto["nome"],
            "editar_produto_atual": {
                "nome": produto["nome"],
                "custo_unitario": float(produto["custo_unitario"] or 0),
                "preco_venda": float(produto["preco_venda"] or 0),
                "sku": produto.get("sku"),
            },
        }
        salvar_sessao(conn, numero_autorizado_id, "editar_campo_escolha", dados)
        return texto_menu_campos_editar(produto["nome"])

    # ── ETAPA: edição de produto — escolhendo quais campos mudar (aceita "2,3") — TAREFA 1 ──
    if etapa == "editar_campo_escolha":
        indices = parse_selecao_multipla(texto, len(CAMPOS_EDITAVEIS_PRODUTO))
        if not indices:
            return (f"Manda o número do campo (1 a {len(CAMPOS_EDITAVEIS_PRODUTO)}). "
                     "Pra mais de um, separe por vírgula, ex: 2,3.")
        fila_campos = [CAMPOS_EDITAVEIS_PRODUTO[i - 1][0] for i in indices]
        dados["editar_campos_fila"] = fila_campos
        dados["editar_campos_novos"] = {}
        campo_atual = dados["editar_campos_fila"].pop(0)
        dados["editar_campo_atual"] = campo_atual
        salvar_sessao(conn, numero_autorizado_id, "editar_campo_valor", dados)
        return f"Novo valor para *{CAMPO_LABEL_PRODUTO[campo_atual]}*?"

    # ── ETAPA: edição de produto — pedindo o novo valor de cada campo, um de cada vez — TAREFA 1 ──
    if etapa == "editar_campo_valor":
        campo_atual = dados.get("editar_campo_atual")
        ok, valor = parse_valor_campo_produto(campo_atual, texto)
        if not ok:
            if campo_atual in CAMPOS_NUMERICOS_PRODUTO:
                return "Manda só o número, por favor."
            return "Manda um valor válido, por favor."
        dados.setdefault("editar_campos_novos", {})[campo_atual] = valor
        dados.pop("editar_campo_atual", None)

        fila = dados.get("editar_campos_fila", [])
        if fila:
            proximo = fila.pop(0)
            dados["editar_campos_fila"] = fila
            dados["editar_campo_atual"] = proximo
            salvar_sessao(conn, numero_autorizado_id, "editar_campo_valor", dados)
            return f"Novo valor para *{CAMPO_LABEL_PRODUTO[proximo]}*?"

        salvar_sessao(conn, numero_autorizado_id, "editar_campo_confirmar", dados)
        return montar_resumo_edicao_produto(dados)

    # ── ETAPA: edição de produto — confirmação final e gravação (só campos escolhidos) — TAREFA 1 ──
    if etapa == "editar_campo_confirmar":
        if texto_low in ("sim", "s", "confirmo", "confirmar"):
            aplicar_edicao_produto(conn, cliente["id"], dados["editar_produto_id"], dados.get("editar_campos_novos", {}))
            nome_final = dados.get("editar_campos_novos", {}).get("nome", dados.get("editar_produto_nome"))
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return f"✅ *{nome_final}* atualizado com sucesso!\n\n" + resposta_menu(modulos)
        if texto_low in ("não", "nao", "n"):
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Cancelado.\n\n" + resposta_menu(modulos)
        return "Responda SIM ou NÃO."

    # ── ETAPA: ajuste de estoque de matéria-prima (opção 10 → 3) ──
    if etapa == "ajuste_mp_escolha":
        materias_ids = dados.get("materias_ids", [])
        indices = parse_selecao_multipla(texto, len(materias_ids))
        if not indices:
            return (f"Manda o número da matéria-prima (1 a {len(materias_ids)}). "
                     "Pra mais de uma, separe por vírgula, ex: 1,3.")
        escolhidas_ids = [materias_ids[i - 1] for i in indices]
        primeira = db_one(conn, "SELECT * FROM materias_primas WHERE id = %s AND cliente_id = %s",
                           (escolhidas_ids[0], cliente["id"]))
        if not primeira:
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Essa matéria-prima não existe mais.\n\n" + resposta_menu(modulos)
        dados["fila_materias_ids"] = escolhidas_ids[1:]
        dados["carrinho_mp"] = []
        dados["materia_prima_id"] = primeira["id"]
        dados["materia_prima_nome"] = primeira["nome"]
        dados["materia_prima_unidade"] = primeira["unidade"]
        salvar_sessao(conn, numero_autorizado_id, "ajuste_mp_quantidade", dados)
        return (f"Estoque atual de *{primeira['nome']}*: {fmt_num(primeira['estoque_atual'])} {primeira['unidade']}.\n"
                f"Qual a nova quantidade em estoque?")

    if etapa == "ajuste_mp_quantidade":
        try:
            quantidade = float(texto.replace(",", "."))
            assert quantidade >= 0
        except (ValueError, AssertionError):
            return "Manda só o número da nova quantidade, por favor."
        dados.setdefault("carrinho_mp", []).append({
            "materia_prima_id": dados["materia_prima_id"], "nome": dados["materia_prima_nome"],
            "unidade": dados["materia_prima_unidade"], "quantidade": quantidade,
        })

        fila = dados.get("fila_materias_ids", [])
        if fila:
            proximo_id = fila.pop(0)
            proxima = db_one(conn, "SELECT * FROM materias_primas WHERE id = %s AND cliente_id = %s",
                              (proximo_id, cliente["id"]))
            dados["fila_materias_ids"] = fila
            if not proxima:
                return "Essa matéria-prima não existe mais. Vou pular pra próxima."
            dados["materia_prima_id"] = proxima["id"]
            dados["materia_prima_nome"] = proxima["nome"]
            dados["materia_prima_unidade"] = proxima["unidade"]
            salvar_sessao(conn, numero_autorizado_id, "ajuste_mp_quantidade", dados)
            return (f"Estoque atual de *{proxima['nome']}*: {fmt_num(proxima['estoque_atual'])} {proxima['unidade']}.\n"
                    f"Qual a nova quantidade em estoque?")

        salvar_sessao(conn, numero_autorizado_id, "confirmando_mp_ajuste", dados)
        linhas = "\n".join(f"- {item['nome']} → {fmt_num(item['quantidade'])} {item['unidade']}"
                            for item in dados["carrinho_mp"])
        return f"Confirma os ajustes de estoque?\n{linhas}\n\nResponda SIM ou NÃO."

    if etapa == "confirmando_mp_ajuste":
        if texto_low in ("sim", "s", "confirmo", "confirmar"):
            alertas = []
            for item in dados.get("carrinho_mp", []):
                _, alerta = aplicar_movimentacao_materia_prima(
                    conn, cliente["id"], item["materia_prima_id"], numero_autorizado_id, "ajuste",
                    item["quantidade"], 0, origem="whatsapp"
                )
                if alerta:
                    alertas.append(alerta)
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            resposta = "✅ Estoque de matéria-prima atualizado com sucesso!"
            if alertas:
                resposta += "\n\n" + "\n".join(alertas)
            return resposta + "\n\n" + resposta_menu(modulos)
        if texto_low in ("não", "nao", "n"):
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Cancelado.\n\n" + resposta_menu(modulos)
        return "Responda SIM ou NÃO."

    # ── ETAPA: agenda — submenu (criar / listar hoje / listar semana / cancelar) ──
    if etapa == "agenda_menu":
        if texto.strip() == "1":
            salvar_sessao(conn, numero_autorizado_id, "agenda_criar_titulo", {})
            return "Qual o título do compromisso? (ex: Consulta, Reunião, Dentista)"

        if texto.strip() == "2":
            salvar_sessao(conn, numero_autorizado_id, "agenda_menu", {})
            return texto_lista_compromissos_periodo(conn, cliente["id"], dias=0) + "\n\n" + AGENDA_MENU_TEXTO

        if texto.strip() == "3":
            salvar_sessao(conn, numero_autorizado_id, "agenda_menu", {})
            return texto_lista_compromissos_periodo(conn, cliente["id"], dias=7) + "\n\n" + AGENDA_MENU_TEXTO

        if texto.strip() == "4":
            compromissos = listar_compromissos_futuros_cliente(conn, cliente["id"])
            if not compromissos:
                salvar_sessao(conn, numero_autorizado_id, "agenda_menu", {})
                return "Você não tem compromissos futuros agendados.\n\n" + AGENDA_MENU_TEXTO
            dados = {"agenda_cancelar_ids": [c["id"] for c in compromissos]}
            salvar_sessao(conn, numero_autorizado_id, "agenda_cancelar_escolha", dados)
            itens = [
                {"nome": f"{c['titulo']} — {c['data'].strftime('%d/%m')} às {str(c['hora_inicio'])[:5]}"}
                for c in compromissos
            ]
            return montar_lista_numerada(itens, "Qual compromisso cancelar?")

        return "Responda 1 (criar), 2 (hoje), 3 (semana) ou 4 (cancelar)."

    # ── ETAPA: agenda — criar compromisso: título ──
    if etapa == "agenda_criar_titulo":
        titulo = texto.strip()
        if not titulo:
            return "Manda um título válido, por favor."
        dados["agenda_titulo"] = titulo
        clientes = listar_clientes_negocio_cliente(conn, cliente["id"])
        if clientes:
            dados["agenda_clientes_opcoes"] = [{"id": c["id"], "nome": c["nome"]} for c in clientes]
            salvar_sessao(conn, numero_autorizado_id, "agenda_criar_cliente", dados)
            return montar_lista_numerada(
                clientes, "Vincular a algum cliente? (opcional)",
                rodape="Responda o número, ou digite PULAR."
            )
        salvar_sessao(conn, numero_autorizado_id, "agenda_criar_data", dados)
        return "Qual a data? (formato DD/MM ou DD/MM/AAAA)"

    # ── ETAPA: agenda — criar compromisso: cliente vinculado (opcional) ──
    if etapa == "agenda_criar_cliente":
        if texto_low in ("pular", "não", "nao", "n"):
            dados["agenda_cliente_negocio_id"] = None
            dados.pop("agenda_clientes_opcoes", None)
            salvar_sessao(conn, numero_autorizado_id, "agenda_criar_data", dados)
            return "Qual a data? (formato DD/MM ou DD/MM/AAAA)"
        opcoes_cliente = dados.get("agenda_clientes_opcoes", [])
        try:
            idx = int(texto.strip())
            assert 1 <= idx <= len(opcoes_cliente)
        except (ValueError, AssertionError):
            return f"Manda o número do cliente (1 a {len(opcoes_cliente)}), ou digite PULAR."
        escolhido = opcoes_cliente[idx - 1]
        dados["agenda_cliente_negocio_id"] = escolhido["id"]
        dados["agenda_cliente_negocio_nome"] = escolhido["nome"]
        dados.pop("agenda_clientes_opcoes", None)
        salvar_sessao(conn, numero_autorizado_id, "agenda_criar_data", dados)
        return "Qual a data? (formato DD/MM ou DD/MM/AAAA)"

    # ── ETAPA: agenda — criar compromisso: data ──
    if etapa == "agenda_criar_data":
        data_obj = parse_data_br(texto)
        if not data_obj:
            return "Data inválida. Manda no formato DD/MM ou DD/MM/AAAA."
        dados["agenda_data"] = data_obj.isoformat()
        salvar_sessao(conn, numero_autorizado_id, "agenda_criar_hora", dados)
        return "Que horas? (formato HH:MM)"

    # ── ETAPA: agenda — criar compromisso: hora ──
    if etapa == "agenda_criar_hora":
        m = re.fullmatch(r"(\d{1,2}):(\d{2})", texto.strip())
        if not m:
            return "Hora inválida. Manda no formato HH:MM (ex: 14:30)."
        h, mi = int(m.group(1)), int(m.group(2))
        if h > 23 or mi > 59:
            return "Hora inválida. Manda no formato HH:MM (ex: 14:30)."
        dados["agenda_hora"] = f"{h:02d}:{mi:02d}"
        salvar_sessao(conn, numero_autorizado_id, "agenda_criar_lembrete", dados)
        return "Quer receber um lembrete antes por WhatsApp? Responda quantos minutos antes (ex: 30), ou NÃO."

    # ── ETAPA: agenda — criar compromisso: lembrete ──
    if etapa == "agenda_criar_lembrete":
        if texto_low in ("não", "nao", "n"):
            dados["agenda_lembrete_minutos"] = None
        else:
            try:
                minutos = int(texto.strip())
                assert minutos > 0
            except (ValueError, AssertionError):
                return "Manda só o número de minutos antes (ex: 30), ou responda NÃO."
            dados["agenda_lembrete_minutos"] = minutos
        salvar_sessao(conn, numero_autorizado_id, "agenda_criar_confirmar", dados)
        return montar_resumo_agenda_confirmacao(conn, cliente["id"], dados)

    # ── ETAPA: agenda — criar compromisso: confirmação final e gravação ──
    if etapa == "agenda_criar_confirmar":
        if texto_low in ("sim", "s", "confirmo", "confirmar"):
            data_obj = datetime.fromisoformat(dados["agenda_data"]).date()
            hora_inicio = dados["agenda_hora"]
            hora_fim = calcular_hora_fim(hora_inicio)
            db_exec(conn, """
                INSERT INTO agenda_compromissos
                    (cliente_id, cliente_negocio_id, titulo, data, hora_inicio, hora_fim, lembrete_minutos_antes)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (cliente["id"], dados.get("agenda_cliente_negocio_id"), dados["agenda_titulo"],
                  data_obj, hora_inicio, hora_fim, dados.get("agenda_lembrete_minutos")))
            salvar_sessao(conn, numero_autorizado_id, "agenda_menu", {})
            return "✅ Compromisso agendado com sucesso!\n\n" + AGENDA_MENU_TEXTO
        if texto_low in ("não", "nao", "n"):
            salvar_sessao(conn, numero_autorizado_id, "agenda_menu", {})
            return "Cancelado.\n\n" + AGENDA_MENU_TEXTO
        return "Responda SIM ou NÃO."

    # ── ETAPA: agenda — cancelar um compromisso existente ──
    if etapa == "agenda_cancelar_escolha":
        ids = dados.get("agenda_cancelar_ids", [])
        try:
            idx = int(texto.strip())
            assert 1 <= idx <= len(ids)
        except (ValueError, AssertionError):
            return f"Manda o número do compromisso (1 a {len(ids)})."
        compromisso_id = ids[idx - 1]
        db_exec(conn, "UPDATE agenda_compromissos SET status = 'cancelado' WHERE id = %s AND cliente_id = %s",
                (compromisso_id, cliente["id"]))
        salvar_sessao(conn, numero_autorizado_id, "agenda_menu", {})
        return "✅ Compromisso cancelado.\n\n" + AGENDA_MENU_TEXTO

    # ── ETAPA: calculadora de custos fixos — coletando conta por conta até PRONTO — TAREFA 3 ──
    if etapa == "custosfixos_item":
        if texto_low == "pronto":
            itens = dados.get("custosfixos_itens", [])
            if not itens:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return "Nenhuma conta adicionada. Calculadora cancelada.\n\n" + resposta_menu(modulos)
            total = sum(item["valor"] for item in itens)  # soma sempre em Python
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return montar_texto_resultado_custos_fixos(itens, total) + "\n\n" + resposta_menu(modulos)

        resultado = parse_conta_fixa_texto(texto)
        if not resultado:
            return ("Não entendi. Manda o nome e o valor da conta (ex: Aluguel 2000), "
                     "ou digite PRONTO para terminar.")
        nome_conta, valor_conta = resultado
        dados.setdefault("custosfixos_itens", []).append({"nome": nome_conta, "valor": valor_conta})
        salvar_sessao(conn, numero_autorizado_id, "custosfixos_item", dados)
        total_parcial = sum(item["valor"] for item in dados["custosfixos_itens"])
        return (
            f"Adicionado: {nome_conta} — R$ {valor_conta:.2f}\n"
            f"Total parcial: R$ {total_parcial:.2f}\n\n"
            "Manda a próxima conta, ou digite PRONTO para terminar."
        )

    # ── ETAPA: escolhendo o(s) produto(s) da lista numerada ──
    if etapa.endswith("_produto"):
        tipo = dados.get("tipo")
        produtos_ids = dados.get("produtos_ids", [])
        indices = parse_selecao_multipla(texto, len(produtos_ids))
        if not indices:
            return (f"Manda o número do produto (1 a {len(produtos_ids)}). "
                     "Pra mais de um, separe por vírgula, ex: 1,3.")
        escolhidos_ids = [produtos_ids[i - 1] for i in indices]

        if tipo == "consulta":
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            blocos = []
            for pid in escolhidos_ids:
                produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s", (pid,))
                if not produto:
                    continue
                blocos.append(
                    f"📦 *{produto['nome']}*\n"
                    f"Estoque atual: {fmt_num(produto['estoque_atual'])} {produto['unidade']}\n"
                    f"Custo: R$ {produto['custo_unitario']:.2f}\n"
                    f"Preço de venda: R$ {produto['preco_venda']:.2f}"
                )
            return "\n\n".join(blocos) + "\n\n" + resposta_menu(modulos)

        # entrada / venda / ajuste: monta a fila e processa item por item
        primeiro = db_one(conn, "SELECT * FROM produtos WHERE id = %s", (escolhidos_ids[0],))
        dados["fila_produtos_ids"] = escolhidos_ids[1:]
        dados["carrinho"] = []
        dados["produto_id"] = primeiro["id"]
        dados["produto_nome"] = primeiro["nome"]
        salvar_sessao(conn, numero_autorizado_id, f"{tipo}_quantidade", dados)
        return f"Quantidade de *{primeiro['nome']}*?"

    # ── ETAPA: pedindo quantidade (de um item do carrinho) ──
    if etapa.endswith("_quantidade"):
        tipo = dados.get("tipo")
        try:
            quantidade = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número da quantidade, por favor."
        dados["quantidade"] = quantidade

        if tipo == "ajuste":
            dados.setdefault("carrinho", []).append({
                "produto_id": dados["produto_id"], "produto_nome": dados["produto_nome"],
                "quantidade": quantidade, "valor_unitario": 0,
            })
            return avancar_fila_ou_confirmar(conn, numero_autorizado_id, dados, tipo)

        # entrada / venda: sugere o valor já cadastrado no produto
        produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s", (dados["produto_id"],))
        valor_sugerido = float((produto["custo_unitario"] if tipo == "entrada" else produto["preco_venda"]) or 0)
        dados["valor_sugerido"] = valor_sugerido
        salvar_sessao(conn, numero_autorizado_id, f"{tipo}_valor", dados)
        rotulo = "custo unitário" if tipo == "entrada" else "valor unitário de venda"
        if valor_sugerido > 0:
            return (
                f"{rotulo.capitalize()} de *{dados['produto_nome']}*: R$ {valor_sugerido:.2f} (já cadastrado no painel).\n"
                f"Responda OK para manter esse valor, ou digite outro valor."
            )
        return f"Qual o {rotulo} (R$) de *{dados['produto_nome']}*?"

    # ── ETAPA: pedindo/confirmando valor de um item do carrinho ──
    if etapa.endswith("_valor"):
        tipo = dados.get("tipo")
        valor_sugerido = dados.get("valor_sugerido", 0) or 0
        if texto_low in ("ok", "sim", "s", "manter", "confirmo") and valor_sugerido > 0:
            valor = valor_sugerido
        else:
            try:
                valor = float(texto.replace(",", "."))
            except ValueError:
                if valor_sugerido > 0:
                    return "Responda OK para manter o valor sugerido, ou digite um número."
                return "Manda o valor em número, por favor."

        dados.setdefault("carrinho", []).append({
            "produto_id": dados["produto_id"], "produto_nome": dados["produto_nome"],
            "quantidade": dados["quantidade"], "valor_unitario": valor,
        })
        return avancar_fila_ou_confirmar(conn, numero_autorizado_id, dados, tipo)

    # ── ETAPA: venda — identificar o cliente (opcional, TAREFA 2) ──
    if etapa == "venda_cliente":
        if texto_low != "pular":
            cliente_negocio = buscar_ou_criar_cliente_negocio(conn, cliente["id"], texto)
            if cliente_negocio:
                dados["cliente_negocio_id"] = cliente_negocio["id"]
                dados["cliente_negocio_nome"] = cliente_negocio["nome"]
        salvar_sessao(conn, numero_autorizado_id, "confirmando", dados)
        texto_confirmacao = montar_texto_confirmacao_carrinho(dados, "venda")
        if dados.get("cliente_negocio_nome"):
            texto_confirmacao = f"Cliente: {dados['cliente_negocio_nome']}\n" + texto_confirmacao
        return texto_confirmacao

    # ── ETAPA: cadastro de matéria-prima (opção 6) ──
    if etapa == "mp_nome":
        dados["mp_nome"] = texto.strip()
        salvar_sessao(conn, numero_autorizado_id, "mp_unidade", dados)
        return "Qual a unidade de medida? (ex: kg, g, l, ml, un) — ou digite PULAR para usar 'un'"

    if etapa == "mp_unidade":
        dados["mp_unidade"] = texto.strip() if texto_low != "pular" else "un"
        salvar_sessao(conn, numero_autorizado_id, "mp_custo", dados)
        return "Qual o custo unitário (R$)? — digite 0 se ainda não souber"

    if etapa == "mp_custo":
        try:
            custo = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número do custo, por favor."
        dados["mp_custo"] = custo
        salvar_sessao(conn, numero_autorizado_id, "mp_estoque", dados)
        return "Qual o estoque inicial dessa matéria-prima?"

    if etapa == "mp_estoque":
        try:
            estoque = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número do estoque, por favor."
        dados["mp_estoque"] = estoque
        salvar_sessao(conn, numero_autorizado_id, "mp_estoque_minimo", dados)
        return (
            "Quer receber um aviso quando essa matéria-prima ficar baixa? "
            "Se sim, digite o estoque mínimo (ex: 1). Se não quiser usar isso, digite PULAR."
        )

    if etapa == "mp_estoque_minimo":
        if texto_low == "pular":
            dados["mp_estoque_minimo"] = None
        else:
            try:
                dados["mp_estoque_minimo"] = float(texto.replace(",", "."))
            except ValueError:
                return "Manda só o número do estoque mínimo, ou digite PULAR pra não usar essa opção."
        salvar_sessao(conn, numero_autorizado_id, "confirmando_mp", dados)
        linha_minimo = (f" | alerta abaixo de {fmt_num(dados['mp_estoque_minimo'])}"
                         if dados.get("mp_estoque_minimo") is not None else "")
        return (
            f"Confirma o cadastro?\n"
            f"*{dados['mp_nome']}* | {dados['mp_unidade']} | custo R$ {dados['mp_custo']:.2f} | "
            f"estoque inicial {fmt_num(dados['mp_estoque'])}{linha_minimo}\n\nResponda SIM ou NÃO."
        )

    if etapa == "confirmando_mp":
        if texto_low in ("sim", "s", "confirmo", "confirmar"):
            existente = buscar_materia_prima_por_nome(conn, cliente["id"], dados["mp_nome"])
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            if existente:
                return f"⚠️ Já existe uma matéria-prima chamada '{existente['nome']}'. Use o painel pra editar.\n\n" + resposta_menu(modulos)
            db_exec(conn, """
                INSERT INTO materias_primas (cliente_id, nome, unidade, custo_unitario, estoque_atual, estoque_minimo)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (cliente["id"], dados["mp_nome"], dados["mp_unidade"], dados["mp_custo"], dados["mp_estoque"],
                  dados.get("mp_estoque_minimo")))
            return "✅ Matéria-prima cadastrada com sucesso!\n\n" + resposta_menu(modulos)
        if texto_low in ("não", "nao", "n"):
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Cancelado.\n\n" + resposta_menu(modulos)
        return "Responda SIM ou NÃO."

    # ── ETAPA: cadastro de produto (opção 8) ──
    if etapa == "prod_nome":
        dados["prod_nome"] = texto.strip()
        salvar_sessao(conn, numero_autorizado_id, "prod_unidade", dados)
        return "Qual a unidade de medida? (ex: un, kg, l) — ou digite PULAR para usar 'un'"

    if etapa == "prod_unidade":
        dados["prod_unidade"] = texto.strip() if texto_low != "pular" else "un"
        salvar_sessao(conn, numero_autorizado_id, "prod_custo_modo", dados)
        return (
            "Você já sabe o custo e o preço de venda, ou quer ajuda pra calcular?\n"
            "1 - Já sei\n"
            "2 - Quer ajuda pra calcular"
        )

    # ── ETAPA: escolhe entre informar custo/preço direto ou usar a calculadora (TAREFA 1) ──
    if etapa == "prod_custo_modo":
        if texto_low in ("1", "ja sei", "já sei", "sei"):
            salvar_sessao(conn, numero_autorizado_id, "prod_custo", dados)
            return "Qual o custo unitário (R$)? — digite 0 se ainda não souber"
        if texto_low in ("2", "quero ajuda", "quer ajuda", "ajuda", "calcular", "não sei", "nao sei"):
            salvar_sessao(conn, numero_autorizado_id, "calc_pergunta_receita", dados)
            return (
                "Vamos calcular! Esse produto vai ter uma ficha técnica (receita com matérias-primas)? "
                "Se sim, calculo o custo variável automaticamente a partir dela.\n"
                "Se não, você digita o custo variável por unidade direto.\n\nResponda SIM ou NÃO."
            )
        return "Responda 1 (já sei o custo/preço) ou 2 (quero ajuda pra calcular)."

    # ── ETAPA: calculadora — quer montar a ficha técnica agora? ──
    if etapa == "calc_pergunta_receita":
        if texto_low in ("sim", "s"):
            materias = listar_materias_primas_cliente(conn, cliente["id"])
            if not materias:
                dados["calc_receita_itens"] = []
                salvar_sessao(conn, numero_autorizado_id, "calc_custo_variavel_manual", dados)
                return (
                    "Você ainda não tem matéria-prima cadastrada, então não dá pra montar a ficha técnica agora.\n"
                    "Qual o custo variável por unidade (R$)? (ingredientes, embalagem etc. — ou 0 se não tiver)"
                )
            dados["calc_receita_itens"] = []
            dados["calc_materias_ids"] = [m["id"] for m in materias]
            salvar_sessao(conn, numero_autorizado_id, "calc_receita_item_escolha", dados)
            return montar_lista_numerada(
                materias, "Qual matéria-prima entra na receita?",
                rodape="Responda com o número, ou digite PRONTO quando terminar."
            )
        if texto_low in ("não", "nao", "n"):
            dados["calc_receita_itens"] = []
            salvar_sessao(conn, numero_autorizado_id, "calc_custo_variavel_manual", dados)
            return "Qual o custo variável por unidade (R$)? (ingredientes, embalagem etc. — ou 0 se não tiver)"
        return "Responda SIM ou NÃO."

    # ── ETAPA: calculadora — montando a ficha técnica ──
    if etapa == "calc_receita_item_escolha":
        if texto_low == "pronto":
            dados["prod_quer_receita"] = bool(dados.get("calc_receita_itens"))
            if dados["prod_quer_receita"]:
                dados["receita_itens"] = dados["calc_receita_itens"]
                dados["calc_receita_pronta"] = True
            salvar_sessao(conn, numero_autorizado_id, "calc_custo_fixo", dados)
            return (
                "Qual o custo fixo mensal total do negócio (aluguel, luz, embalagem, mão de obra somados)?\n"
                "(esses valores não ficam salvos, é só pra calcular o preço agora)"
            )
        materias_ids = dados.get("calc_materias_ids", [])
        try:
            idx = int(texto.strip())
            assert 1 <= idx <= len(materias_ids)
        except (ValueError, AssertionError):
            return f"Manda só o número da matéria-prima (1 a {len(materias_ids)}), ou PRONTO para terminar."
        materia = db_one(conn, "SELECT * FROM materias_primas WHERE id = %s", (materias_ids[idx - 1],))
        if not materia:
            return "Essa matéria-prima não existe mais. Escolha outro número ou digite PRONTO."
        dados["calc_item_atual"] = {
            "id": materia["id"], "nome": materia["nome"], "unidade": materia["unidade"],
            "custo_unitario": float(materia["custo_unitario"] or 0),
        }
        salvar_sessao(conn, numero_autorizado_id, "calc_receita_item_qtd", dados)
        return f"Quantos {materia['unidade']} de *{materia['nome']}* vão em 1 unidade do produto?"

    if etapa == "calc_receita_item_qtd":
        try:
            quantidade = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número da quantidade, por favor."
        item = dados["calc_item_atual"]
        dados.setdefault("calc_receita_itens", []).append({
            "materia_prima_id": item["id"], "nome": item["nome"], "unidade": item["unidade"],
            "quantidade": quantidade, "custo_unitario": item["custo_unitario"],
        })
        dados.pop("calc_item_atual", None)
        materias = listar_materias_primas_cliente(conn, cliente["id"])
        dados["calc_materias_ids"] = [m["id"] for m in materias]
        salvar_sessao(conn, numero_autorizado_id, "calc_receita_item_escolha", dados)
        return montar_lista_numerada(
            materias, "Adicionado! Mais alguma matéria-prima?",
            rodape="Responda com o número, ou digite PRONTO para terminar."
        )

    # ── ETAPA: calculadora — custo variável informado direto (sem ficha técnica) ──
    if etapa == "calc_custo_variavel_manual":
        try:
            valor = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número do custo variável, por favor."
        dados["custo_variavel"] = valor
        salvar_sessao(conn, numero_autorizado_id, "calc_custo_fixo", dados)
        return (
            "Qual o custo fixo mensal total do negócio (aluguel, luz, embalagem, mão de obra somados)?\n"
            "(esses valores não ficam salvos, é só pra calcular o preço agora)"
        )

    # ── ETAPA: calculadora — custo fixo mensal ──
    if etapa == "calc_custo_fixo":
        try:
            valor = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número do custo fixo mensal, por favor."
        dados["calc_custo_fixo_mensal"] = valor
        salvar_sessao(conn, numero_autorizado_id, "calc_volume", dados)
        return "Quantas unidades desse produto você espera vender por mês?"

    # ── ETAPA: calculadora — volume esperado de vendas ──
    if etapa == "calc_volume":
        try:
            valor = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número do volume esperado, por favor."
        if valor <= 0:
            return "O volume esperado precisa ser maior que zero (pra dar pra ratear o custo fixo)."
        dados["calc_volume_esperado"] = valor
        salvar_sessao(conn, numero_autorizado_id, "calc_margem", dados)
        return "Qual a margem de lucro desejada em %? (sugestão: 30% — digite PULAR para usar 30%)"

    # ── ETAPA: calculadora — margem desejada e cálculo final ──
    if etapa == "calc_margem":
        if texto_low == "pular":
            margem = 30.0
        else:
            try:
                margem = float(texto.replace(",", "."))
            except ValueError:
                return "Manda só o número da margem (%), ou PULAR para usar 30%."
        dados["calc_margem"] = margem
        resultado = calcular_resultado_calculadora(dados)
        dados["calc_resultado"] = resultado
        salvar_sessao(conn, numero_autorizado_id, "calc_confirmar", dados)
        return texto_resultado_calculadora(dados, resultado)

    # ── ETAPA: calculadora — confirma usar o custo/preço calculado ──
    if etapa == "calc_confirmar":
        if texto_low in ("sim", "s", "confirmo", "confirmar"):
            resultado = dados.get("calc_resultado", {})
            dados["prod_custo"] = resultado.get("custo_total_unitario", 0)
            dados["prod_preco"] = resultado.get("preco_sugerido", 0)
            # Se veio pelo caminho "custo variável manual" (sem ficha técnica), prod_quer_receita
            # nunca foi marcado — sem isso, a etapa prod_estoque pergunta de novo "usa receita?".
            if "prod_quer_receita" not in dados:
                dados["prod_quer_receita"] = bool(dados.get("calc_receita_itens"))
            salvar_sessao(conn, numero_autorizado_id, "prod_estoque", dados)
            return "Qual o estoque inicial desse produto?"
        if texto_low in ("não", "nao", "n"):
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Cadastro cancelado.\n\n" + resposta_menu(modulos)
        return "Responda SIM ou NÃO."

    if etapa == "prod_custo":
        try:
            custo = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número do custo, por favor."
        dados["prod_custo"] = custo
        salvar_sessao(conn, numero_autorizado_id, "prod_preco", dados)
        return "Qual o valor de venda (R$)? — digite 0 se ainda não souber"

    if etapa == "prod_preco":
        try:
            preco = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número do valor de venda, por favor."
        dados["prod_preco"] = preco
        salvar_sessao(conn, numero_autorizado_id, "prod_estoque", dados)
        return "Qual o estoque inicial desse produto?"

    if etapa == "prod_estoque":
        try:
            estoque = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número do estoque, por favor."
        dados["prod_estoque"] = estoque

        # Se veio da calculadora (ou da IA), "usa receita?" já foi respondido — pula a pergunta.
        if "prod_quer_receita" in dados:
            salvar_sessao(conn, numero_autorizado_id, "produto_aguardando_confirmacao", dados)
            return (
                f"Confirma o cadastro?\n"
                f"*{dados['prod_nome']}* | {dados['prod_unidade']} | custo R$ {dados['prod_custo']:.2f} | "
                f"venda R$ {dados['prod_preco']:.2f} | estoque inicial {fmt_num(dados['prod_estoque'])}"
                f"\n\nResponda SIM ou NÃO."
            )

        salvar_sessao(conn, numero_autorizado_id, "prod_quer_receita", dados)
        return (
            "Esse produto usa alguma matéria-prima na receita (ex: farinha, embalagem)? "
            "Se sim, a matéria-prima é descontada sozinha a cada venda. Responda SIM ou NÃO."
        )

    if etapa == "prod_quer_receita":
        if texto_low in ("sim", "s"):
            dados["prod_quer_receita"] = True
        elif texto_low in ("não", "nao", "n"):
            dados["prod_quer_receita"] = False
        else:
            return "Responda SIM ou NÃO."
        salvar_sessao(conn, numero_autorizado_id, "produto_aguardando_confirmacao", dados)
        return (
            f"Confirma o cadastro?\n"
            f"*{dados['prod_nome']}* | {dados['prod_unidade']} | custo R$ {dados['prod_custo']:.2f} | "
            f"venda R$ {dados['prod_preco']:.2f} | estoque inicial {fmt_num(dados['prod_estoque'])}"
            f"\n\nResponda SIM ou NÃO."
        )

    if etapa == "produto_aguardando_confirmacao":
        if texto_low in ("sim", "s", "confirmo", "confirmar"):
            existente = buscar_produto_por_nome(conn, cliente["id"], dados["prod_nome"])
            if existente:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return f"⚠️ Já existe um produto chamado '{existente['nome']}'. Use o painel pra editar.\n\n" + resposta_menu(modulos)
            produto_novo = db_exec(conn, """
                INSERT INTO produtos (cliente_id, nome, unidade, custo_unitario, preco_venda, estoque_atual)
                VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
            """, (cliente["id"], dados["prod_nome"], dados["prod_unidade"], dados["prod_custo"],
                  dados["prod_preco"], dados["prod_estoque"]))

            if not dados.get("prod_quer_receita"):
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return "✅ Produto cadastrado com sucesso!\n\n" + resposta_menu(modulos)

            # TAREFA 1 — a receita já foi montada durante a calculadora: salva direto, sem perguntar de novo.
            if dados.get("calc_receita_pronta") and dados.get("receita_itens"):
                for item in dados["receita_itens"]:
                    db_exec(conn, """
                        INSERT INTO receita_itens (produto_id, materia_prima_id, quantidade_necessaria)
                        VALUES (%s,%s,%s)
                    """, (produto_novo["id"], item["materia_prima_id"], item["quantidade"]))
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return ("✅ Produto cadastrado com receita e preço já calculados!\n\n"
                        "A partir de agora, vender esse produto já desconta a matéria-prima automaticamente.\n\n"
                        ) + resposta_menu(modulos)

            # Quis vincular receita — encadeia direto no mesmo fluxo da opção 7,
            # já com o produto recém-criado.
            materias = listar_materias_primas_cliente(conn, cliente["id"])
            if not materias:
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return ("✅ Produto cadastrado com sucesso!\n\n"
                        "Só que você ainda não tem matéria-prima cadastrada. Cadastre uma (opção 6) "
                        "e depois monte a receita pela opção 7.\n\n") + resposta_menu(modulos)

            dados_receita = {
                "receita_produto_id": produto_novo["id"], "receita_produto_nome": produto_novo["nome"],
                "receita_itens": [], "materias_ids": [m["id"] for m in materias],
            }
            salvar_sessao(conn, numero_autorizado_id, "receita_item_escolha", dados_receita)
            return "✅ Produto cadastrado! Agora vamos montar a receita.\n\n" + montar_lista_numerada(
                materias, f"Qual matéria-prima entra em *{produto_novo['nome']}*?",
                rodape="Responda com o número, ou digite PRONTO quando terminar."
            )
        if texto_low in ("não", "nao", "n"):
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Cancelado.\n\n" + resposta_menu(modulos)
        return "Responda SIM ou NÃO."

    # ── ETAPA: editar receita de um produto (opção 9 — Editar Matéria-Prima) ──
    if etapa == "receita_produto_escolha":
        produtos_ids = dados.get("produtos_ids", [])
        try:
            idx = int(texto.strip())
            assert 1 <= idx <= len(produtos_ids)
        except (ValueError, AssertionError):
            return f"Manda só o número do produto (1 a {len(produtos_ids)})."
        produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s", (produtos_ids[idx - 1],))
        if not produto:
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Esse produto não existe mais.\n\n" + resposta_menu(modulos)

        materias = listar_materias_primas_cliente(conn, cliente["id"])
        if not materias:
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Você ainda não tem matéria-prima cadastrada. Cadastre uma primeiro (opção 4).\n\n" + resposta_menu(modulos)

        receita_atual = db_all(conn, """
            SELECT r.materia_prima_id, r.quantidade_necessaria, m.nome, m.unidade
            FROM receita_itens r
            JOIN materias_primas m ON m.id = r.materia_prima_id
            WHERE r.produto_id = %s
            ORDER BY m.nome
        """, (produto["id"],))
        receita_itens = [
            {"materia_prima_id": i["materia_prima_id"], "nome": i["nome"], "unidade": i["unidade"],
             "quantidade": float(i["quantidade_necessaria"])}
            for i in receita_atual
        ]

        dados = {
            "receita_produto_id": produto["id"], "receita_produto_nome": produto["nome"],
            "receita_itens": receita_itens,
        }
        salvar_sessao(conn, numero_autorizado_id, "receita_menu_edicao", dados)
        return montar_menu_edicao_receita(dados)

    if etapa == "receita_menu_edicao":
        if texto_low == "pronto":
            if not dados.get("receita_itens"):
                salvar_sessao(conn, numero_autorizado_id, "menu", {})
                return "Nenhum item na receita. Nada foi salvo.\n\n" + resposta_menu(modulos)
            salvar_sessao(conn, numero_autorizado_id, "confirmando_receita", dados)
            linhas = "\n".join(f"- {fmt_num(i['quantidade'])} {i['unidade']} de {i['nome']}" for i in dados["receita_itens"])
            return f"Confirma a receita de *{dados['receita_produto_nome']}*?\n{linhas}\n\nResponda SIM ou NÃO."

        if texto_low == "novo":
            usadas_ids = {i["materia_prima_id"] for i in dados.get("receita_itens", [])}
            materias = [m for m in listar_materias_primas_cliente(conn, cliente["id"]) if m["id"] not in usadas_ids]
            if not materias:
                return "Todas as matérias-primas cadastradas já estão nessa receita.\n\n" + montar_menu_edicao_receita(dados)
            dados["materias_ids"] = [m["id"] for m in materias]
            salvar_sessao(conn, numero_autorizado_id, "receita_item_escolha", dados)
            return montar_lista_numerada(
                materias, "Qual matéria-prima entra na receita?",
                rodape="Responda com o número, ou digite CANCELAR para voltar."
            )

        if texto_low.startswith("remover"):
            partes = texto.strip().split()
            try:
                pos = int(partes[1])
                assert 1 <= pos <= len(dados.get("receita_itens", []))
            except (IndexError, ValueError, AssertionError):
                return f"Manda *remover N*, com N de 1 a {len(dados.get('receita_itens', []))}.\n\n" + montar_menu_edicao_receita(dados)
            removido = dados["receita_itens"].pop(pos - 1)
            salvar_sessao(conn, numero_autorizado_id, "receita_menu_edicao", dados)
            return f"Removido: {fmt_num(removido['quantidade'])} {removido['unidade']} de {removido['nome']}.\n\n" + montar_menu_edicao_receita(dados)

        try:
            pos = int(texto.strip())
            assert 1 <= pos <= len(dados.get("receita_itens", []))
        except (ValueError, AssertionError):
            return "Não entendi. " + montar_menu_edicao_receita(dados)
        item = dados["receita_itens"][pos - 1]
        dados["receita_editar_idx"] = pos - 1
        salvar_sessao(conn, numero_autorizado_id, "receita_editar_qtd", dados)
        return (f"Quantidade atual de *{item['nome']}*: {fmt_num(item['quantidade'])} {item['unidade']}.\n"
                f"Manda a nova quantidade de {item['unidade']} por unidade do produto.")

    if etapa == "receita_editar_qtd":
        try:
            quantidade = float(texto.replace(",", "."))
            assert quantidade > 0
        except (ValueError, AssertionError):
            return "Manda só o número da nova quantidade, por favor."
        idx = dados.get("receita_editar_idx")
        dados["receita_itens"][idx]["quantidade"] = quantidade
        dados.pop("receita_editar_idx", None)
        salvar_sessao(conn, numero_autorizado_id, "receita_menu_edicao", dados)
        return "Quantidade atualizada!\n\n" + montar_menu_edicao_receita(dados)

    if etapa == "receita_item_escolha":
        if texto_low == "cancelar":
            salvar_sessao(conn, numero_autorizado_id, "receita_menu_edicao", dados)
            return montar_menu_edicao_receita(dados)

        materias_ids = dados.get("materias_ids", [])
        try:
            idx = int(texto.strip())
            assert 1 <= idx <= len(materias_ids)
        except (ValueError, AssertionError):
            return f"Manda só o número da matéria-prima (1 a {len(materias_ids)}), ou CANCELAR para voltar."
        materia = db_one(conn, "SELECT * FROM materias_primas WHERE id = %s", (materias_ids[idx - 1],))
        if not materia:
            return "Essa matéria-prima não existe mais. Escolha outro número ou digite CANCELAR."
        dados["receita_item_atual"] = {"id": materia["id"], "nome": materia["nome"], "unidade": materia["unidade"]}
        salvar_sessao(conn, numero_autorizado_id, "receita_item_qtd", dados)
        return f"Quantos {materia['unidade']} de *{materia['nome']}* vão em 1 unidade do produto?"

    if etapa == "receita_item_qtd":
        try:
            quantidade = float(texto.replace(",", "."))
            assert quantidade > 0
        except (ValueError, AssertionError):
            return "Manda só o número da quantidade, por favor."
        item = dados["receita_item_atual"]
        dados.setdefault("receita_itens", []).append({
            "materia_prima_id": item["id"], "nome": item["nome"], "unidade": item["unidade"], "quantidade": quantidade
        })
        dados.pop("receita_item_atual", None)
        dados.pop("materias_ids", None)
        salvar_sessao(conn, numero_autorizado_id, "receita_menu_edicao", dados)
        return "Adicionado!\n\n" + montar_menu_edicao_receita(dados)

    if etapa == "confirmando_receita":
        if texto_low in ("sim", "s", "confirmo", "confirmar"):
            produto_id = dados["receita_produto_id"]
            db_exec(conn, "DELETE FROM receita_itens WHERE produto_id = %s", (produto_id,))
            for item in dados["receita_itens"]:
                db_exec(conn, """
                    INSERT INTO receita_itens (produto_id, materia_prima_id, quantidade_necessaria)
                    VALUES (%s,%s,%s)
                """, (produto_id, item["materia_prima_id"], item["quantidade"]))
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "✅ Receita salva com sucesso! A partir de agora, vender esse produto já desconta a matéria-prima automaticamente.\n\n" + resposta_menu(modulos)
        if texto_low in ("não", "nao", "n"):
            salvar_sessao(conn, numero_autorizado_id, "receita_menu_edicao", dados)
            return "Ok, voltando pra edição.\n\n" + montar_menu_edicao_receita(dados)
        return "Responda SIM ou NÃO."

    # ── ETAPA: confirmando ──
    if etapa == "confirmando":
        if texto_low in ("sim", "s", "confirmo", "confirmar"):
            tipo = dados.get("tipo")
            carrinho = dados.get("carrinho")
            todos_alertas = []
            if carrinho:
                for item in carrinho:
                    _, alertas_item = aplicar_movimentacao(
                        conn, cliente["id"], item["produto_id"], numero_autorizado_id,
                        tipo, item["quantidade"], item.get("valor_unitario", 0),
                        origem="formulario", mensagem_original=texto,
                        cliente_negocio_id=dados.get("cliente_negocio_id")
                    )
                    todos_alertas.extend(alertas_item)
            else:
                # compatibilidade com o modo IA, que ainda manda um único item
                _, alertas_item = aplicar_movimentacao(
                    conn, cliente["id"], dados["produto_id"], numero_autorizado_id,
                    tipo, dados["quantidade"], dados.get("valor_unitario", 0),
                    origem="formulario", mensagem_original=texto,
                    cliente_negocio_id=dados.get("cliente_negocio_id")
                )
                todos_alertas.extend(alertas_item)
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            resposta = "✅ Registrado com sucesso!"
            if todos_alertas:
                resposta += "\n\n" + "\n".join(todos_alertas)
            return resposta + "\n\n" + resposta_menu(modulos)
        if texto_low in ("não", "nao", "n"):
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "Cancelado.\n\n" + resposta_menu(modulos)
        return "Responda SIM ou NÃO."

    # ── ETAPA: configurando horário(s) do resumo automático (opção 8) ──
    if etapa == "config_resumo_horarios":
        if texto_low in ("desativar", "desligar", "remover", "cancelar_config"):
            salvar_config_resumo_automatico(conn, cliente["id"], None, None, ativo=False)
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return "🔕 Resumo automático desativado.\n\n" + resposta_menu(modulos)

        horarios = parse_horarios(texto)
        if horarios is None:
            return (
                "Formato de horário inválido 🤔\n"
                "Digite até 2 horários no formato HH:MM separados por vírgula (ex: 12:00,20:00), "
                "ou *desativar* para desligar."
            )
        horario_1 = horarios[0]
        horario_2 = horarios[1] if len(horarios) > 1 else None
        salvar_config_resumo_automatico(conn, cliente["id"], horario_1, horario_2, ativo=True)
        salvar_sessao(conn, numero_autorizado_id, "menu", {})
        resumo_horarios = " e ".join(horarios)
        return (
            f"✅ Resumo automático configurado para {resumo_horarios} (todos os números autorizados recebem).\n\n"
            + resposta_menu(modulos)
        )

    # ── ETAPA: orçamento (opção 1) — escolhendo o(s) produto(s) ──
    if etapa == "orc_escolha_produtos":
        produtos_ids = dados.get("produtos_ids", [])
        indices = parse_selecao_multipla(texto, len(produtos_ids))
        if not indices:
            return (f"Manda o número do produto (1 a {len(produtos_ids)}). "
                     "Pra mais de um, separe por vírgula, ex: 1,3.")
        escolhidos_ids = [produtos_ids[i - 1] for i in indices]
        primeiro = db_one(conn, "SELECT * FROM produtos WHERE id = %s", (escolhidos_ids[0],))
        dados["fila_orc_produtos_ids"] = escolhidos_ids[1:]
        dados["orc_itens"] = []
        dados["produto_id"] = primeiro["id"]
        dados["produto_nome"] = primeiro["nome"]
        salvar_sessao(conn, numero_autorizado_id, "orc_qtd_item", dados)
        return f"Quantidade de *{primeiro['nome']}*?"

    # ── ETAPA: orçamento — pedindo quantidade de cada item ──
    if etapa == "orc_qtd_item":
        try:
            quantidade = float(texto.replace(",", "."))
        except ValueError:
            return "Manda só o número da quantidade, por favor."
        if quantidade <= 0:
            return "A quantidade precisa ser maior que zero."
        dados.setdefault("orc_itens", []).append({
            "produto_id": dados["produto_id"], "quantidade": quantidade,
        })
        fila = dados.get("fila_orc_produtos_ids", [])
        if fila:
            proximo_id = fila.pop(0)
            produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s", (proximo_id,))
            dados["fila_orc_produtos_ids"] = fila
            dados["produto_id"] = produto["id"]
            dados["produto_nome"] = produto["nome"]
            salvar_sessao(conn, numero_autorizado_id, "orc_qtd_item", dados)
            return f"Quantidade de *{produto['nome']}*?"
        salvar_sessao(conn, numero_autorizado_id, "orc_desconto", dados)
        return (
            "Quer ajustar o preço?\n"
            "Digite um valor pra dar *desconto* (ex: 10 ou 10%), "
            "um valor pra *aumentar* (ex: +10 ou +10%), "
            "ou *NÃO* pra manter igual."
        )

    # ── ETAPA: orçamento — desconto ──
    if etapa == "orc_desconto":
        try:
            _, subtotal_atual = calcular_itens_orcamento(conn, cliente["id"], dados.get("orc_itens", []))
        except ValueError as e:
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return f"❌ {e}\n\n" + resposta_menu(modulos)
        resultado_ajuste = parse_ajuste_preco_texto(texto, subtotal_atual)
        if resultado_ajuste is None:
            return (
                "Digite um valor pra *desconto* (ex: 10 ou 10%), "
                "pra *aumento* (ex: +10 ou +10%), ou *NÃO* pra manter igual."
            )
        ajuste_tipo, ajuste_valor_informado = resultado_ajuste
        dados["orc_desconto_tipo"] = ajuste_tipo
        dados["orc_desconto_valor_informado"] = ajuste_valor_informado
        salvar_sessao(conn, numero_autorizado_id, "orc_nome_cliente", dados)
        return "Nome do cliente pra colocar no orçamento? (digite PULAR se não quiser informar)"

    # ── ETAPA: orçamento — nome do cliente (última pergunta) e geração do texto final ──
    if etapa == "orc_nome_cliente":
        nome_cliente = None if texto_low == "pular" else texto.strip()
        cliente_negocio_id = None
        if nome_cliente:
            cliente_negocio = buscar_ou_criar_cliente_negocio(conn, cliente["id"], nome_cliente)
            if cliente_negocio:
                cliente_negocio_id = cliente_negocio["id"]
        try:
            itens_detalhados, subtotal = calcular_itens_orcamento(conn, cliente["id"], dados.get("orc_itens", []))
        except ValueError as e:
            salvar_sessao(conn, numero_autorizado_id, "menu", {})
            return f"❌ {e}\n\n" + resposta_menu(modulos)
        ajuste_tipo = dados.get("orc_desconto_tipo")
        ajuste_valor_informado = dados.get("orc_desconto_valor_informado", 0.0) or 0.0
        ajuste_calculado = calcular_ajuste_preco(subtotal, ajuste_tipo, ajuste_valor_informado)
        total = subtotal + ajuste_calculado
        texto_formatado = montar_texto_orcamento(
            cliente, itens_detalhados, subtotal, ajuste_tipo, ajuste_calculado,
            ajuste_valor_informado, total, observacoes=None, nome_cliente=nome_cliente
        )
        salvar_orcamento_db(conn, cliente["id"], nome_cliente, itens_detalhados, subtotal,
                             ajuste_tipo, ajuste_calculado, total, texto_formatado, None,
                             cliente_negocio_id=cliente_negocio_id)
        salvar_sessao(conn, numero_autorizado_id, "menu", {})
        return (
            "📋 Segue o orçamento! Confira os detalhes:\n\n"
            "━━━━━━━━━━━━━━━\n"
            f"{texto_formatado}\n"
            "━━━━━━━━━━━━━━━"
        )

    # fallback de segurança
    salvar_sessao(conn, numero_autorizado_id, "menu", {})
    return resposta_menu(modulos)

async def preparar_confirmacao_ia(conn, cliente, numero_autorizado_id, extraido, texto_original, modulos=None):
    tipo = extraido.get("tipo")
    nome_produto = extraido.get("produto")
    quantidade = extraido.get("quantidade")
    valor_unitario = extraido.get("valor_unitario") or 0

    if not nome_produto or quantidade is None:
        return "Entendi que é sobre estoque, mas faltou produto ou quantidade. Pode reescrever?\n\n" + resposta_menu(modulos)

    produto = buscar_produto_por_nome(conn, cliente["id"], nome_produto)
    if not produto:
        produto = db_exec(conn, "INSERT INTO produtos (cliente_id, nome) VALUES (%s,%s) RETURNING *",
                           (cliente["id"], nome_produto.strip()))

    dados = {
        "tipo": tipo, "produto_id": produto["id"], "produto_nome": produto["nome"],
        "quantidade": float(quantidade), "valor_unitario": float(valor_unitario or 0),
    }

    # TAREFA 2 — se a mensagem livre já mencionou o cliente numa venda, vincula sem perguntar de novo
    cliente_texto = extraido.get("cliente")
    if tipo == "venda" and cliente_texto:
        cliente_negocio = buscar_ou_criar_cliente_negocio(conn, cliente["id"], cliente_texto)
        if cliente_negocio:
            dados["cliente_negocio_id"] = cliente_negocio["id"]
            dados["cliente_negocio_nome"] = cliente_negocio["nome"]

    salvar_sessao(conn, numero_autorizado_id, "confirmando", dados)
    total = round(dados["quantidade"] * dados["valor_unitario"], 2)
    acao = {"entrada": "Entrada de", "venda": "Venda de", "saida": "Saída de"}[tipo]
    resposta = (
        f"Confirma?\n{acao} {fmt_num(dados['quantidade'])} × *{produto['nome']}* "
        f"a R$ {dados['valor_unitario']:.2f} (total R$ {total:.2f})\n\nResponda SIM ou NÃO."
    )
    if dados.get("cliente_negocio_nome"):
        resposta = f"Cliente: {dados['cliente_negocio_nome']}\n" + resposta
    return resposta

async def iniciar_cadastro_produto_ia(conn, cliente, numero_autorizado_id, extraido) -> Optional[str]:
    """TAREFA 1 — modo IA: dispara o cadastro de produto (com ou sem a calculadora)
    a partir de uma extração livre. Preenche o que conseguiu identificar e entra
    na etapa seguinte da MESMA máquina de estados usada pelo formulário, pedindo
    só o que ainda falta. Retorna None se não achou nome do produto (nesse caso o
    chamador simplesmente ignora e segue o fluxo normal)."""
    nome = extraido.get("nome")
    if not nome or not nome.strip():
        return None

    dados = {"prod_nome": nome.strip(), "prod_unidade": extraido.get("unidade") or "un"}

    custo_direto = extraido.get("custo_unitario")
    preco_direto = extraido.get("preco_venda")
    quer_calc = extraido.get("quer_calculadora")
    tem_dados_calc = any(
        extraido.get(k) is not None
        for k in ("custo_variavel_unitario", "custo_fixo_mensal", "volume_esperado_mensal", "margem_percentual")
    )

    if quer_calc or (custo_direto is None and preco_direto is None and tem_dados_calc):
        dados["custo_variavel"] = float(extraido.get("custo_variavel_unitario") or 0)
        dados["prod_quer_receita"] = False  # modo IA não monta ficha técnica na hora, só calcula
        if extraido.get("custo_fixo_mensal") is not None:
            dados["calc_custo_fixo_mensal"] = float(extraido["custo_fixo_mensal"])
        if extraido.get("volume_esperado_mensal") is not None:
            dados["calc_volume_esperado"] = float(extraido["volume_esperado_mensal"])
        dados["calc_margem"] = float(extraido["margem_percentual"]) if extraido.get("margem_percentual") is not None else 30.0

        if "calc_custo_fixo_mensal" not in dados:
            salvar_sessao(conn, numero_autorizado_id, "calc_custo_fixo", dados)
            return (
                f"Beleza, vamos calcular o preço de *{dados['prod_nome']}*!\n"
                "Qual o custo fixo mensal total do negócio (aluguel, luz, embalagem, mão de obra somados)?\n"
                "(esses valores não ficam salvos, é só pra calcular o preço agora)"
            )
        if "calc_volume_esperado" not in dados:
            salvar_sessao(conn, numero_autorizado_id, "calc_volume", dados)
            return f"Quantas unidades de *{dados['prod_nome']}* você espera vender por mês?"

        resultado = calcular_resultado_calculadora(dados)
        dados["calc_resultado"] = resultado
        salvar_sessao(conn, numero_autorizado_id, "calc_confirmar", dados)
        return texto_resultado_calculadora(dados, resultado)

    # modo direto — já tem (ou não) custo/preço; segue o cadastro pedindo só o que falta
    dados["prod_custo"] = float(custo_direto or 0)
    dados["prod_preco"] = float(preco_direto or 0)
    salvar_sessao(conn, numero_autorizado_id, "prod_estoque", dados)
    return (
        f"Beleza! Vou cadastrar *{dados['prod_nome']}* "
        f"(custo R$ {dados['prod_custo']:.2f}, venda R$ {dados['prod_preco']:.2f}).\n"
        "Qual o estoque inicial desse produto?"
    )

async def iniciar_edicao_produto_ia(conn, cliente, numero_autorizado_id, extraido, modulos=None) -> Optional[str]:
    """TAREFA 1 — modo IA: reconhece um pedido de edição de produto numa mensagem
    livre (ex: 'muda o preço do bolo pra 30 e o nome pra Bolo Premium') e monta
    o mesmo resumo de confirmação usado no fluxo do formulário (etapa
    'editar_campo_confirmar'), aplicando só os campos que a IA identificou.
    Retorna None se não achou o produto mencionado ou nenhuma alteração concreta
    — nesse caso o chamador segue o fluxo normal."""
    nome_produto = extraido.get("produto_editar")
    if not nome_produto or not nome_produto.strip():
        return None

    produto = buscar_produto_por_nome(conn, cliente["id"], nome_produto)
    if not produto:
        return (f"Não encontrei nenhum produto chamado '{nome_produto}'. "
                 "Confira o nome ou use a opção 10 do menu.\n\n") + resposta_menu(modulos)

    campos_novos = {}
    if extraido.get("nome_novo"):
        campos_novos["nome"] = str(extraido["nome_novo"]).strip()
    if extraido.get("custo_unitario_novo") is not None:
        try:
            campos_novos["custo_unitario"] = float(extraido["custo_unitario_novo"])
        except (TypeError, ValueError):
            pass
    if extraido.get("preco_venda_novo") is not None:
        try:
            campos_novos["preco_venda"] = float(extraido["preco_venda_novo"])
        except (TypeError, ValueError):
            pass
    if extraido.get("sku_novo"):
        campos_novos["sku"] = str(extraido["sku_novo"]).strip()

    if not campos_novos:
        return None  # a IA marcou "edicao_produto" mas não trouxe nenhum campo concreto pra mudar

    dados = {
        "editar_produto_id": produto["id"],
        "editar_produto_nome": produto["nome"],
        "editar_produto_atual": {
            "nome": produto["nome"],
            "custo_unitario": float(produto["custo_unitario"] or 0),
            "preco_venda": float(produto["preco_venda"] or 0),
            "sku": produto.get("sku"),
        },
        "editar_campos_novos": campos_novos,
    }
    salvar_sessao(conn, numero_autorizado_id, "editar_campo_confirmar", dados)
    return montar_resumo_edicao_produto(dados)

def iniciar_calculadora_custos_fixos_ia(extraido, modulos=None) -> Optional[str]:
    """TAREFA 3 — modo IA: reconhece uma lista de contas fixas numa mensagem livre
    (ex: 'aluguel 2000, luz 300, água 100') e soma tudo em Python — nunca confia
    na soma que a IA possa ter feito. Não usa/altera sessão: é resposta única.
    Retorna None se não veio nenhuma conta válida — o chamador segue o fluxo normal."""
    contas = extraido.get("contas_fixas")
    if not contas:
        return None
    itens = []
    for c in contas:
        nome = (c or {}).get("nome")
        valor = (c or {}).get("valor")
        if not nome or valor is None:
            continue
        try:
            itens.append({"nome": str(nome).strip(), "valor": float(valor)})
        except (TypeError, ValueError):
            continue
    if not itens:
        return None
    total = sum(item["valor"] for item in itens)  # soma sempre em Python, nunca confia na IA
    return montar_texto_resultado_custos_fixos(itens, total) + "\n\n" + resposta_menu(modulos)

def gerar_visao_geral(conn, cliente_id: int, modulos=None) -> str:
    """Lista todos os produtos e matérias-primas com o estoque atual,
    sinalizando com ⚠️ quem tem estoque mínimo definido e já ficou abaixo dele."""
    produtos = db_all(conn, """
        SELECT nome, unidade, estoque_atual FROM produtos
        WHERE cliente_id = %s AND ativo = TRUE ORDER BY nome
    """, (cliente_id,))
    materias = db_all(conn, """
        SELECT nome, unidade, estoque_atual, estoque_minimo FROM materias_primas
        WHERE cliente_id = %s AND ativo = TRUE ORDER BY nome
    """, (cliente_id,))

    blocos = ["📊 *Visão geral do estoque*"]

    blocos.append("\n📦 *Produtos*")
    if produtos:
        for p in produtos:
            blocos.append(f"- {p['nome']}: {fmt_num(p['estoque_atual'])} {p['unidade']}")
    else:
        blocos.append("Nenhum produto cadastrado ainda.")

    blocos.append("\n🧂 *Matérias-primas*")
    if materias:
        for m in materias:
            abaixo = m["estoque_minimo"] is not None and float(m["estoque_atual"]) < float(m["estoque_minimo"])
            marca = "⚠️ " if abaixo else "- "
            blocos.append(f"{marca}{m['nome']}: {fmt_num(m['estoque_atual'])} {m['unidade']}")
    else:
        blocos.append("Nenhuma matéria-prima cadastrada ainda.")

    return "\n".join(blocos) + "\n\n" + resposta_menu(modulos)

async def gerar_resumo_dia(conn, cliente_id: int, modulos=None) -> str:
    hoje = datetime.utcnow().date()
    linhas = db_all(conn, """
        SELECT tipo, COALESCE(SUM(quantidade),0) qtd, COALESCE(SUM(valor_total),0) total
        FROM movimentacoes
        WHERE cliente_id = %s AND criado_em::date = %s
        GROUP BY tipo
    """, (cliente_id, hoje))
    if not linhas:
        return f"📊 Nenhuma movimentação hoje ({hoje.strftime('%d/%m')}).\n\n" + resposta_menu(modulos)
    txt = f"📊 *Resumo de hoje ({hoje.strftime('%d/%m')})*\n"
    for l in linhas:
        txt += f"- {l['tipo'].capitalize()}: {fmt_num(l['qtd'])} un | R$ {float(l['total']):.2f}\n"
    return txt + "\n" + resposta_menu(modulos)

# ─────────────────────────────────────────
#  RESUMO AUTOMÁTICO — configuração e disparo agendado
# ─────────────────────────────────────────
def parse_horarios(texto: str):
    """Aceita até 2 horários no formato HH:MM separados por vírgula/espaço
    (ex: '12:00,20:00' ou '12:00 20:00'). Retorna lista de strings 'HH:MM'
    normalizadas, ou None se algo for inválido."""
    partes = [p.strip() for p in re.split(r"[,\s]+", texto.strip()) if p.strip()]
    if not partes or len(partes) > 2:
        return None
    horarios = []
    for p in partes:
        m = re.fullmatch(r"(\d{1,2}):(\d{2})", p)
        if not m:
            return None
        h, mi = int(m.group(1)), int(m.group(2))
        if h > 23 or mi > 59:
            return None
        hhmm = f"{h:02d}:{mi:02d}"
        if hhmm not in horarios:
            horarios.append(hhmm)
    return horarios or None

def obter_config_resumo_automatico(conn, cliente_id: int):
    return db_one(conn, "SELECT * FROM resumo_automatico_config WHERE cliente_id = %s", (cliente_id,))

def formatar_horarios_config(config) -> str:
    if not config or not config.get("ativo") or not (config.get("horario_1") or config.get("horario_2")):
        return "nenhum configurado"
    horarios = [str(h)[:5] for h in (config.get("horario_1"), config.get("horario_2")) if h]
    return " e ".join(horarios)

def salvar_config_resumo_automatico(conn, cliente_id: int, horario_1, horario_2, ativo: bool):
    db_exec(conn, """
        INSERT INTO resumo_automatico_config (cliente_id, horario_1, horario_2, ativo, atualizado_em)
        VALUES (%s, %s, %s, %s, NOW())
        ON CONFLICT (cliente_id) DO UPDATE
        SET horario_1 = EXCLUDED.horario_1,
            horario_2 = EXCLUDED.horario_2,
            ativo = EXCLUDED.ativo,
            atualizado_em = NOW()
    """, (cliente_id, horario_1, horario_2, ativo))

async def disparar_resumo_automatico_cliente(conn, cliente_id: int):
    resumo = await gerar_resumo_dia(conn, cliente_id)
    numeros = db_all(conn, "SELECT numero FROM numeros_autorizados WHERE cliente_id = %s AND ativo = TRUE",
                      (cliente_id,))
    for n in numeros:
        await enviar_whatsapp(n["numero"], "⏰ *Resumo automático*\n\n" + resumo)

async def checar_e_disparar_resumos_automaticos():
    """Roda 1x por minuto. Compara o horário atual (America/Sao_Paulo) com os
    horários configurados por cada cliente e dispara o resumo do dia quando
    bate, evitando reenvio duplicado no mesmo dia via ultimo_envio_1/2."""
    agora = datetime.now(TIMEZONE_PADRAO)
    hoje = agora.date()
    hora_atual = agora.strftime("%H:%M")

    conn = get_conn_raw()
    try:
        configs = db_all(conn, "SELECT * FROM resumo_automatico_config WHERE ativo = TRUE")
        for c in configs:
            for campo_horario, campo_envio in (("horario_1", "ultimo_envio_1"), ("horario_2", "ultimo_envio_2")):
                horario = c.get(campo_horario)
                if not horario:
                    continue
                if str(horario)[:5] != hora_atual:
                    continue
                if c.get(campo_envio) == hoje:
                    continue  # já disparou esse horário hoje
                try:
                    await disparar_resumo_automatico_cliente(conn, c["cliente_id"])
                except Exception as e:
                    print(f"⚠️ Erro ao disparar resumo automático (cliente {c['cliente_id']}): {e}")
                db_exec(conn, f"UPDATE resumo_automatico_config SET {campo_envio} = %s WHERE cliente_id = %s",
                        (hoje, c["cliente_id"]))
    finally:
        conn.close()

async def loop_relogio_resumo_automatico():
    while True:
        try:
            await checar_e_disparar_resumos_automaticos()
        except Exception as e:
            print(f"⚠️ Erro no relógio de resumo automático: {e}")
        await asyncio.sleep(60)

# ─────────────────────────────────────────
#  AGENDA — lembrete automático por WhatsApp
# ─────────────────────────────────────────
async def checar_e_disparar_lembretes_agenda():
    """Roda 1x por minuto. Olha os compromissos de hoje com lembrete configurado
    e ainda não enviado (lembrete_enviado = FALSE), calcula se já está na janela
    de disparo (hora do compromisso menos os minutos de antecedência) e dispara
    o WhatsApp, marcando lembrete_enviado = TRUE pra não reenviar."""
    agora = datetime.now(TIMEZONE_PADRAO)
    hoje = agora.date()
    agora_min = agora.hour * 60 + agora.minute

    conn = get_conn_raw()
    try:
        compromissos = db_all(conn, """
            SELECT a.id, a.cliente_id, a.titulo, a.hora_inicio, a.notas, a.lembrete_minutos_antes
            FROM agenda_compromissos a
            WHERE a.data = %s AND a.status = 'agendado'
              AND a.lembrete_minutos_antes IS NOT NULL AND a.lembrete_enviado = FALSE
        """, (hoje,))
        for c in compromissos:
            h, mi = map(int, str(c["hora_inicio"])[:5].split(":"))
            compromisso_min = h * 60 + mi
            disparo_min = compromisso_min - int(c["lembrete_minutos_antes"])
            if agora_min < disparo_min:
                continue  # ainda não chegou a hora do lembrete
            if agora_min > disparo_min + 60:
                # passou muito da janela (ex: sistema ficou fora do ar) — marca como
                # enviado sem disparar, pra não mandar um lembrete extemporâneo horas depois
                db_exec(conn, "UPDATE agenda_compromissos SET lembrete_enviado = TRUE WHERE id = %s", (c["id"],))
                continue
            try:
                numeros = db_all(conn, "SELECT numero FROM numeros_autorizados WHERE cliente_id = %s AND ativo = TRUE",
                                  (c["cliente_id"],))
                texto = f"⏰ *Lembrete de compromisso*\n\n📌 {c['titulo']}\n🕐 Hoje às {str(c['hora_inicio'])[:5]}"
                if c.get("notas"):
                    texto += f"\n📝 {c['notas']}"
                for n in numeros:
                    await enviar_whatsapp(n["numero"], texto)
                db_exec(conn, "UPDATE agenda_compromissos SET lembrete_enviado = TRUE WHERE id = %s", (c["id"],))
            except Exception as e:
                print(f"⚠️ Erro ao disparar lembrete da agenda (compromisso {c['id']}): {e}")
    finally:
        conn.close()

async def loop_relogio_lembretes_agenda():
    while True:
        try:
            await checar_e_disparar_lembretes_agenda()
        except Exception as e:
            print(f"⚠️ Erro no relógio de lembretes da agenda: {e}")
        await asyncio.sleep(60)

# ─────────────────────────────────────────
#  STARTUP — Criar tabelas se não existirem
# ─────────────────────────────────────────
def criar_tabelas_resumo_automatico():
    """Cria a tabela resumo_automatico_config no startup SE ela não existir.
    Seguro: usa IF NOT EXISTS, nunca deleta ou altera dados existentes."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS resumo_automatico_config (
                cliente_id      INTEGER PRIMARY KEY REFERENCES clientes(id) ON DELETE CASCADE,
                horario_1       TIME NULL,
                horario_2       TIME NULL,
                ativo           BOOLEAN NOT NULL DEFAULT TRUE,
                ultimo_envio_1  DATE NULL,
                ultimo_envio_2  DATE NULL,
                atualizado_em   TIMESTAMP NOT NULL DEFAULT NOW()
            );
        """)
        conn.commit()
        print("✅ Tabela resumo_automatico_config: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela: {e}")
    finally:
        conn.close()

def criar_tabelas_agendamentos():
    """Cria a tabela agendamentos no startup SE ela não existir."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS agendamentos (
                id              SERIAL PRIMARY KEY,
                cliente_id      INTEGER NOT NULL REFERENCES clientes(id) ON DELETE CASCADE,
                produto_id      INTEGER NOT NULL REFERENCES produtos(id) ON DELETE CASCADE,
                data_agendamento TIMESTAMP NOT NULL,
                quantidade      FLOAT NOT NULL,
                notas           TEXT,
                criado_em       TIMESTAMP NOT NULL DEFAULT NOW()
            );
        """)
        conn.commit()
        print("✅ Tabela agendamentos: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela agendamentos: {e}")
    finally:
        conn.close()

def criar_tabelas_orcamentos():
    """Cria a tabela orcamentos no startup SE ela não existir."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS orcamentos (
                id              SERIAL PRIMARY KEY,
                cliente_id      INTEGER NOT NULL REFERENCES clientes(id) ON DELETE CASCADE,
                nome_cliente    TEXT,
                itens           JSONB NOT NULL,
                subtotal        NUMERIC NOT NULL,
                desconto_tipo   TEXT NULL,
                desconto_valor  NUMERIC NULL,
                total           NUMERIC NOT NULL,
                texto_formatado TEXT NOT NULL,
                observacoes     TEXT,
                criado_em       TIMESTAMP NOT NULL DEFAULT NOW()
            );
        """)
        conn.commit()
        print("✅ Tabela orcamentos: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela orcamentos: {e}")
    finally:
        conn.close()

def criar_tabela_clientes_negocio():
    """Cria a tabela clientes_negocio (TAREFA 2) e as colunas opcionais
    cliente_negocio_id em orcamentos/movimentacoes, SE ainda não existirem.
    Mesmo padrão 'self-healing' das outras tabelas novas — o migration.sql
    também cobre isso pra quem preferir rodar a migração manualmente."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS clientes_negocio (
                id          SERIAL PRIMARY KEY,
                cliente_id  INTEGER NOT NULL REFERENCES clientes(id) ON DELETE CASCADE,
                nome        TEXT NOT NULL,
                telefone    TEXT,
                criado_em   TIMESTAMP NOT NULL DEFAULT NOW()
            );
        """)
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS ux_clientes_negocio_cliente_telefone
            ON clientes_negocio (cliente_id, telefone) WHERE telefone IS NOT NULL;
        """)
        cur.execute("ALTER TABLE orcamentos ADD COLUMN IF NOT EXISTS cliente_negocio_id INTEGER NULL REFERENCES clientes_negocio(id);")
        cur.execute("ALTER TABLE movimentacoes ADD COLUMN IF NOT EXISTS cliente_negocio_id INTEGER NULL REFERENCES clientes_negocio(id);")
        conn.commit()
        print("✅ Tabela clientes_negocio e colunas relacionadas: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela clientes_negocio: {e}")
    finally:
        conn.close()

def criar_tabela_custos_fixos():
    """Cria a tabela custos_fixos (TAREFA 5) no startup SE ainda não existir.
    Mesmo padrão 'self-healing' das outras tabelas novas: usa IF NOT EXISTS,
    nunca deleta ou altera dados existentes."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS custos_fixos (
                id             SERIAL PRIMARY KEY,
                cliente_id     INTEGER NOT NULL REFERENCES clientes(id) ON DELETE CASCADE,
                nome_da_conta  TEXT NOT NULL,
                valor          NUMERIC NOT NULL DEFAULT 0,
                atualizado_em  TIMESTAMP NOT NULL DEFAULT NOW()
            );
        """)
        conn.commit()
        print("✅ Tabela custos_fixos: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela custos_fixos: {e}")
    finally:
        conn.close()

def criar_tabela_agenda_compromissos():
    """Cria a tabela agenda_compromissos (agenda genérica de compromissos, não
    confundir com `agendamentos`, que é reposição de estoque) no startup SE
    ainda não existir. Mesmo padrão 'self-healing' das outras tabelas novas."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS agenda_compromissos (
                id                      SERIAL PRIMARY KEY,
                cliente_id              INTEGER NOT NULL REFERENCES clientes(id) ON DELETE CASCADE,
                cliente_negocio_id      INTEGER NULL REFERENCES clientes_negocio(id) ON DELETE SET NULL,
                titulo                  TEXT NOT NULL,
                data                    DATE NOT NULL,
                hora_inicio             TIME NOT NULL,
                hora_fim                TIME NULL,
                status                  TEXT NOT NULL DEFAULT 'agendado',
                notas                   TEXT,
                lembrete_minutos_antes  INTEGER NULL,
                lembrete_enviado        BOOLEAN NOT NULL DEFAULT FALSE,
                criado_em               TIMESTAMP NOT NULL DEFAULT NOW()
            );
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS ix_agenda_compromissos_cliente_data
            ON agenda_compromissos (cliente_id, data);
        """)
        conn.commit()
        print("✅ Tabela agenda_compromissos: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela agenda_compromissos: {e}")
    finally:
        conn.close()

def criar_coluna_modulos_clientes():
    """Adiciona a coluna `modulos` (array de texto) na tabela clientes, SE ainda
    não existir. Guarda quais módulos (estoque/agenda) aquele cliente tem ativo.
    Nullable com default ['estoque'] pra não quebrar clientes que já existem hoje.
    Mesmo padrão 'self-healing' das outras migrações — ALTER TABLE ADD COLUMN IF NOT EXISTS."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            ALTER TABLE clientes
            ADD COLUMN IF NOT EXISTS modulos TEXT[] DEFAULT ARRAY['estoque'];
        """)
        cur.execute("""
            UPDATE clientes SET modulos = ARRAY['estoque'] WHERE modulos IS NULL;
        """)
        conn.commit()
        print("✅ Coluna clientes.modulos: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar coluna clientes.modulos: {e}")
    finally:
        conn.close()

def criar_coluna_movimentacao_id_materia_prima():
    """Adiciona a coluna `movimentacao_id` (FK -> movimentacoes.id) na tabela
    movimentacoes_materia_prima, SE ainda não existir. Necessária desde que
    baixar_materia_prima_por_receita() passou a gravar de qual venda (movimentacao)
    cada baixa de matéria-prima se originou, pra permitir reverter com precisão
    (TAREFA 2 — dashboard). Mesmo padrão 'self-healing' das outras migrações —
    ALTER TABLE ADD COLUMN IF NOT EXISTS, nunca deleta ou altera dados existentes."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            ALTER TABLE movimentacoes_materia_prima
            ADD COLUMN IF NOT EXISTS movimentacao_id INTEGER NULL REFERENCES movimentacoes(id) ON DELETE SET NULL;
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS ix_movimentacoes_materia_prima_movimentacao_id
            ON movimentacoes_materia_prima (movimentacao_id);
        """)
        conn.commit()
        print("✅ Coluna movimentacoes_materia_prima.movimentacao_id: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar coluna movimentacoes_materia_prima.movimentacao_id: {e}")
    finally:
        conn.close()

def criar_coluna_atendimento_cliente_final():
    """Adiciona a coluna `atendimento_cliente_final_ativado` na tabela clientes,
    SE ainda não existir. Controla o toggle do dashboard que liga/desliga o
    atendimento automático pro cliente final (visitante) no mesmo número
    WhatsApp. Default FALSE pra não mudar comportamento de ninguém que já
    está rodando. Mesmo padrão 'self-healing' das outras migrações."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            ALTER TABLE clientes
            ADD COLUMN IF NOT EXISTS atendimento_cliente_final_ativado BOOLEAN NOT NULL DEFAULT FALSE;
        """)
        conn.commit()
        print("✅ Coluna clientes.atendimento_cliente_final_ativado: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar coluna clientes.atendimento_cliente_final_ativado: {e}")
    finally:
        conn.close()

def criar_tabela_sessoes_cliente_final():
    """Cria a tabela sessoes_cliente_final — sessão de conversa separada da
    sessoes_conversa (que é só pra número autorizado/funcionário). Guarda em
    que etapa do menu simplificado (ver produtos, orçamento, falar com
    atendente) cada visitante está, por empresa. Mesmo padrão 'self-healing'."""
    conn = get_conn_raw()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS sessoes_cliente_final (
                id               SERIAL PRIMARY KEY,
                cliente_id       INTEGER NOT NULL REFERENCES clientes(id) ON DELETE CASCADE,
                numero           TEXT NOT NULL,
                etapa_atual      TEXT NOT NULL DEFAULT 'menu_cliente_final',
                dados_parciais   JSONB NOT NULL DEFAULT '{}',
                criado_em        TIMESTAMP NOT NULL DEFAULT NOW(),
                atualizado_em    TIMESTAMP NOT NULL DEFAULT NOW(),
                UNIQUE (cliente_id, numero)
            );
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS ix_sessoes_cliente_final_cliente_numero
            ON sessoes_cliente_final (cliente_id, numero);
        """)
        conn.commit()
        print("✅ Tabela sessoes_cliente_final: OK")
    except Exception as e:
        print(f"⚠️ Erro ao criar tabela sessoes_cliente_final: {e}")
    finally:
        conn.close()

# ─────────────────────────────────────────
#  LIFESPAN
# ─────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    # 1️⃣ Criar tabelas necessárias
    criar_tabelas_resumo_automatico()
    criar_tabelas_agendamentos()
    criar_tabelas_orcamentos()
    criar_tabela_clientes_negocio()
    criar_tabela_custos_fixos()
    criar_tabela_agenda_compromissos()
    criar_coluna_modulos_clientes()
    criar_coluna_movimentacao_id_materia_prima()
    criar_coluna_atendimento_cliente_final()
    criar_tabela_sessoes_cliente_final()
    
    # 2️⃣ Testar conexão com banco
    try:
        conn = get_conn_raw()
        conn.close()
        print("✅ Conexão com banco OK")
    except Exception as e:
        print(f"⚠️ Não foi possível conectar ao banco no startup: {e}")

    # 3️⃣ Iniciar os "relógios" em background
    tarefa_relogio = asyncio.create_task(loop_relogio_resumo_automatico())
    tarefa_relogio_agenda = asyncio.create_task(loop_relogio_lembretes_agenda())
    yield
    
    # 4️⃣ Cleanup: cancelar os relógios quando o app fecha
    tarefa_relogio.cancel()
    tarefa_relogio_agenda.cancel()
    for tarefa in (tarefa_relogio, tarefa_relogio_agenda):
        try:
            await tarefa
        except asyncio.CancelledError:
            pass

app = FastAPI(title="Estoque WPP", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

# ═════════════════════════════════════════
#  AUTH — CLIENTE
# ═════════════════════════════════════════
@app.post("/auth/login")
def login(body: LoginBody, conn=Depends(get_db)):
    cliente = db_one(conn, "SELECT * FROM clientes WHERE email = %s AND ativo = TRUE", (body.email,))
    if not cliente or not bcrypt.checkpw(body.senha.encode(), cliente["senha_hash"].encode()):
        raise HTTPException(status_code=401, detail="Email ou senha inválidos")
    token = criar_token(cliente["id"], cliente["email"])
    return {
        "access_token": token,
        "cliente_id": cliente["id"],
        "cliente": {k: v for k, v in cliente.items() if k != "senha_hash"},
    }

@app.get("/auth/me")
def me(cliente=Depends(get_current_cliente)):
    return {k: v for k, v in cliente.items() if k != "senha_hash"}

@app.patch("/cliente/atendimento-cliente-final")
def alterar_atendimento_cliente_final(body: AtendimentoClienteFinalBody,
                                       cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Liga/desliga o atendimento automático simplificado (menu de produtos,
    orçamento e 'falar com atendente') pro cliente final da empresa, no mesmo
    número WhatsApp. Só afeta números que já estão cadastrados em
    clientes-negocio dessa empresa — números aleatórios continuam bloqueados."""
    db_exec(conn, "UPDATE clientes SET atendimento_cliente_final_ativado = %s WHERE id = %s",
            (body.ativado, cliente["id"]))
    return {"ok": True, "atendimento_cliente_final_ativado": body.ativado}

# ═════════════════════════════════════════
#  ADMIN — gestão de clientes/números/conexão
# ═════════════════════════════════════════
@app.get("/admin/clientes")
def admin_listar_clientes(conn=Depends(get_db), _admin=Depends(check_admin)):
    clientes = db_all(conn, "SELECT id, nome_negocio, email, plano, ativo, modulos, criado_em FROM clientes ORDER BY id DESC")
    for c in clientes:
        c["numeros"] = db_all(conn, "SELECT * FROM numeros_autorizados WHERE cliente_id = %s", (c["id"],))
    return clientes

@app.post("/admin/clientes")
def admin_criar_cliente(body: AdminCriarClienteBody, conn=Depends(get_db), _admin=Depends(check_admin)):
    if db_one(conn, "SELECT id FROM clientes WHERE email = %s", (body.email,)):
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    senha_hash = bcrypt.hashpw(body.senha.encode(), bcrypt.gensalt()).decode()
    modulos = body.modulos if body.modulos else ["estoque"]
    cliente = db_exec(conn, """
        INSERT INTO clientes (nome_negocio, email, senha_hash, plano, modulos)
        VALUES (%s,%s,%s,%s,%s) RETURNING id, nome_negocio, email, plano, ativo, modulos
    """, (body.nome_negocio, body.email, senha_hash, body.plano, modulos))
    return cliente

@app.patch("/admin/clientes/{cliente_id}/plano")
def admin_trocar_plano(cliente_id: int, body: AdminPlanoBody, conn=Depends(get_db), _admin=Depends(check_admin)):
    db_exec(conn, "UPDATE clientes SET plano = %s WHERE id = %s", (body.plano, cliente_id))
    return {"ok": True}

@app.patch("/admin/clientes/{cliente_id}/modulos")
def admin_trocar_modulos(cliente_id: int, body: AdminModulosBody, conn=Depends(get_db), _admin=Depends(check_admin)):
    modulos_validos = [m for m in body.modulos if m in ("estoque", "agenda")]
    if not modulos_validos:
        raise HTTPException(status_code=400, detail="Informe ao menos um módulo válido (estoque, agenda)")
    db_exec(conn, "UPDATE clientes SET modulos = %s WHERE id = %s", (modulos_validos, cliente_id))
    return {"ok": True}

@app.patch("/admin/clientes/{cliente_id}/ativo")
def admin_toggle_ativo(cliente_id: int, body: AdminAtivoBody, conn=Depends(get_db), _admin=Depends(check_admin)):
    db_exec(conn, "UPDATE clientes SET ativo = %s WHERE id = %s", (body.ativo, cliente_id))
    return {"ok": True}

@app.delete("/admin/clientes/{cliente_id}")
def admin_deletar_cliente(cliente_id: int, conn=Depends(get_db), _admin=Depends(check_admin)):
    db_exec(conn, "DELETE FROM clientes WHERE id = %s", (cliente_id,))
    return {"ok": True}

@app.post("/admin/clientes/{cliente_id}/numeros")
def admin_add_numero(cliente_id: int, body: AdminNumeroBody, conn=Depends(get_db), _admin=Depends(check_admin)):
    numero = normalizar_numero(body.numero)
    if db_one(conn, "SELECT id FROM numeros_autorizados WHERE numero = %s", (numero,)):
        raise HTTPException(status_code=400, detail="Número já cadastrado")
    reg = db_exec(conn, """
        INSERT INTO numeros_autorizados (cliente_id, numero, nome) VALUES (%s,%s,%s) RETURNING *
    """, (cliente_id, numero, body.nome))
    return reg

@app.delete("/admin/numeros/{numero_id}")
def admin_del_numero(numero_id: int, conn=Depends(get_db), _admin=Depends(check_admin)):
    db_exec(conn, "DELETE FROM numeros_autorizados WHERE id = %s", (numero_id,))
    return {"ok": True}

@app.get("/admin/conexao")
def admin_get_conexao(conn=Depends(get_db), _admin=Depends(check_admin)):
    conexao = db_one(conn, "SELECT * FROM conexao_bot WHERE id = 1")
    try:
        async def _status():
            async with httpx.AsyncClient(timeout=10) as client:
                r = await client.get(f"{BAILEYS_URL}/status")
                return r.json()
    except Exception:
        pass
    return conexao

@app.post("/admin/conexao/resetar")
async def admin_resetar_conexao(conn=Depends(get_db), _admin=Depends(check_admin)):
    async with httpx.AsyncClient(timeout=15) as client:
        try:
            await client.post(f"{BAILEYS_URL}/resetar")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Falha ao resetar sessão no Baileys: {e}")
    db_exec(conn, "UPDATE conexao_bot SET status = 'trocando', atualizado_em = NOW() WHERE id = 1")
    return {"ok": True, "mensagem": "Sessão resetada. Acesse /whatsapp/qrcode para escanear o novo QR."}

@app.post("/admin/conexao/notificar-clientes")
async def admin_notificar_clientes(conn=Depends(get_db), _admin=Depends(check_admin)):
    conexao = db_one(conn, "SELECT * FROM conexao_bot WHERE id = 1")
    numeros = db_all(conn, "SELECT numero FROM numeros_autorizados WHERE ativo = TRUE")
    texto = (
        "📢 Aviso: este é o número oficial atualizado do sistema de estoque. "
        "Salve este contato para continuar registrando suas movimentações por aqui."
    )
    enviados = 0
    for n in numeros:
        await enviar_whatsapp(n["numero"], texto)
        enviados += 1
    return {"ok": True, "notificados": enviados}

@app.get("/admin/log-trocas")
def admin_log_trocas(conn=Depends(get_db), _admin=Depends(check_admin)):
    return db_all(conn, "SELECT * FROM log_troca_numero ORDER BY criado_em DESC")

# Chamado internamente pelo Baileys quando conecta com um número (ver webhook abaixo)
class ConexaoAtualizaBody(BaseModel):
    numero: str

@app.post("/admin/conexao/atualizar")
def conexao_atualizar(body: ConexaoAtualizaBody, conn=Depends(get_db)):
    numero = normalizar_numero(body.numero)
    atual = db_one(conn, "SELECT * FROM conexao_bot WHERE id = 1")
    if atual and atual.get("numero_atual") and atual["numero_atual"] != numero:
        db_exec(conn, "INSERT INTO log_troca_numero (numero_antigo, numero_novo) VALUES (%s,%s)",
                (atual["numero_atual"], numero))
    db_exec(conn, "UPDATE conexao_bot SET numero_atual = %s, status = 'conectado', atualizado_em = NOW() WHERE id = 1",
            (numero,))
    return {"ok": True}

# ═════════════════════════════════════════
#  PRODUTOS (painel do cliente)
# ═════════════════════════════════════════
@app.get("/produtos")
def listar_produtos(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    return db_all(conn, "SELECT * FROM produtos WHERE cliente_id = %s AND ativo = TRUE ORDER BY nome",
                  (cliente["id"],))

@app.post("/produtos")
def criar_produto(body: ProdutoBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    return db_exec(conn, """
        INSERT INTO produtos (cliente_id, nome, sku, custo_unitario, preco_venda, estoque_atual, unidade)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
    """, (cliente["id"], body.nome, body.sku, body.custo_unitario, body.preco_venda, body.estoque_atual, body.unidade))

@app.put("/produtos/{produto_id}")
def atualizar_produto(produto_id: int, body: ProdutoBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    db_exec(conn, """
        UPDATE produtos SET nome=%s, sku=%s, custo_unitario=%s, preco_venda=%s, estoque_atual=%s, unidade=%s
        WHERE id=%s AND cliente_id=%s
    """, (body.nome, body.sku, body.custo_unitario, body.preco_venda, body.estoque_atual, body.unidade,
          produto_id, cliente["id"]))
    return {"ok": True}

@app.delete("/produtos/{produto_id}")
def deletar_produto(produto_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    db_exec(conn, "UPDATE produtos SET ativo = FALSE WHERE id = %s AND cliente_id = %s", (produto_id, cliente["id"]))
    return {"ok": True}

@app.patch("/produtos/{produto_id}/valores")
def editar_valores_produto(produto_id: int, body: ProdutoEditarBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Endpoint rápido pra editar só custo/preço (inline editing)"""
    produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s", (produto_id, cliente["id"]))
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    
    if body.custo_unitario is not None:
        db_exec(conn, "UPDATE produtos SET custo_unitario=%s WHERE id=%s", (body.custo_unitario, produto_id))
    if body.preco_venda is not None:
        db_exec(conn, "UPDATE produtos SET preco_venda=%s WHERE id=%s", (body.preco_venda, produto_id))
    
    return {"ok": True}

# ═════════════════════════════════════════
#  MATÉRIAS-PRIMAS (painel do cliente)
# ═════════════════════════════════════════
@app.get("/materias-primas")
def listar_materias_primas(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    return db_all(conn, "SELECT * FROM materias_primas WHERE cliente_id = %s AND ativo = TRUE ORDER BY nome",
                  (cliente["id"],))

@app.post("/materias-primas")
def criar_materia_prima(body: MateriaPrimaBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    return db_exec(conn, """
        INSERT INTO materias_primas (cliente_id, nome, sku, custo_unitario, estoque_atual, unidade, estoque_minimo)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *
    """, (cliente["id"], body.nome, body.sku, body.custo_unitario, body.estoque_atual, body.unidade, body.estoque_minimo))

@app.put("/materias-primas/{materia_id}")
def atualizar_materia_prima(materia_id: int, body: MateriaPrimaBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    db_exec(conn, """
        UPDATE materias_primas SET nome=%s, sku=%s, custo_unitario=%s, estoque_atual=%s, unidade=%s, estoque_minimo=%s
        WHERE id=%s AND cliente_id=%s
    """, (body.nome, body.sku, body.custo_unitario, body.estoque_atual, body.unidade, body.estoque_minimo,
          materia_id, cliente["id"]))
    return {"ok": True}

@app.delete("/materias-primas/{materia_id}")
def deletar_materia_prima(materia_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    db_exec(conn, "UPDATE materias_primas SET ativo = FALSE WHERE id = %s AND cliente_id = %s",
            (materia_id, cliente["id"]))
    return {"ok": True}

@app.post("/materias-primas/movimentacao")
def criar_movimentacao_materia_prima_manual(body: MovimentacaoMateriaPrimaBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    materia = db_one(conn, "SELECT * FROM materias_primas WHERE id = %s AND cliente_id = %s",
                      (body.materia_prima_id, cliente["id"]))
    if not materia:
        raise HTTPException(status_code=404, detail="Matéria-prima não encontrada")
    _, alerta = aplicar_movimentacao_materia_prima(conn, cliente["id"], body.materia_prima_id, None, body.tipo,
                                                    body.quantidade, body.valor_unitario, origem="manual_painel")
    return {"ok": True, "alerta": alerta}

@app.get("/movimentacoes-materia-prima")
def listar_movimentacoes_materia_prima(
    materia_prima_id: Optional[int] = None, tipo: Optional[str] = None,
    de: Optional[str] = None, ate: Optional[str] = None,
    cliente=Depends(get_current_cliente), conn=Depends(get_db)
):
    sql = """
        SELECT mv.*, m.nome AS materia_prima_nome, n.nome AS registrado_por, p.nome AS produto_origem_nome
        FROM movimentacoes_materia_prima mv
        JOIN materias_primas m ON m.id = mv.materia_prima_id
        LEFT JOIN numeros_autorizados n ON n.id = mv.numero_autorizado_id
        LEFT JOIN produtos p ON p.id = mv.produto_id_origem
        WHERE mv.cliente_id = %s
    """
    params = [cliente["id"]]
    if materia_prima_id:
        sql += " AND mv.materia_prima_id = %s"; params.append(materia_prima_id)
    if tipo:
        sql += " AND mv.tipo = %s"; params.append(tipo)
    if de:
        sql += " AND mv.criado_em::date >= %s"; params.append(de)
    if ate:
        sql += " AND mv.criado_em::date <= %s"; params.append(ate)
    sql += " ORDER BY mv.criado_em DESC LIMIT 500"
    return db_all(conn, sql, tuple(params))

# ═════════════════════════════════════════
#  RECEITA / FICHA TÉCNICA (produto ⇄ matérias-primas)
# ═════════════════════════════════════════
@app.get("/produtos/{produto_id}/receita")
def obter_receita(produto_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s", (produto_id, cliente["id"]))
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    return db_all(conn, """
        SELECT r.id, r.materia_prima_id, r.quantidade_necessaria,
               m.nome AS materia_prima_nome, m.unidade, m.estoque_atual, m.custo_unitario
        FROM receita_itens r
        JOIN materias_primas m ON m.id = r.materia_prima_id
        WHERE r.produto_id = %s
        ORDER BY m.nome
    """, (produto_id,))

@app.put("/produtos/{produto_id}/receita")
def salvar_receita(produto_id: int, body: ReceitaBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s", (produto_id, cliente["id"]))
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    db_exec(conn, "DELETE FROM receita_itens WHERE produto_id = %s", (produto_id,))
    for item in body.itens:
        materia = db_one(conn, "SELECT id FROM materias_primas WHERE id = %s AND cliente_id = %s",
                          (item.materia_prima_id, cliente["id"]))
        if not materia:
            raise HTTPException(status_code=400, detail=f"Matéria-prima {item.materia_prima_id} não encontrada")
        db_exec(conn, """
            INSERT INTO receita_itens (produto_id, materia_prima_id, quantidade_necessaria)
            VALUES (%s,%s,%s)
        """, (produto_id, item.materia_prima_id, item.quantidade_necessaria))
    return {"ok": True}

# ═════════════════════════════════════════
#  MOVIMENTAÇÕES / DASHBOARD
# ═════════════════════════════════════════
@app.get("/movimentacoes")
def listar_movimentacoes(
    produto_id: Optional[int] = None, tipo: Optional[str] = None,
    de: Optional[str] = None, ate: Optional[str] = None,
    cliente=Depends(get_current_cliente), conn=Depends(get_db)
):
    sql = """
        SELECT m.*, p.nome AS produto_nome, n.nome AS registrado_por, cn.nome AS cliente_negocio_nome
        FROM movimentacoes m
        JOIN produtos p ON p.id = m.produto_id
        LEFT JOIN numeros_autorizados n ON n.id = m.numero_autorizado_id
        LEFT JOIN clientes_negocio cn ON cn.id = m.cliente_negocio_id
        WHERE m.cliente_id = %s
    """
    params = [cliente["id"]]
    if produto_id:
        sql += " AND m.produto_id = %s"; params.append(produto_id)
    if tipo:
        sql += " AND m.tipo = %s"; params.append(tipo)
    if de:
        sql += " AND m.criado_em::date >= %s"; params.append(de)
    if ate:
        sql += " AND m.criado_em::date <= %s"; params.append(ate)
    sql += " ORDER BY m.criado_em DESC LIMIT 500"
    return db_all(conn, sql, tuple(params))

@app.delete("/movimentacoes/{movimentacao_id}")
def deletar_movimentacao(movimentacao_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """TAREFA 2 — exclui uma venda registrada revertendo o efeito dela no estoque:
    devolve a quantidade vendida ao produto e, se o produto tinha receita (ficha
    técnica), devolve também a matéria-prima que tinha sido baixada automaticamente.
    Só funciona para movimentações do tipo 'venda' — outros tipos (entrada, saída,
    ajuste) continuam sem exclusão pelo painel, pra não abrir brecha de descontrole
    de estoque em fluxos que não foram pedidos aqui."""
    mov = db_one(conn, "SELECT * FROM movimentacoes WHERE id = %s AND cliente_id = %s",
                 (movimentacao_id, cliente["id"]))
    if not mov:
        raise HTTPException(status_code=404, detail="Movimentação não encontrada")
    if mov["tipo"] != "venda":
        raise HTTPException(status_code=400, detail="Só é possível excluir movimentações do tipo venda por aqui")

    # 1) Devolve a matéria-prima baixada automaticamente pela receita (se houver)
    baixas_mp = db_all(conn, "SELECT * FROM movimentacoes_materia_prima WHERE movimentacao_id = %s",
                        (movimentacao_id,))
    for baixa in baixas_mp:
        db_exec(conn, "UPDATE materias_primas SET estoque_atual = estoque_atual + %s WHERE id = %s",
                (baixa["quantidade"], baixa["materia_prima_id"]))
    if baixas_mp:
        db_exec(conn, "DELETE FROM movimentacoes_materia_prima WHERE movimentacao_id = %s", (movimentacao_id,))

    # 2) Devolve a quantidade vendida ao estoque do produto
    db_exec(conn, "UPDATE produtos SET estoque_atual = estoque_atual + %s WHERE id = %s",
            (mov["quantidade"], mov["produto_id"]))

    # 3) Remove o registro da venda
    db_exec(conn, "DELETE FROM movimentacoes WHERE id = %s", (movimentacao_id,))
    return {"ok": True}

@app.post("/movimentacoes")
def criar_movimentacao_manual(body: MovimentacaoManualBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s", (body.produto_id, cliente["id"]))
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    _, alertas = aplicar_movimentacao(conn, cliente["id"], body.produto_id, None, body.tipo, body.quantidade,
                                       body.valor_unitario, origem="manual_admin",
                                       cliente_negocio_id=body.cliente_negocio_id)
    return {"ok": True, "alertas": alertas}

# ═════════════════════════════════════════
#  CLIENTES DO NEGÓCIO (painel do cliente) — TAREFA 2
# ═════════════════════════════════════════
@app.get("/clientes-negocio")
def listar_clientes_negocio(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    return db_all(conn, "SELECT * FROM clientes_negocio WHERE cliente_id = %s ORDER BY nome",
                  (cliente["id"],))

@app.post("/clientes-negocio")
def criar_cliente_negocio(body: ClienteNegocioBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    telefone = re.sub(r"\D", "", body.telefone) if body.telefone else None
    if telefone:
        existente = db_one(conn, "SELECT * FROM clientes_negocio WHERE cliente_id = %s AND telefone = %s",
                            (cliente["id"], telefone))
        if existente:
            raise HTTPException(status_code=400, detail="Já existe um cliente com esse telefone")
    return db_exec(conn, """
        INSERT INTO clientes_negocio (cliente_id, nome, telefone) VALUES (%s,%s,%s) RETURNING *
    """, (cliente["id"], body.nome, telefone))

@app.get("/clientes-negocio/{cliente_negocio_id}/historico")
def historico_cliente_negocio(cliente_negocio_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Retorna todos os orçamentos e vendas vinculados a esse cliente, ordenados por data."""
    cliente_negocio = db_one(conn, "SELECT * FROM clientes_negocio WHERE id = %s AND cliente_id = %s",
                              (cliente_negocio_id, cliente["id"]))
    if not cliente_negocio:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    orcamentos = db_all(conn, """
        SELECT id, 'orcamento' AS tipo_registro, nome_cliente, itens, subtotal, total, criado_em
        FROM orcamentos WHERE cliente_id = %s AND cliente_negocio_id = %s
    """, (cliente["id"], cliente_negocio_id))
    vendas = db_all(conn, """
        SELECT m.id, 'venda' AS tipo_registro, p.nome AS produto_nome, m.quantidade,
               m.valor_unitario, m.valor_total, m.criado_em
        FROM movimentacoes m JOIN produtos p ON p.id = m.produto_id
        WHERE m.cliente_id = %s AND m.cliente_negocio_id = %s AND m.tipo = 'venda'
    """, (cliente["id"], cliente_negocio_id))

    historico = sorted(orcamentos + vendas, key=lambda r: r["criado_em"], reverse=True)
    return {"cliente": cliente_negocio, "historico": historico}

@app.put("/clientes-negocio/{cliente_negocio_id}")
def atualizar_cliente_negocio(cliente_negocio_id: int, body: ClienteNegocioBody,
                               cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """TAREFA 4 — edita nome/telefone de um cliente do negócio."""
    cliente_negocio = db_one(conn, "SELECT * FROM clientes_negocio WHERE id = %s AND cliente_id = %s",
                              (cliente_negocio_id, cliente["id"]))
    if not cliente_negocio:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    telefone = re.sub(r"\D", "", body.telefone) if body.telefone else None
    if telefone:
        existente = db_one(conn, """
            SELECT * FROM clientes_negocio WHERE cliente_id = %s AND telefone = %s AND id != %s
        """, (cliente["id"], telefone, cliente_negocio_id))
        if existente:
            raise HTTPException(status_code=400, detail="Já existe um cliente com esse telefone")

    db_exec(conn, "UPDATE clientes_negocio SET nome = %s, telefone = %s WHERE id = %s",
            (body.nome, telefone, cliente_negocio_id))
    return {"ok": True}

@app.delete("/clientes-negocio/{cliente_negocio_id}")
def deletar_cliente_negocio(cliente_negocio_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """TAREFA 4 — exclui o cliente do negócio SEM apagar vendas/orçamentos antigos:
    apenas desvincula (cliente_negocio_id = NULL) pra manter o histórico intacto."""
    cliente_negocio = db_one(conn, "SELECT * FROM clientes_negocio WHERE id = %s AND cliente_id = %s",
                              (cliente_negocio_id, cliente["id"]))
    if not cliente_negocio:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")

    db_exec(conn, "UPDATE movimentacoes SET cliente_negocio_id = NULL WHERE cliente_negocio_id = %s",
            (cliente_negocio_id,))
    db_exec(conn, "UPDATE orcamentos SET cliente_negocio_id = NULL WHERE cliente_negocio_id = %s",
            (cliente_negocio_id,))
    db_exec(conn, "DELETE FROM clientes_negocio WHERE id = %s", (cliente_negocio_id,))
    return {"ok": True}

# ═════════════════════════════════════════
#  CUSTOS FIXOS + CALCULADORA DE PREÇO (painel do cliente) — TAREFA 5
# ═════════════════════════════════════════
@app.get("/custos-fixos")
def listar_custos_fixos(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    return db_all(conn, "SELECT * FROM custos_fixos WHERE cliente_id = %s ORDER BY nome_da_conta",
                  (cliente["id"],))

@app.post("/custos-fixos")
def criar_custo_fixo(body: CustoFixoBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    return db_exec(conn, """
        INSERT INTO custos_fixos (cliente_id, nome_da_conta, valor, atualizado_em)
        VALUES (%s,%s,%s,NOW()) RETURNING *
    """, (cliente["id"], body.nome_da_conta, body.valor))

@app.get("/custos-fixos/export.csv")
def export_custos_fixos_csv(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    linhas = db_all(conn, """
        SELECT nome_da_conta, valor, atualizado_em FROM custos_fixos
        WHERE cliente_id = %s ORDER BY nome_da_conta
    """, (cliente["id"],))
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Conta", "Valor", "Atualizado em"])
    for l in linhas:
        writer.writerow([l["nome_da_conta"], l["valor"], l["atualizado_em"]])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=custos-fixos.csv"}
    )

@app.get("/custos-fixos/export.excel")
def export_custos_fixos_excel(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    linhas = db_all(conn, """
        SELECT nome_da_conta, valor, atualizado_em FROM custos_fixos
        WHERE cliente_id = %s ORDER BY nome_da_conta
    """, (cliente["id"],))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Custos Fixos"
    ws.append(["Conta", "Valor", "Atualizado em"])
    for l in linhas:
        ws.append([l["nome_da_conta"], l["valor"], l["atualizado_em"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=custos-fixos.xlsx"}
    )

@app.post("/custos-fixos/calculadora")
def calcular_preco_com_custos_fixos(body: CalculadoraCustoBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Calculadora de preço do dashboard: usa o total de custos fixos JÁ SALVO no
    banco (soma da tabela custos_fixos) em vez de pedir pra redigitar tudo de novo,
    diferente de como funciona hoje no bot do WhatsApp. Reaproveita a mesma fórmula
    de `calcular_resultado_calculadora`."""
    custo_variavel = float(body.custo_variavel or 0)
    if body.produto_id:
        produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s",
                          (body.produto_id, cliente["id"]))
        if not produto:
            raise HTTPException(status_code=404, detail="Produto não encontrado")
        if not custo_variavel:
            custo_variavel = float(produto["custo_unitario"] or 0)

    total_fixo = db_one(conn, "SELECT COALESCE(SUM(valor),0) AS total FROM custos_fixos WHERE cliente_id = %s",
                         (cliente["id"],))
    custo_fixo_mensal = float(total_fixo["total"])

    dados = {
        "custo_variavel": custo_variavel,
        "calc_custo_fixo_mensal": custo_fixo_mensal,
        "calc_volume_esperado": body.volume_esperado,
        "calc_margem": body.margem if body.margem is not None else 30.0,
    }
    resultado = calcular_resultado_calculadora(dados)
    resultado["custo_fixo_mensal_total"] = round(custo_fixo_mensal, 2)
    return resultado

@app.put("/custos-fixos/{custo_id}")
def atualizar_custo_fixo(custo_id: int, body: CustoFixoBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    custo = db_one(conn, "SELECT * FROM custos_fixos WHERE id = %s AND cliente_id = %s", (custo_id, cliente["id"]))
    if not custo:
        raise HTTPException(status_code=404, detail="Custo fixo não encontrado")
    db_exec(conn, "UPDATE custos_fixos SET nome_da_conta = %s, valor = %s, atualizado_em = NOW() WHERE id = %s",
            (body.nome_da_conta, body.valor, custo_id))
    return {"ok": True}

@app.delete("/custos-fixos/{custo_id}")
def deletar_custo_fixo(custo_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    custo = db_one(conn, "SELECT * FROM custos_fixos WHERE id = %s AND cliente_id = %s", (custo_id, cliente["id"]))
    if not custo:
        raise HTTPException(status_code=404, detail="Custo fixo não encontrado")
    db_exec(conn, "DELETE FROM custos_fixos WHERE id = %s", (custo_id,))
    return {"ok": True}

@app.get("/dashboard/resumo")
def dashboard_resumo(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    produtos = db_all(conn, "SELECT * FROM produtos WHERE cliente_id = %s AND ativo = TRUE", (cliente["id"],))
    valor_em_estoque = sum(float(p["estoque_atual"]) * float(p["custo_unitario"]) for p in produtos)

    hoje = datetime.utcnow().date()
    inicio_mes = hoje.replace(day=1)
    vendas_mes = db_one(conn, """
        SELECT COALESCE(SUM(valor_total),0) total, COALESCE(SUM(quantidade),0) qtd
        FROM movimentacoes WHERE cliente_id = %s AND tipo = 'venda' AND criado_em::date >= %s
    """, (cliente["id"], inicio_mes))
    custo_vendido_mes = db_all(conn, """
        SELECT m.produto_id, m.quantidade, p.custo_unitario
        FROM movimentacoes m JOIN produtos p ON p.id = m.produto_id
        WHERE m.cliente_id = %s AND m.tipo = 'venda' AND m.criado_em::date >= %s
    """, (cliente["id"], inicio_mes))
    custo_total_mes = sum(float(r["quantidade"]) * float(r["custo_unitario"]) for r in custo_vendido_mes)
    margem_mes = float(vendas_mes["total"]) - custo_total_mes

    return {
        "total_produtos": len(produtos),
        "valor_em_estoque": round(valor_em_estoque, 2),
        "produtos_estoque_baixo": [p for p in produtos if float(p["estoque_atual"]) <= 0],
        "vendas_mes_total": round(float(vendas_mes["total"]), 2),
        "vendas_mes_qtd": float(vendas_mes["qtd"]),
        "margem_mes": round(margem_mes, 2),
    }

@app.get("/dashboard/export.csv")
def export_csv(de: Optional[str] = None, ate: Optional[str] = None,
                cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    sql = """
        SELECT m.criado_em, p.nome AS produto, m.tipo, m.quantidade, m.valor_unitario, m.valor_total, m.origem
        FROM movimentacoes m JOIN produtos p ON p.id = m.produto_id
        WHERE m.cliente_id = %s
    """
    params = [cliente["id"]]
    if de:
        sql += " AND m.criado_em::date >= %s"; params.append(de)
    if ate:
        sql += " AND m.criado_em::date <= %s"; params.append(ate)
    sql += " ORDER BY m.criado_em DESC"
    linhas = db_all(conn, sql, tuple(params))

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Data", "Produto", "Tipo", "Quantidade", "Valor Unitário", "Valor Total", "Origem"])
    for l in linhas:
        writer.writerow([l["criado_em"], l["produto"], l["tipo"], l["quantidade"],
                          l["valor_unitario"], l["valor_total"], l["origem"]])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=movimentacoes.csv"}
    )

@app.get("/dashboard/export.excel")
def export_excel(de: Optional[str] = None, ate: Optional[str] = None,
                  cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    sql = """
        SELECT m.criado_em, p.nome AS produto, m.tipo, m.quantidade, m.valor_unitario, m.valor_total, m.origem
        FROM movimentacoes m JOIN produtos p ON p.id = m.produto_id
        WHERE m.cliente_id = %s
    """
    params = [cliente["id"]]
    if de:
        sql += " AND m.criado_em::date >= %s"; params.append(de)
    if ate:
        sql += " AND m.criado_em::date <= %s"; params.append(ate)
    sql += " ORDER BY m.criado_em DESC"
    linhas = db_all(conn, sql, tuple(params))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentações"
    ws.append(["Data", "Produto", "Tipo", "Quantidade", "Valor Unitário", "Valor Total", "Origem"])
    for l in linhas:
        ws.append([l["criado_em"], l["produto"], l["tipo"], l["quantidade"],
                    l["valor_unitario"], l["valor_total"], l["origem"]])
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=movimentacoes.xlsx"}
    )

@app.get("/dashboard/template-excel")
def template_excel(cliente=Depends(get_current_cliente)):
    """Retorna um template Excel vazio pra cliente preencher e importar.
    Tem 3 abas: Produtos, Matérias-Primas e Receita — a aba Receita é o que
    liga um produto às matérias-primas dele (o vínculo em si), usando o
    NOME de cada um pra achar o registro certo (não precisa saber o ID)."""
    wb = openpyxl.Workbook()

    def formatar_header(ws):
        for cell in ws[1]:
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    ws_produtos = wb.active
    ws_produtos.title = "Produtos"
    ws_produtos.append(["Nome do Produto", "SKU", "Custo Unitário", "Preço de Venda", "Estoque Atual", "Unidade"])
    ws_produtos.append(["Exemplo: Hambúrguer", "HAM-001", "5.50", "15.00", "10", "un"])
    formatar_header(ws_produtos)

    ws_materias = wb.create_sheet("Matérias-Primas")
    ws_materias.append(["Nome da Matéria-Prima", "SKU", "Custo Unitário", "Estoque Atual", "Unidade"])
    ws_materias.append(["Exemplo: Pão", "PAO-001", "0.80", "50", "un"])
    formatar_header(ws_materias)

    ws_receita = wb.create_sheet("Receita")
    ws_receita.append(["Nome do Produto", "Nome da Matéria-Prima", "Quantidade por Unidade"])
    ws_receita.append(["Exemplo: Hambúrguer", "Exemplo: Pão", "1"])
    ws_receita.append(["Exemplo: Hambúrguer", "Exemplo: Carne", "0.15"])
    formatar_header(ws_receita)
    ws_receita.column_dimensions["A"].width = 25
    ws_receita.column_dimensions["B"].width = 25
    ws_receita.column_dimensions["C"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template-produtos.xlsx"}
    )

@app.post("/produtos/importar-excel")
async def importar_excel(file: UploadFile = File(...), cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Importa produtos, matérias-primas e a receita (vínculo produto ↔ matéria-prima)
    de um arquivo Excel. As abas Matérias-Primas e Receita são opcionais — um arquivo
    só com a aba Produtos (template antigo) continua funcionando normalmente."""
    try:
        conteudo = await file.read()
        planilhas = pd.read_excel(io.BytesIO(conteudo), sheet_name=None)

        # Aceita tanto o arquivo novo (3 abas) quanto o antigo (1 aba sem nome específico)
        df_produtos = planilhas.get("Produtos")
        if df_produtos is None:
            # Template antigo: uma aba só, sem nome "Produtos" — usa a primeira
            primeira_aba = next(iter(planilhas.values()))
            df_produtos = primeira_aba
        df_materias = planilhas.get("Matérias-Primas")
        df_receita = planilhas.get("Receita")

        colunas_esperadas = ["Nome do Produto", "SKU", "Custo Unitário", "Preço de Venda", "Estoque Atual", "Unidade"]
        for col in colunas_esperadas:
            if col not in df_produtos.columns:
                raise HTTPException(status_code=400, detail=f"Coluna ausente na aba Produtos: {col}")

        produtos_criados = 0
        for idx, row in df_produtos.iterrows():
            nome = str(row["Nome do Produto"]).strip()
            if not nome or nome.startswith("Exemplo"):
                continue

            try:
                db_exec(conn, """
                    INSERT INTO produtos (cliente_id, nome, sku, custo_unitario, preco_venda, estoque_atual, unidade, ativo)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
                """, (
                    cliente["id"],
                    nome,
                    str(row["SKU"]).strip() if pd.notna(row["SKU"]) else "",
                    float(row["Custo Unitário"]) if pd.notna(row["Custo Unitário"]) else 0,
                    float(row["Preço de Venda"]) if pd.notna(row["Preço de Venda"]) else 0,
                    float(row["Estoque Atual"]) if pd.notna(row["Estoque Atual"]) else 0,
                    str(row["Unidade"]).strip() if pd.notna(row["Unidade"]) else "un"
                ))
                produtos_criados += 1
            except Exception as e:
                print(f"Erro ao importar linha {idx} (Produtos): {e}")

        materias_criadas = 0
        if df_materias is not None:
            colunas_materias = ["Nome da Matéria-Prima", "SKU", "Custo Unitário", "Estoque Atual", "Unidade"]
            for col in colunas_materias:
                if col not in df_materias.columns:
                    raise HTTPException(status_code=400, detail=f"Coluna ausente na aba Matérias-Primas: {col}")

            for idx, row in df_materias.iterrows():
                nome = str(row["Nome da Matéria-Prima"]).strip()
                if not nome or nome.startswith("Exemplo"):
                    continue
                try:
                    # Evita duplicar se a matéria-prima já existe (importante pra reimportar o mesmo
                    # arquivo depois de editar só a aba Receita, sem duplicar o cadastro)
                    ja_existe = buscar_materia_prima_por_nome(conn, cliente["id"], nome)
                    if ja_existe:
                        continue
                    db_exec(conn, """
                        INSERT INTO materias_primas (cliente_id, nome, sku, custo_unitario, estoque_atual, unidade, ativo)
                        VALUES (%s, %s, %s, %s, %s, %s, TRUE)
                    """, (
                        cliente["id"],
                        nome,
                        str(row["SKU"]).strip() if pd.notna(row["SKU"]) else "",
                        float(row["Custo Unitário"]) if pd.notna(row["Custo Unitário"]) else 0,
                        float(row["Estoque Atual"]) if pd.notna(row["Estoque Atual"]) else 0,
                        str(row["Unidade"]).strip() if pd.notna(row["Unidade"]) else "un"
                    ))
                    materias_criadas += 1
                except Exception as e:
                    print(f"Erro ao importar linha {idx} (Matérias-Primas): {e}")

        receitas_vinculadas = 0
        receitas_com_erro = []
        if df_receita is not None:
            colunas_receita = ["Nome do Produto", "Nome da Matéria-Prima", "Quantidade por Unidade"]
            for col in colunas_receita:
                if col not in df_receita.columns:
                    raise HTTPException(status_code=400, detail=f"Coluna ausente na aba Receita: {col}")

            # Agrupa por produto pra gravar a receita inteira de uma vez (mesma lógica do
            # endpoint manual PUT /produtos/{id}/receita — substitui a receita anterior do produto)
            itens_por_produto = {}
            for idx, row in df_receita.iterrows():
                nome_produto = str(row["Nome do Produto"]).strip()
                nome_materia = str(row["Nome da Matéria-Prima"]).strip()
                if not nome_produto or nome_produto.startswith("Exemplo"):
                    continue
                if not pd.notna(row["Quantidade por Unidade"]):
                    continue
                try:
                    quantidade = float(row["Quantidade por Unidade"])
                except (ValueError, TypeError):
                    receitas_com_erro.append(f"linha {idx + 2}: quantidade inválida")
                    continue

                produto = buscar_produto_por_nome(conn, cliente["id"], nome_produto)
                materia = buscar_materia_prima_por_nome(conn, cliente["id"], nome_materia)
                if not produto:
                    receitas_com_erro.append(f"linha {idx + 2}: produto '{nome_produto}' não encontrado")
                    continue
                if not materia:
                    receitas_com_erro.append(f"linha {idx + 2}: matéria-prima '{nome_materia}' não encontrada")
                    continue

                itens_por_produto.setdefault(produto["id"], []).append(
                    {"materia_prima_id": materia["id"], "quantidade_necessaria": quantidade}
                )

            for produto_id, itens in itens_por_produto.items():
                db_exec(conn, "DELETE FROM receita_itens WHERE produto_id = %s", (produto_id,))
                for item in itens:
                    db_exec(conn, """
                        INSERT INTO receita_itens (produto_id, materia_prima_id, quantidade_necessaria)
                        VALUES (%s,%s,%s)
                    """, (produto_id, item["materia_prima_id"], item["quantidade_necessaria"]))
                    receitas_vinculadas += 1

        resultado = {
            "ok": True,
            "produtos_criados": produtos_criados,
            "materias_primas_criadas": materias_criadas,
            "receitas_vinculadas": receitas_vinculadas,
        }
        if receitas_com_erro:
            resultado["avisos"] = receitas_com_erro
        return resultado
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo: {str(e)}")

# ═════════════════════════════════════════
#  AGENDAMENTOS
# ═════════════════════════════════════════
@app.post("/agendamentos")
def criar_agendamento(body: AgendamentoBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Criar novo agendamento"""
    produto = db_one(conn, "SELECT * FROM produtos WHERE id = %s AND cliente_id = %s", (body.produto_id, cliente["id"]))
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    
    # Combina data e hora em um datetime
    try:
        data_hora_str = f"{body.data_agendamento} {body.hora}:00"
        data_agendamento = datetime.fromisoformat(data_hora_str)
    except ValueError:
        raise HTTPException(status_code=400, detail="Data/hora inválida (use YYYY-MM-DD e HH:MM)")
    
    db_exec(conn, """
        INSERT INTO agendamentos (cliente_id, produto_id, data_agendamento, quantidade, notas, criado_em)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (cliente["id"], body.produto_id, data_agendamento, body.quantidade, body.notas or "", datetime.utcnow()))
    
    return {"ok": True}

@app.get("/agendamentos")
def listar_agendamentos(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Listar agendamentos do cliente"""
    return db_all(conn, """
        SELECT a.id, a.produto_id, a.data_agendamento, a.quantidade, a.notas, p.nome AS produto_nome
        FROM agendamentos a
        JOIN produtos p ON p.id = a.produto_id
        WHERE a.cliente_id = %s
        ORDER BY a.data_agendamento DESC
    """, (cliente["id"],))

@app.delete("/agendamentos/{agendamento_id}")
def deletar_agendamento(agendamento_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Deletar agendamento"""
    agendamento = db_one(conn, "SELECT * FROM agendamentos WHERE id = %s AND cliente_id = %s", (agendamento_id, cliente["id"]))
    if not agendamento:
        raise HTTPException(status_code=404, detail="Agendamento não encontrado")
    db_exec(conn, "DELETE FROM agendamentos WHERE id = %s", (agendamento_id,))
    return {"ok": True}

# ═════════════════════════════════════════
#  AGENDA / COMPROMISSOS
#  (⚠️ Não confundir com /agendamentos acima, que é reposição de estoque —
#  esta aqui é a agenda genérica de compromissos, tabela agenda_compromissos)
# ═════════════════════════════════════════
@app.post("/agenda")
def criar_compromisso_agenda(body: AgendaCompromissoBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Cria um novo compromisso. Pode ou não estar vinculado a um cliente_negocio.
    Conflito de horário só gera aviso (aviso_conflito na resposta) — não bloqueia."""
    try:
        data_obj = datetime.fromisoformat(body.data).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Data inválida (use YYYY-MM-DD)")
    if not re.fullmatch(r"\d{1,2}:\d{2}", body.hora_inicio):
        raise HTTPException(status_code=400, detail="Hora inválida (use HH:MM)")
    hora_fim = calcular_hora_fim(body.hora_inicio, body.hora_fim, body.duracao_minutos)

    cliente_negocio_id = None
    if body.cliente_negocio_id:
        cn = db_one(conn, "SELECT id FROM clientes_negocio WHERE id = %s AND cliente_id = %s",
                    (body.cliente_negocio_id, cliente["id"]))
        if not cn:
            raise HTTPException(status_code=404, detail="Cliente não encontrado")
        cliente_negocio_id = cn["id"]

    conflitos = verificar_conflito_agenda(conn, cliente["id"], data_obj, body.hora_inicio, hora_fim)

    registro = db_exec(conn, """
        INSERT INTO agenda_compromissos
            (cliente_id, cliente_negocio_id, titulo, data, hora_inicio, hora_fim, notas, lembrete_minutos_antes)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *
    """, (cliente["id"], cliente_negocio_id, body.titulo, data_obj, body.hora_inicio, hora_fim,
          body.notas, body.lembrete_minutos_antes))

    resultado = dict(registro)
    if conflitos:
        resultado["aviso_conflito"] = [f"{c['titulo']} ({str(c['hora_inicio'])[:5]})" for c in conflitos]
    return resultado

@app.get("/agenda")
def listar_agenda(inicio: str, fim: str, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Lista compromissos (não cancelados ou não) entre duas datas — usado pelo
    calendário do painel (dia/semana/mês manda o intervalo visível)."""
    try:
        data_inicio = datetime.fromisoformat(inicio).date()
        data_fim = datetime.fromisoformat(fim).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Datas inválidas (use YYYY-MM-DD)")
    return db_all(conn, """
        SELECT a.*, cn.nome AS cliente_negocio_nome
        FROM agenda_compromissos a
        LEFT JOIN clientes_negocio cn ON cn.id = a.cliente_negocio_id
        WHERE a.cliente_id = %s AND a.data BETWEEN %s AND %s
        ORDER BY a.data, a.hora_inicio
    """, (cliente["id"], data_inicio, data_fim))

@app.put("/agenda/{compromisso_id}")
def editar_compromisso_agenda(compromisso_id: int, body: AgendaCompromissoBody,
                               cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    existente = db_one(conn, "SELECT * FROM agenda_compromissos WHERE id = %s AND cliente_id = %s",
                        (compromisso_id, cliente["id"]))
    if not existente:
        raise HTTPException(status_code=404, detail="Compromisso não encontrado")
    try:
        data_obj = datetime.fromisoformat(body.data).date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Data inválida (use YYYY-MM-DD)")
    if not re.fullmatch(r"\d{1,2}:\d{2}", body.hora_inicio):
        raise HTTPException(status_code=400, detail="Hora inválida (use HH:MM)")
    hora_fim = calcular_hora_fim(body.hora_inicio, body.hora_fim, body.duracao_minutos)

    cliente_negocio_id = None
    if body.cliente_negocio_id:
        cn = db_one(conn, "SELECT id FROM clientes_negocio WHERE id = %s AND cliente_id = %s",
                    (body.cliente_negocio_id, cliente["id"]))
        if not cn:
            raise HTTPException(status_code=404, detail="Cliente não encontrado")
        cliente_negocio_id = cn["id"]

    conflitos = verificar_conflito_agenda(conn, cliente["id"], data_obj, body.hora_inicio, hora_fim,
                                           excluir_id=compromisso_id)

    # Se mudou data/hora/lembrete, reseta lembrete_enviado pra poder disparar de novo.
    lembrete_enviado = existente["lembrete_enviado"]
    if (str(existente["hora_inicio"])[:5] != body.hora_inicio
            or existente["lembrete_minutos_antes"] != body.lembrete_minutos_antes
            or str(existente["data"]) != str(data_obj)):
        lembrete_enviado = False

    novo_status = body.status if body.status in ("agendado", "concluido", "cancelado") else existente["status"]

    registro = db_exec(conn, """
        UPDATE agenda_compromissos
        SET titulo = %s, data = %s, hora_inicio = %s, hora_fim = %s, cliente_negocio_id = %s,
            notas = %s, lembrete_minutos_antes = %s, lembrete_enviado = %s, status = %s
        WHERE id = %s RETURNING *
    """, (body.titulo, data_obj, body.hora_inicio, hora_fim, cliente_negocio_id, body.notas,
          body.lembrete_minutos_antes, lembrete_enviado, novo_status, compromisso_id))

    resultado = dict(registro)
    if conflitos:
        resultado["aviso_conflito"] = [f"{c['titulo']} ({str(c['hora_inicio'])[:5]})" for c in conflitos]
    return resultado

@app.patch("/agenda/{compromisso_id}/status")
def mudar_status_compromisso_agenda(compromisso_id: int, body: AgendaStatusBody,
                                     cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    existente = db_one(conn, "SELECT * FROM agenda_compromissos WHERE id = %s AND cliente_id = %s",
                        (compromisso_id, cliente["id"]))
    if not existente:
        raise HTTPException(status_code=404, detail="Compromisso não encontrado")
    if body.status not in ("agendado", "concluido", "cancelado"):
        raise HTTPException(status_code=400, detail="Status inválido")
    db_exec(conn, "UPDATE agenda_compromissos SET status = %s WHERE id = %s", (body.status, compromisso_id))
    return {"ok": True}

@app.delete("/agenda/{compromisso_id}")
def deletar_compromisso_agenda(compromisso_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    existente = db_one(conn, "SELECT * FROM agenda_compromissos WHERE id = %s AND cliente_id = %s",
                        (compromisso_id, cliente["id"]))
    if not existente:
        raise HTTPException(status_code=404, detail="Compromisso não encontrado")
    db_exec(conn, "DELETE FROM agenda_compromissos WHERE id = %s", (compromisso_id,))
    return {"ok": True}

# ═════════════════════════════════════════
#  ORÇAMENTOS
# ═════════════════════════════════════════
@app.post("/orcamentos")
def criar_orcamento(body: OrcamentoBody, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Monta um orçamento a partir de produtos + quantidade, calcula desconto/total
    e já gera o texto formatado pronto pra copiar/enviar."""
    if not body.itens:
        raise HTTPException(status_code=400, detail="Adicione ao menos um produto")

    itens_input = [{"produto_id": item.produto_id, "quantidade": item.quantidade} for item in body.itens]
    try:
        itens_detalhados, subtotal = calcular_itens_orcamento(conn, cliente["id"], itens_input)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    # aceita "valor"/"percentual" (desconto, comportamento antigo do painel) ou já
    # "desconto_valor"/"desconto_percentual"/"aumento_valor"/"aumento_percentual"
    tipo_bruto = body.desconto_tipo
    if tipo_bruto in ("valor", "percentual"):
        ajuste_tipo = f"desconto_{tipo_bruto}"
    elif tipo_bruto in ("desconto_valor", "desconto_percentual", "aumento_valor", "aumento_percentual"):
        ajuste_tipo = tipo_bruto
    else:
        ajuste_tipo = None
    ajuste_valor_informado = float(body.desconto_valor or 0)
    ajuste_calculado = calcular_ajuste_preco(subtotal, ajuste_tipo, ajuste_valor_informado)
    total = subtotal + ajuste_calculado

    # ── Monta o texto formatado (pronto pra copiar/enviar no WhatsApp) ──
    texto_formatado = montar_texto_orcamento(
        cliente, itens_detalhados, subtotal, ajuste_tipo, ajuste_calculado,
        ajuste_valor_informado, total, observacoes=body.observacoes, nome_cliente=body.nome_cliente
    )

    registro = salvar_orcamento_db(conn, cliente["id"], body.nome_cliente, itens_detalhados, subtotal,
                                    ajuste_tipo, ajuste_calculado, total, texto_formatado, body.observacoes,
                                    cliente_negocio_id=body.cliente_negocio_id)
    return registro

@app.get("/orcamentos")
def listar_orcamentos(cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Lista o histórico de orçamentos gerados"""
    return db_all(conn, """
        SELECT id, nome_cliente, itens, subtotal, desconto_tipo, desconto_valor, total, texto_formatado, observacoes, criado_em
        FROM orcamentos
        WHERE cliente_id = %s
        ORDER BY criado_em DESC
    """, (cliente["id"],))

@app.delete("/orcamentos/{orcamento_id}")
def deletar_orcamento(orcamento_id: int, cliente=Depends(get_current_cliente), conn=Depends(get_db)):
    """Remove um orçamento do histórico"""
    orcamento = db_one(conn, "SELECT * FROM orcamentos WHERE id = %s AND cliente_id = %s",
                        (orcamento_id, cliente["id"]))
    if not orcamento:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")
    db_exec(conn, "DELETE FROM orcamentos WHERE id = %s", (orcamento_id,))
    return {"ok": True}

# ═════════════════════════════════════════
#  WHATSAPP — webhook e status
# ═════════════════════════════════════════
@app.post("/webhook/mensagem")
async def webhook_mensagem(payload: dict, conn=Depends(get_db)):
    remote_jid = payload.get("remoteJid", "")
    numero_resolvido = payload.get("numeroResolvido")  # ex: "5511999998888@s.whatsapp.net", quando o remote_jid é @lid
    texto = payload.get("text")
    from_me = payload.get("fromMe", False)
    if from_me or not texto or remote_jid.endswith("@g.us"):
        return {"ok": True}

    # Autorização é checada pelo número de telefone real (resolvido do LID
    # quando disponível). A resposta, porém, SEMPRE vai pro remote_jid
    # original — trocar de endereço no meio da conversa quebra a sessão de
    # criptografia do WhatsApp e a mensagem fica "Aguardando mensagem".
    numero = normalizar_numero(numero_resolvido or remote_jid)
    numero_autorizado = buscar_numero_autorizado(conn, numero)

    if numero_autorizado:
        # ── MODO FUNCIONÁRIO/DONA — menu completo (estoque, vendas, agenda etc) ──
        cliente = db_one(conn, "SELECT * FROM clientes WHERE id = %s AND ativo = TRUE", (numero_autorizado["cliente_id"],))
        if not cliente:
            await enviar_whatsapp(remote_jid, "❌ Conta inativa. Fale com o administrador.")
            return {"ok": True}

        resposta = await processar_texto(conn, cliente, numero_autorizado, texto)
        await enviar_whatsapp(remote_jid, resposta)
        return {"ok": True}

    # ── Não é funcionário: será que é cliente final de alguma empresa? ──
    # Número é reconhecido pela ficha em clientes_negocio (telefone), e só
    # entra no menu simplificado se aquela empresa ligou o toggle no
    # dashboard. Isso evita responder em nome da empresa errada quando o
    # mesmo número compartilhado atende negócios diferentes.
    cliente_negocio = db_one(conn, """
        SELECT cn.*, c.id AS cliente_id, c.nome_negocio, c.ativo AS cliente_ativo,
               c.atendimento_cliente_final_ativado
        FROM clientes_negocio cn
        JOIN clientes c ON c.id = cn.cliente_id
        WHERE cn.telefone = %s AND c.atendimento_cliente_final_ativado = TRUE AND c.ativo = TRUE
        ORDER BY cn.criado_em DESC
        LIMIT 1
    """, (numero,))

    if cliente_negocio:
        cliente = db_one(conn, "SELECT * FROM clientes WHERE id = %s AND ativo = TRUE", (cliente_negocio["cliente_id"],))
        if cliente:
            # ── MODO CLIENTE FINAL — menu simplificado (ver produtos, orçamento, atendente) ──
            resposta = await processar_texto_cliente_final(conn, cliente, numero, texto)
            await enviar_whatsapp(remote_jid, resposta)
            return {"ok": True}

    await enviar_whatsapp(remote_jid, "❌ Número não autorizado. Fale com o administrador do sistema.")
    return {"ok": True}

@app.get("/whatsapp/status")
async def whatsapp_status():
    async with httpx.AsyncClient(timeout=10) as client:
        try:
            r = await client.get(f"{BAILEYS_URL}/status")
            return r.json()
        except Exception:
            return {"connected": False}

@app.get("/whatsapp/qrcode")
async def whatsapp_qrcode():
    async with httpx.AsyncClient(timeout=10) as client:
        try:
            r = await client.get(f"{BAILEYS_URL}/qrcode-raw")
            return r.json()
        except Exception:
            return {"qr": None}

@app.get("/")
def root():
    return FileResponse("frontend/login.html")

@app.get("/login.html")
def pagina_login():
    return FileResponse("frontend/login.html")

@app.get("/admin-login.html")
def pagina_admin_login():
    return FileResponse("frontend/admin-login.html")

@app.get("/dashboard.html")
def pagina_dashboard():
    return FileResponse("frontend/dashboard.html")

@app.get("/admin.html")
def pagina_admin():
    return FileResponse("frontend/admin.html")
